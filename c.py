"""
Full-featured Streamlit CRM — single-file example

Features:
- SQLAlchemy ORM models: Users, Leads, Activities (audit log), Comments
- Secure password hashing (passlib) and simple user management (admin can add users)
- Role-based permissions: salesman, head_of_sales, cto, ceo, admin
- Lead lifecycle: statuses (new, contacted, qualified, lost, won), assign to agent
- CRUD for leads: create, read, update, delete (with activity log)
- Upload XLSX/CSV and map columns; bulk-create leads
- Lead detail view with comments (notes) and history
- Filtering, searching, sorting, and pagination
- Dashboards for CTO (time-series, agent breakdown, status breakdown)
- Reporting and export (CSV / Excel) for CEO and admin
- Dockerfile and requirements notes in header

Run (dev):
1. pip install -r requirements.txt
   requirements.txt content below in comment
2. streamlit run streamlit_crm_full.py

Security note: This example uses an in-app user management for demo only.
For production use central auth (OAuth2 / SSO) and secure DB credentials.
"""

import streamlit as st
import pandas as pd
import sqlalchemy as sa
from sqlalchemy.orm import declarative_base, sessionmaker, relationship
from sqlalchemy import Column, Integer, String, DateTime, Text, ForeignKey, LargeBinary, func
from datetime import datetime, date
import io
import plotly.express as px
from passlib.context import CryptContext
import os
import math
import zipfile
import random

# ----------------- Requirements (put in requirements.txt) -----------------
# streamlit
# pandas
# sqlalchemy
# openpyxl
# plotly
# passlib
# python-dateutil
# --------------------------------------------------------------------------

BASE_DIR = os.path.dirname(__file__) if '__file__' in globals() else '.'
DB_FILE = os.path.join(BASE_DIR, 'crm_full.db')
DATABASE_URL = f"sqlite:///{DB_FILE}"

# ----------------- DB setup -----------------
engine = sa.create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# ----------------- Models -----------------
class User(Base):
    __tablename__ = 'users'
    id = Column(Integer, primary_key=True)
    username = Column(String, unique=True, nullable=False)
    name = Column(String)
    password_hash = Column(String, nullable=False)
    role = Column(String, nullable=False)  # salesman, head_of_sales, cto, ceo, admin
    created_at = Column(DateTime, default=datetime.utcnow)

    # relationship to leads via uploaded_by_id
    uploads = relationship('Lead', back_populates='uploader', cascade="all, delete-orphan")
    # relationship to deals via uploaded_by_id
    deals = relationship('Deal', back_populates='uploader', cascade="all, delete-orphan")

    def verify_password(self, password):
        return pwd_context.verify(password, self.password_hash)

    @classmethod
    def hash_password(cls, password):
        return pwd_context.hash(password)

class Lead(Base):
    __tablename__ = 'leads'
    id = Column(Integer, primary_key=True)
    number = Column(String, index=True)
    name = Column(String)
    sales_agent = Column(String, index=True)
    contact = Column(String)
    case_desc = Column(Text)
    feedback = Column(Text)
    status = Column(String, default='new')  # new, contacted, qualified, lost, won
    assigned_to = Column(String, nullable=True)

    # Keep a readable username string and also a foreign-key to users.id
    uploaded_by = Column(String)                     # username (readable)
    uploaded_by_id = Column(Integer, ForeignKey('users.id'))  # FK
    uploaded_at = Column(DateTime, default=datetime.utcnow)

    # relationships
    uploader = relationship('User', back_populates='uploads')
    activities = relationship('Activity', back_populates='lead', cascade="all, delete-orphan")
    comments = relationship('Comment', back_populates='lead', cascade="all, delete-orphan")

class Activity(Base):
    __tablename__ = 'activities'
    id = Column(Integer, primary_key=True)
    lead_id = Column(Integer, ForeignKey('leads.id'))
    actor = Column(String)
    action = Column(String)
    timestamp = Column(DateTime, default=datetime.utcnow)
    detail = Column(Text)
    lead = relationship('Lead', back_populates='activities')

class Comment(Base):
    __tablename__ = 'comments'
    id = Column(Integer, primary_key=True)
    lead_id = Column(Integer, ForeignKey('leads.id'))
    author = Column(String)
    text = Column(Text)
    created_at = Column(DateTime, default=datetime.utcnow)
    lead = relationship('Lead', back_populates='comments')

class Deal(Base):
    __tablename__ = 'deals'
    id = Column(Integer, primary_key=True)
    customer_name = Column(String, nullable=False)
    phone = Column(String, nullable=False)
    payment_screenshot = Column(LargeBinary, nullable=False)
    uploaded_by = Column(String)                     # username (readable)
    uploaded_by_id = Column(Integer, ForeignKey('users.id'))  # FK
    created_at = Column(DateTime, default=datetime.utcnow)
    uploader = relationship('User', back_populates='deals')

class LoginEvent(Base):
    __tablename__ = 'login_events'
    id = Column(Integer, primary_key=True)
    username = Column(String, nullable=False)
    role = Column(String, nullable=False)
    logged_in_at = Column(DateTime, default=datetime.utcnow, index=True)

class Setting(Base):
    __tablename__ = 'settings'
    key = Column(String, primary_key=True)
    value = Column(String)

Base.metadata.create_all(bind=engine)

# Ensure DB schema matches current models (adds new columns if missing)
def ensure_schema():
    """Lightweight auto-migration for SQLite.
    Adds missing columns on existing tables to avoid OperationalError when models evolve.
    """
    with engine.begin() as conn:
        # Leads table columns
        try:
            rows = conn.exec_driver_sql("PRAGMA table_info('leads')").fetchall()
            existing_columns = {row[1] for row in rows}
        except Exception:
            existing_columns = set()

        # Add columns if they don't exist
        if 'uploaded_by' not in existing_columns:
            conn.exec_driver_sql("ALTER TABLE leads ADD COLUMN uploaded_by VARCHAR")
        if 'uploaded_by_id' not in existing_columns:
            conn.exec_driver_sql("ALTER TABLE leads ADD COLUMN uploaded_by_id INTEGER")
        if 'uploaded_at' not in existing_columns:
            conn.exec_driver_sql("ALTER TABLE leads ADD COLUMN uploaded_at DATETIME")
        if 'assigned_to' not in existing_columns:
            conn.exec_driver_sql("ALTER TABLE leads ADD COLUMN assigned_to VARCHAR")
        if 'status' not in existing_columns:
            conn.exec_driver_sql("ALTER TABLE leads ADD COLUMN status VARCHAR DEFAULT 'new'")

        # Optional: ensure indexes exist (SQLite auto-creates pk index; skip others for simplicity)

        # Deals table columns (create table if not exists then add new columns if missing)
        conn.exec_driver_sql("""
        CREATE TABLE IF NOT EXISTS deals (
            id INTEGER PRIMARY KEY,
            customer_name VARCHAR NOT NULL,
            phone VARCHAR NOT NULL,
            payment_screenshot BLOB NOT NULL,
            uploaded_by VARCHAR,
            uploaded_by_id INTEGER,
            created_at DATETIME
        )
        """)
        try:
            rows_deals = conn.exec_driver_sql("PRAGMA table_info('deals')").fetchall()
            existing_deals_cols = {row[1] for row in rows_deals}
        except Exception:
            existing_deals_cols = set()
        if 'uploaded_by' not in existing_deals_cols:
            conn.exec_driver_sql("ALTER TABLE deals ADD COLUMN uploaded_by VARCHAR")
        if 'uploaded_by_id' not in existing_deals_cols:
            conn.exec_driver_sql("ALTER TABLE deals ADD COLUMN uploaded_by_id INTEGER")
        if 'created_at' not in existing_deals_cols:
            conn.exec_driver_sql("ALTER TABLE deals ADD COLUMN created_at DATETIME")

        # Login events table
        conn.exec_driver_sql("""
        CREATE TABLE IF NOT EXISTS login_events (
            id INTEGER PRIMARY KEY,
            username VARCHAR NOT NULL,
            role VARCHAR NOT NULL,
            logged_in_at DATETIME
        )
        """)

        # Settings table
        conn.exec_driver_sql("""
        CREATE TABLE IF NOT EXISTS settings (
            key VARCHAR PRIMARY KEY,
            value VARCHAR
        )
        """)

ensure_schema()

# ----------------- Utility helpers -----------------

def get_session():
    return SessionLocal()

def get_user_by_username(db, username):
    return db.query(User).filter(User.username == username).first()

def create_user(db, username, password, role='salesman', name=None):
    if get_user_by_username(db, username):
        return None
    user = User(username=username, password_hash=User.hash_password(password), role=role, name=name or username)
    db.add(user)
    db.commit()
    db.refresh(user)
    return user

def update_or_create_user(db, username, password, role='salesman', name=None):
    """Update existing user or create new one"""
    user = get_user_by_username(db, username)
    if user:
        # Update existing user
        user.password_hash = User.hash_password(password)
        user.name = name or user.name
        user.role = role
        db.commit()
        db.refresh(user)
        return user
    else:
        # Create new user
        return create_user(db, username, password, role, name)

def ensure_demo_users():
    db = get_session()
    try:
        # Management roles
        update_or_create_user(db, 'head', 'IQstats@iq-2024', role='head_of_sales', name='Mohamed Akmal')
        update_or_create_user(db, 'cto', 'IQstats@iq-2025', role='cto', name='Omar Samy')
        update_or_create_user(db, 'ceo', 'IQstats@iq-2026', role='ceo', name='ENG Ahmed Essam')
        # Sales team
        update_or_create_user(db, 'toqa', 'IQstats@iq-2027', role='salesman', name='Toqa Amin')
        update_or_create_user(db, 'mahmoud', 'IQstats@iq-2028', role='salesman', name='Mahmoud Fathalla')
        update_or_create_user(db, 'mazen', 'IQstats@iq-2029', role='salesman', name='Mazen Ashraf')
        update_or_create_user(db, 'ahmed_malek', 'IQstats@iq-2030', role='salesman', name='Ahmed Malek')
        update_or_create_user(db, 'youssry', 'IQstats@iq-2031', role='salesman', name='Youssry Hassan')
    finally:
        db.close()

ensure_demo_users()

# ----------------- Activity logger -----------------

def log_activity(db, lead_id, actor, action, detail=None):
    act = Activity(lead_id=lead_id, actor=actor, action=action, detail=detail)
    db.add(act)
    db.commit()

# ----------------- Streamlit UI -----------------

st.set_page_config(page_title='IQ Stats CRM — Full', layout='wide')

# --- Authentication ---
if 'user' not in st.session_state:
    st.session_state['user'] = None

if st.session_state['user'] is None:
    st.title('IQ Stats CRM — Login')
    with st.form('login_form'):
        username = st.text_input('Username')
        password = st.text_input('Password', type='password')
        submitted = st.form_submit_button('Login')
        if submitted:
            db = get_session()
            user = get_user_by_username(db, username)
            if user and user.verify_password(password):
                st.session_state['user'] = user.username
                st.session_state['role'] = user.role
                # log login event
                try:
                    evt = LoginEvent(username=user.username, role=user.role, logged_in_at=datetime.utcnow())
                    db.add(evt)
                    db.commit()
                except Exception:
                    pass
                st.success(f'Welcome, {user.name} ({user.role})')
                st.rerun()
            else:
                st.error('Invalid username or password')


    st.stop()

# Load current user
current_user = None
with get_session() as db:
    current_user = get_user_by_username(db, st.session_state['user'])

if current_user is None:
    st.error('User not found. Please login again.')
    st.session_state.clear()
    st.stop()

role = current_user.role

# Top bar
st.sidebar.title('IQ Stats CRM')
st.sidebar.write(f'Logged in as **{current_user.name}** — *{role}*')
if st.sidebar.button('Logout'):
    for k in list(st.session_state.keys()):
        del st.session_state[k]
    st.rerun()

# Admin: manage users
if role == 'admin':
    st.header('Admin — System Management')
    
    # User Management Section
    st.subheader('👥 User Management')
    db = get_session()
    users = db.query(User).all()
    
    # Display current users
    st.write('**Current Users:**')
    cols = st.columns([1,2,2,1])
    cols[0].write('ID')
    cols[1].write('Username')
    cols[2].write('Name')
    cols[3].write('Role')
    for u in users:
        c0, c1, c2, c3 = st.columns([1,2,2,1])
        c0.write(u.id)
        c1.write(u.username)
        c2.write(u.name)
        c3.write(u.role)

    # Create new user form
    st.markdown('---')
    st.subheader('➕ Create New User')
    with st.form('create_user'):
        col1, col2 = st.columns(2)
        with col1:
            new_username = st.text_input('Username', placeholder='e.g., john_doe')
            new_name = st.text_input('Full Name', placeholder='e.g., John Doe')
            new_role = st.selectbox('Role', ['salesman','head_of_sales','cto','ceo','admin'])
        with col2:
            new_pass = st.text_input('Password', type='password', placeholder='Enter password')
            confirm_pass = st.text_input('Confirm Password', type='password', placeholder='Confirm password')
            st.write('**Password Format:** IQstats@iq-XXXX (recommended)')
        
        create = st.form_submit_button('✅ Create User')
        if create:
            if new_username and new_pass and new_name:
                if new_pass == confirm_pass:
                    res = create_user(db, new_username, new_pass, role=new_role, name=new_name)
                    if res:
                        st.success(f'✅ User "{new_username}" created successfully!')
                        st.rerun()
                    else:
                        st.error('❌ Username already exists')
                else:
                    st.error('❌ Passwords do not match')
            else:
                st.error('❌ Please fill all required fields')
    
    # User Actions
    st.markdown('---')
    st.subheader('🔧 User Actions')
    col1, col2 = st.columns(2)
    
    with col1:
        st.write('**Delete User:**')
        if users:
            user_to_delete = st.selectbox('Select user to delete', 
                                        options=[f"{u.username} ({u.name})" for u in users if u.username != 'admin'],
                                        key='delete_user')
            if st.button('🗑️ Delete User', type='secondary'):
                if user_to_delete:
                    username = user_to_delete.split(' (')[0]
                    user = db.query(User).filter(User.username == username).first()
                    if user:
                        db.delete(user)
                        db.commit()
                        st.success(f'✅ User "{username}" deleted successfully!')
                        st.rerun()
    
    with col2:
        st.write('**Update User Password:**')
        if users:
            user_to_update = st.selectbox('Select user to update', 
                                        options=[f"{u.username} ({u.name})" for u in users],
                                        key='update_user')
            new_password = st.text_input('New Password', type='password', key='new_pwd')
            if st.button('🔐 Update Password', type='secondary'):
                if user_to_update and new_password:
                    username = user_to_update.split(' (')[0]
                    user = db.query(User).filter(User.username == username).first()
                    if user:
                        user.password_hash = User.hash_password(new_password)
                        db.commit()
                        st.success(f'✅ Password updated for "{username}"!')
                        st.rerun()
    
    db.close()
    
    # System Maintenance Section
    st.markdown('---')
    st.subheader('⚙️ System Maintenance')
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write('**Database Management:**')
        if st.button('🔄 Reset Database (Drop & Recreate)', type='primary'):
            try:
                # Close sessions and remove DB file
                try:
                    del st.session_state['user']
                except Exception:
                    pass
                if os.path.exists(DB_FILE):
                    os.remove(DB_FILE)
                # Recreate
                Base.metadata.create_all(bind=engine)
                ensure_schema()
                ensure_demo_users()
                st.success('✅ Database reset successfully! Please reload the app.')
                st.rerun()
            except Exception as e:
                st.error(f'❌ Failed to reset database: {e}')
        
        if st.button('📊 Export All Data', type='secondary'):
            try:
                from datetime import datetime
                import zipfile
                
                now = datetime.now()
                date_str = now.strftime('%Y-%m-%d_%H-%M')
                zip_filename = f"Admin_Data_Export_{date_str}.zip"
                
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    # Export users
                    users_df = pd.DataFrame([{
                        'id': u.id, 'username': u.username, 'name': u.name, 
                        'role': u.role, 'created_at': u.created_at
                    } for u in users])
                    users_buffer = io.BytesIO()
                    with pd.ExcelWriter(users_buffer, engine='openpyxl') as writer:
                        users_df.to_excel(writer, sheet_name='Users', index=False)
                    zip_file.writestr('Users.xlsx', users_buffer.getvalue())
                    
                    # Export leads
                    leads_df, _ = read_leads_df(limit=100000)
                    if not leads_df.empty:
                        leads_buffer = io.BytesIO()
                        with pd.ExcelWriter(leads_buffer, engine='openpyxl') as writer:
                            leads_df.to_excel(writer, sheet_name='Leads', index=False)
                        zip_file.writestr('Leads.xlsx', leads_buffer.getvalue())
                    
                    # Export deals
                    deals_df, _ = read_deals_df(limit=100000)
                    if not deals_df.empty:
                        deals_buffer = io.BytesIO()
                        with pd.ExcelWriter(deals_buffer, engine='openpyxl') as writer:
                            deals_df.to_excel(writer, sheet_name='Deals', index=False)
                        zip_file.writestr('Deals.xlsx', deals_buffer.getvalue())
                    
                    # README
                    readme_content = f"""Admin Data Export
Generated on: {now.strftime('%A, %B %d, %Y at %H:%M')}
Generated by: Admin

Contents:
- Users.xlsx: All user accounts and roles
- Leads.xlsx: All leads data
- Deals.xlsx: All deals data

This is a complete system backup for administrative purposes.
"""
                    zip_file.writestr('README.txt', readme_content)
                
                st.download_button(
                    label=f'📥 Download {zip_filename}',
                    data=zip_buffer.getvalue(),
                    file_name=zip_filename,
                    mime='application/zip'
                )
                st.success('✅ Data export ready!')
                
            except Exception as e:
                st.error(f'❌ Error creating data export: {str(e)}')
    
    with col2:
        st.write('**System Information:**')
        st.info(f"""
        **Database File:** {DB_FILE}
        **Total Users:** {len(users)}
        **Database Size:** {os.path.getsize(DB_FILE) / 1024:.1f} KB
        **Last Modified:** {datetime.fromtimestamp(os.path.getmtime(DB_FILE)).strftime('%Y-%m-%d %H:%M')}
        """)
        
        st.write('**Quick Actions:**')
        if st.button('🔄 Refresh Page', type='secondary'):
            st.rerun()
        
        if st.button('📋 Copy System Info', type='secondary'):
            st.code(f"""
Database: {DB_FILE}
Users: {len(users)}
Size: {os.path.getsize(DB_FILE) / 1024:.1f} KB
Modified: {datetime.fromtimestamp(os.path.getmtime(DB_FILE)).strftime('%Y-%m-%d %H:%M')}
            """)
    
    st.stop()

# Common functions for leads

def read_leads_df(filters=None, search=None, order_by='uploaded_at', desc=True, limit=100, offset=0):
    db = get_session()
    q = db.query(Lead)
    if filters:
        for k, v in filters.items():
            if k == 'sales_agent' and v != 'All':
                q = q.filter(Lead.sales_agent == v)
            if k == 'status' and v != 'All':
                q = q.filter(Lead.status == v)
    if search:
        like = f"%{search}%"
        q = q.filter(sa.or_(Lead.name.ilike(like), Lead.number.ilike(like), Lead.contact.ilike(like)))
    total = q.count()
    if order_by:
        col = getattr(Lead, order_by)
        q = q.order_by(col.desc() if desc else col)
    q = q.offset(offset).limit(limit)
    df = pd.read_sql(q.statement, q.session.bind)
    db.close()
    return df, total

def read_deals_df(filters=None, search=None, order_by='created_at', desc=True, limit=1000, offset=0):
    db = get_session()
    q = db.query(Deal)
    if filters:
        for k, v in filters.items():
            if k == 'uploaded_by' and v != 'All':
                q = q.filter(Deal.uploaded_by == v)
    if search:
        like = f"%{search}%"
        q = q.filter(sa.or_(Deal.customer_name.ilike(like), Deal.phone.ilike(like)))
    total = q.count()
    if order_by:
        col = getattr(Deal, order_by)
        q = q.order_by(col.desc() if desc else col)
    q = q.offset(offset).limit(limit)
    # Exclude large binary column when reading to DataFrame
    cols = [Deal.id, Deal.customer_name, Deal.phone, Deal.uploaded_by, Deal.uploaded_by_id, Deal.created_at]
    q = db.query(*cols).filter(Deal.id.in_(db.query(Deal.id).subquery())) if True else q
    df = pd.read_sql(q.statement, q.session.bind)
    db.close()
    return df, total

# ----------------- Export helpers -----------------
def build_deals_excel_with_images(deals):
    """Create an Excel workbook with deals and embedded screenshots if possible.
    Returns bytes, or None if image embedding is unavailable.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
    except Exception:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = 'deals'
    headers = ['id', 'customer_name', 'phone', 'uploaded_by', 'created_at', 'screenshot']
    ws.append(headers)

    # Wider column for screenshot
    try:
        ws.column_dimensions['F'].width = 40
    except Exception:
        pass

    for idx, d in enumerate(deals, start=2):
        ws.cell(row=idx, column=1, value=d.id)
        ws.cell(row=idx, column=2, value=d.customer_name)
        ws.cell(row=idx, column=3, value=d.phone)
        ws.cell(row=idx, column=4, value=d.uploaded_by)
        ws.cell(row=idx, column=5, value=(d.created_at.strftime('%Y-%m-%d %H:%M') if d.created_at else None))
        # Try embedding image
        if getattr(d, 'payment_screenshot', None):
            try:
                img_stream = io.BytesIO(d.payment_screenshot)
                img = XLImage(img_stream)
                cell = f'F{idx}'
                ws.add_image(img, cell)
                # Increase row height for visibility
                try:
                    ws.row_dimensions[idx].height = 120
                except Exception:
                    pass
            except Exception:
                ws.cell(row=idx, column=6, value='[image available]')
        else:
            ws.cell(row=idx, column=6, value='')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def build_deals_images_zip(deals):
    """Create a zip containing deals' screenshots. Returns bytes."""
    def _detect_ext(content: bytes) -> str:
        # PNG
        if content[:8] == b'\x89PNG\r\n\x1a\n':
            return 'png'
        # JPEG
        if content[:3] == b'\xff\xd8\xff':
            return 'jpg'
        # WEBP (RIFF....WEBP)
        if len(content) >= 12 and content[:4] == b'RIFF' and content[8:12] == b'WEBP':
            return 'webp'
        return 'png'
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        for d in deals:
            content = getattr(d, 'payment_screenshot', None)
            if not content:
                continue
            ext = _detect_ext(content)
            zf.writestr(f'deal_{d.id}_payment.{ext}', content)
    return buf.getvalue()

# Layout by role
if role == 'salesman':
    st.header('Sales')
    tab_leads, tab_deals = st.tabs(['Upload & Manage Leads', 'Done Deals'])
    with tab_leads:
        # Check central lock for uploads
        with get_session() as _db_set:
            lock_row = _db_set.query(Setting).filter(Setting.key == 'uploads_locked').first()
            uploads_locked = (lock_row and (lock_row.value or '').lower() in ('1','true','yes','on'))
        uploaded_file = None
        if uploads_locked:
            st.warning('Uploads are temporarily disabled by CTO/Admin. You can still manage existing leads below.')
        else:
            uploaded_file = st.file_uploader('Upload XLSX/CSV with headers: number, name, sales agent, CONTACT, CASE, FEED BACK', type=['xlsx','xls','csv'])
        # Provide a ready-to-use template
        if st.button('Download Excel template'):
            template = pd.DataFrame({
                'number': [],
                'name': [],
                'sales agent': [],
                'contact': [],
                'case': [],
                'feed back': [],
                'how to contact': [],
            })
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                template.to_excel(writer, index=False, sheet_name='template')
            st.download_button('Download template.xlsx', data=buf.getvalue(), file_name='crm_leads_template.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        if uploaded_file is not None:
            try:
                if uploaded_file.name.endswith('.csv'):
                    df = pd.read_csv(uploaded_file)
                else:
                    df = pd.read_excel(uploaded_file)
            except Exception as e:
                st.error(f'Failed to read file: {e}')
                st.stop()

            st.subheader('Preview')
            st.dataframe(df.head())

        if not uploads_locked and st.button('Save to CRM'):
            # map columns
            col_map = {c.lower().strip(): c for c in df.columns}
            mapping = {}
            expected = ['number','name','sales agent','contact','case','feed back']
            for k in expected:
                if k in col_map:
                    mapping[col_map[k]] = k
            alt_map = {
                'sales_agent':'sales agent','salesagent':'sales agent',
                'feedback':'feed back','case_desc':'case',
                'contact_number':'contact','phone':'contact',
                'how to contact':'contact','how_to_contact':'contact','contact method':'contact'
            }
            for k,v in alt_map.items():
                if k in col_map and v not in mapping.values():
                    mapping[col_map[k]] = v
            df_ren = df.rename(columns=mapping)
            for tgt in ['number','name','sales agent','contact','case','feed back']:
                if tgt not in df_ren.columns:
                    df_ren[tgt] = None
            leads = []
            db = get_session()
            for _, row in df_ren.iterrows():
                lead = Lead(number=str(row['number']) if row['number'] is not None else None,
                            name=str(row['name']) if row['name'] is not None else None,
                            sales_agent=str(row['sales agent']) if row['sales agent'] is not None else current_user.username,
                            contact=str(row['contact']) if row['contact'] is not None else None,
                            case_desc=str(row['case']) if row['case'] is not None else None,
                            feedback=str(row['feed back']) if row['feed back'] is not None else None,
                            uploaded_by=current_user.username,
                            uploaded_by_id=current_user.id,
                            uploaded_at=datetime.utcnow())
                db.add(lead)
                db.flush()
                log_activity(db, lead.id, current_user.username, 'upload', detail='Bulk upload')
                leads.append(lead)
            db.commit()
            db.close()
            st.success(f'Saved {len(leads)} leads')

        st.markdown('---')
        st.subheader('My Leads')
        page = st.number_input('Page', min_value=1, value=1, key='leads_page')
        page_size = st.selectbox('Page size', [10,25,50], index=0, key='leads_page_size')
        offset = (page-1)*page_size
        df, total = read_leads_df(filters={'sales_agent': current_user.username}, limit=page_size, offset=offset)
        st.write(f'Total: {total}')
        if not df.empty:
            # Prepare editable table
            contact_options = ['call', 'call and whatsapp', "didn't reach"]
            status_options = ['new','contacted','qualified','lost','won']
            case_options = ['general', 'pricing', 'technical', 'support', 'complaint', 'other']
            feedback_options = ['positive', 'neutral', 'negative', 'not interested', 'call later', 'wrong number', 'closed won', 'closed lost', 'other']
            # pull sales users for dropdowns
            with get_session() as _db:
                sales_users = [u.username for u in _db.query(User).filter(User.role=='salesman').all()]
            editable_cols = ['number','name','sales_agent','contact','case_desc','feedback','status','assigned_to','comment_text']
            base_cols = ['id','number','name','sales_agent','contact','case_desc','feedback','status','assigned_to']
            present_cols = [c for c in base_cols if c in df.columns]
            df_edit = df[present_cols].copy()
            # Ensure all expected columns exist in editor
            for c in base_cols:
                if c not in df_edit.columns:
                    df_edit[c] = None
            # UI-only column for creating new comments
            df_edit['comment_text'] = ''
            df_edit['sales_agent'] = df_edit['sales_agent'].fillna(current_user.username)
            df_edit['assigned_to'] = df_edit['assigned_to'].fillna(current_user.username)
            edited = st.data_editor(
                df_edit,
                column_config={
                    'contact': st.column_config.SelectboxColumn('HOW TO CONTACT', options=contact_options, help='How did you contact?'),
                    'status': st.column_config.SelectboxColumn('Status', options=status_options),
                    'sales_agent': st.column_config.SelectboxColumn('Sales Agent', options=sales_users),
                    'assigned_to': st.column_config.SelectboxColumn('Assigned To', options=sales_users),
                    'case_desc': st.column_config.SelectboxColumn('CASE', options=sorted(set(case_options + [c for c in df_edit['case_desc'].dropna().astype(str).tolist()] ))),
                    'feedback': st.column_config.SelectboxColumn('FEED BACK', options=sorted(set(feedback_options + [c for c in df_edit['feedback'].dropna().astype(str).tolist()] ))),
                    'comment_text': st.column_config.TextColumn('Comment (new)'),
                },
                disabled=['id'],
                use_container_width=True,
                key='leads_editor'
            )
            if st.button('Save table changes', key='save_table_changes'):
                import pandas as _pd
                db = get_session()
                try:
                    orig = df_edit.set_index('id')
                    cur = edited.set_index('id', drop=False)
                    # Updates for existing ids
                    intersect_ids = [i for i in cur.index if _pd.notna(i)]
                    for lead_id in intersect_ids:
                        if lead_id not in orig.index:
                            continue
                        row_orig = orig.loc[lead_id]
                        row_new = cur.loc[lead_id]
                        changed = False
                        lead = db.query(Lead).get(int(lead_id))
                        for col in editable_cols:
                            if col == 'comment_text':
                                continue
                            new_val = row_new[col]
                            old_val = row_orig[col]
                            if (isinstance(new_val, float) and _pd.isna(new_val)) or new_val == '':
                                new_val = None
                            if (isinstance(old_val, float) and _pd.isna(old_val)) or old_val == '':
                                old_val = None
                            if new_val != old_val:
                                setattr(lead, col if col != 'case_desc' else 'case_desc', new_val)
                                changed = True
                        if changed:
                            db.add(lead)
                            db.commit()
                            log_activity(db, lead.id, current_user.username, 'edit', detail='Edited via table')
                        # Add a comment if provided
                        comment_text = (row_new.get('comment_text') or '').strip()
                        if comment_text:
                            com = Comment(lead_id=lead.id, author=current_user.username, text=comment_text)
                            db.add(com)
                            db.commit()
                            log_activity(db, lead.id, current_user.username, 'comment', detail=comment_text)
                    # Inserts for new rows (id is NaN)
                    new_rows = edited[_pd.isna(edited['id'])]
                    for _, r in new_rows.iterrows():
                        lead = Lead(
                            number=r.get('number') or None,
                            name=r.get('name') or None,
                            sales_agent=r.get('sales_agent') or current_user.username,
                            contact=r.get('contact') or None,
                            case_desc=r.get('case_desc') or None,
                            feedback=r.get('feedback') or None,
                            status=r.get('status') or 'new',
                            assigned_to=r.get('assigned_to') or None,
                            uploaded_by=current_user.username,
                            uploaded_by_id=current_user.id,
                            uploaded_at=datetime.utcnow()
                        )
                        db.add(lead)
                        db.commit()
                        log_activity(db, lead.id, current_user.username, 'create', detail='Added via table')
                        comment_text = (r.get('comment_text') or '').strip()
                        if comment_text:
                            com = Comment(lead_id=lead.id, author=current_user.username, text=comment_text)
                            db.add(com)
                            db.commit()
                            log_activity(db, lead.id, current_user.username, 'comment', detail=comment_text)
                finally:
                    db.close()
                st.success('Changes saved')
        else:
            st.info('No leads yet')

    with tab_deals:
        st.subheader('Done Deals — Add new')
        with st.form('deal_form'):
            customer_name = st.text_input('Customer name')
            phone = st.text_input('Phone number')
            screenshot_file = st.file_uploader('Payment screenshot (image)', type=['png','jpg','jpeg','webp'])
            submit_deal = st.form_submit_button('Save deal')
            if submit_deal:
                if not customer_name or not phone or not screenshot_file:
                    st.error('Please provide name, phone, and payment screenshot')
                else:
                    try:
                        screenshot_bytes = screenshot_file.read()
                        db = get_session()
                        deal = Deal(
                            customer_name=customer_name,
                            phone=phone,
                            payment_screenshot=screenshot_bytes,
                            uploaded_by=current_user.username,
                            uploaded_by_id=current_user.id,
                        )
                        db.add(deal)
                        db.commit()
                        st.success('Deal saved')
                        db.close()
                    except Exception as e:
                        st.error(f'Failed to save deal: {e}')

        st.subheader('My Deals')
        # Simple list with preview and download
        try:
            db = get_session()
            # Load only metadata first
            deals = db.query(Deal).filter(Deal.uploaded_by==current_user.username).order_by(Deal.created_at.desc()).all()
            for d in deals:
                with st.expander(f"{d.created_at:%Y-%m-%d %H:%M} — {d.customer_name} ({d.phone})"):
                    st.write(f"Salesman: {d.uploaded_by}")
                    # Display image if possible
                    try:
                        st.image(d.payment_screenshot, caption='Payment screenshot', use_container_width=True)
                    except Exception:
                        st.download_button('Download screenshot', data=d.payment_screenshot, file_name=f'deal_{d.id}_payment.png')

            # Downloads: Excel with embedded images (if supported) and ZIP of images
            if deals:
                excel_bytes = build_deals_excel_with_images(deals)
                if excel_bytes:
                    st.download_button('Download my deals (Excel with images)', data=excel_bytes, file_name='my_deals_with_images.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                zip_bytes = build_deals_images_zip(deals)
                st.download_button('Download my deal screenshots (ZIP)', data=zip_bytes, file_name='my_deal_screenshots.zip', mime='application/zip')
        finally:
            try:
                db.close()
            except Exception:
                pass

elif role == 'head_of_sales':
    st.header('Head of Sales — Overview')
    db = get_session()
    agents = [r[0] for r in db.query(Lead.sales_agent).distinct().all() if r[0]]
    db.close()
    sel_agent = st.selectbox('Filter by agent', options=['All'] + agents)
    sel_status = st.selectbox('Filter by status', options=['All','new','contacted','qualified','lost','won'])
    search = st.text_input('Search (name, number, contact)')
    page = st.number_input('Page', min_value=1, value=1)
    page_size = st.selectbox('Page size', [10,25,50], index=1)
    offset = (page-1)*page_size
    filters = {}
    if sel_agent: filters['sales_agent'] = sel_agent
    if sel_status: filters['status'] = sel_status
    df, total = read_leads_df(filters=filters, search=search, limit=page_size, offset=offset)
    st.write(f'Total matches: {total}')
    st.dataframe(df)
    if not df.empty:
        # quick KPIs
        st.subheader('KPIs')
        c1, c2, c3 = st.columns(3)
        c1.metric('Total Leads', total)
        vc = df['sales_agent'].value_counts()
        top_agent = (vc.idxmax() if not vc.empty else '—')
        c2.metric('Top Agent', top_agent)
        c3.metric('Unique Contacts', int(df['contact'].nunique()))

    st.markdown('---')
    st.subheader('Recently Assigned to Salesmen')
    with get_session() as _db_recent:
        recent = (
            _db_recent.query(Lead)
            .filter(Lead.assigned_to.isnot(None))
            .order_by(Lead.uploaded_at.desc())
            .limit(50)
            .all()
        )
        if recent:
            rec_df = pd.DataFrame([
                {
                    'id': l.id,
                    'name': l.name,
                    'number': l.number,
                    'assigned_to': l.assigned_to,
                    'status': l.status,
                    'uploaded_at': l.uploaded_at
                } for l in recent
            ])
            st.dataframe(rec_df)
        else:
            st.info('No assignments yet.')

elif role == 'cto':
    st.header('CTO Dashboard — Analytics')
    # --- CTO uploader (centralized uploads) ---
    with st.expander('Upload Leads (CTO)'):
        # Salesmen list for default assignment
        with get_session() as _db_cto_sales:
            cto_salesmen = [u.username for u in _db_cto_sales.query(User).filter(User.role=='salesman').order_by(User.username.asc()).all()]
        default_agent_opt = ['(keep from file)'] + cto_salesmen
        default_agent = st.selectbox('Default sales agent (optional)', options=default_agent_opt)
        cto_uploaded = st.file_uploader('XLSX/CSV with headers: number, name, sales agent, CONTACT, CASE, FEED BACK', type=['xlsx','xls','csv'], key='cto_uploader')
        col_left, col_right = st.columns(2)
        with col_left:
            if st.button('Download Excel template (CTO)'):
                template = pd.DataFrame({
                    'number': [],
                    'name': [],
                    'sales agent': [],
                    'contact': [],
                    'case': [],
                    'feed back': [],
                    'how to contact': [],
                })
                buf_t = io.BytesIO()
                with pd.ExcelWriter(buf_t, engine='openpyxl') as writer:
                    template.to_excel(writer, index=False, sheet_name='template')
                st.download_button('Download CTO template.xlsx', data=buf_t.getvalue(), file_name='crm_leads_template_cto.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        if cto_uploaded is not None:
            try:
                if cto_uploaded.name.endswith('.csv'):
                    cto_df = pd.read_csv(cto_uploaded)
                else:
                    cto_df = pd.read_excel(cto_uploaded)
            except Exception as e:
                st.error(f'Failed to read file: {e}')
                cto_df = None
            if cto_df is not None:
                st.dataframe(cto_df.head())
                if st.button('Save uploaded leads (CTO)'):
                    col_map = {c.lower().strip(): c for c in cto_df.columns}
                    mapping = {}
                    expected = ['number','name','sales agent','contact','case','feed back']
                    for k in expected:
                        if k in col_map:
                            mapping[col_map[k]] = k
                    alt_map = {
                        'sales_agent':'sales agent','salesagent':'sales agent',
                        'feedback':'feed back','case_desc':'case',
                        'contact_number':'contact','phone':'contact',
                        'how to contact':'contact','how_to_contact':'contact','contact method':'contact'
                    }
                    for k,v in alt_map.items():
                        if k in col_map and v not in mapping.values():
                            mapping[col_map[k]] = v
                    df_ren = cto_df.rename(columns=mapping)
                    for tgt in ['number','name','sales agent','contact','case','feed back']:
                        if tgt not in df_ren.columns:
                            df_ren[tgt] = None
                    saved = 0
                    with get_session() as db:
                        for _, row in df_ren.iterrows():
                            agent_from_file = str(row['sales agent']) if row['sales agent'] is not None else None
                            final_agent = agent_from_file
                            if (not final_agent or final_agent.strip()=='') and default_agent != '(keep from file)':
                                final_agent = default_agent
                            lead = Lead(
                                number=str(row['number']) if row['number'] is not None else None,
                                name=str(row['name']) if row['name'] is not None else None,
                                sales_agent=str(final_agent) if final_agent else None,
                                contact=str(row['contact']) if row['contact'] is not None else None,
                                case_desc=str(row['case']) if row['case'] is not None else None,
                                feedback=str(row['feed back']) if row['feed back'] is not None else None,
                                uploaded_by=current_user.username,
                                uploaded_by_id=getattr(current_user, 'id', None),
                                uploaded_at=datetime.utcnow(),
                                assigned_to=str(final_agent) if final_agent else None,
                            )
                            db.add(lead)
                            db.flush()
                            log_activity(db, lead.id, current_user.username, 'upload', detail='CTO upload')
                            saved += 1
                        db.commit()
                    st.success(f'Saved {saved} leads')
                    st.rerun()

    df_all, total = read_leads_df(limit=10000, offset=0)
    if df_all.empty:
        st.info('No data yet. Upload leads (Sales tab) or generate demo leads below.')
        with st.expander('Demo tools (dev)'):
            n = st.number_input('How many demo leads to generate?', min_value=10, max_value=2000, value=200, step=10)
            if st.button('Generate demo leads'):
                try:
                    with get_session() as db:
                        sales_users = [u.username for u in db.query(User).filter(User.role=='salesman').all()]
                        if not sales_users:
                            u = create_user(db, 'sales_auto', 'pass', role='salesman', name='Sales Auto')
                            sales_users = [u.username]
                        contact_opts = ['call', 'call and whatsapp', "didn't reach"]
                        case_opts = ['general', 'pricing', 'technical', 'support', 'complaint', 'other']
                        feedback_opts = ['positive', 'neutral', 'negative', 'not interested', 'call later', 'wrong number', 'closed won', 'closed lost', 'other']
                        status_opts = ['new','contacted','qualified','won','lost']
                        now = datetime.utcnow()
                        for i in range(int(n)):
                            agent = random.choice(sales_users)
                            lead = Lead(
                                number=str(100000 + random.randint(0, 999999)),
                                name=f"Lead {i+1}",
                                sales_agent=agent,
                                contact=random.choice(contact_opts),
                                case_desc=random.choice(case_opts),
                                feedback=random.choice(feedback_opts),
                                status=random.choice(status_opts),
                                assigned_to=agent,
                                uploaded_by='demo',
                                uploaded_at=now - pd.Timedelta(days=random.randint(0, 60), hours=random.randint(0,23))
                            )
                            db.add(lead)
                        db.commit()
                    st.success('Demo leads generated. Reloading dashboard...')
                    st.rerun()
                except Exception as e:
                    st.error(f'Failed to generate demo leads: {e}')
    else:
        # Helpers for categorical charts
        def _normalize_text(value: str):
            if value is None:
                return None
            try:
                text = str(value).strip().lower()
            except Exception:
                return None
            if text in {'', 'na', 'n/a', 'none', 'null', '-'}:
                return None
            return text

        def _plot_top_categories(df: pd.DataFrame, column: str, title: str, top_n: int = 20):
            series = df[column].dropna().astype(str).map(_normalize_text).dropna()
            if series.empty:
                st.info(f'No values in {column} to plot.')
                return
            counts = series.value_counts().head(top_n).reset_index()
            counts.columns = [column, 'count']
            fig_local = px.bar(counts, x=column, y='count', title=title)
            st.plotly_chart(fig_local, use_container_width=True)

        df_all['uploaded_at'] = pd.to_datetime(df_all['uploaded_at'])
        df_all['date'] = df_all['uploaded_at'].dt.date

        # Filters
        min_date = df_all['date'].min()
        max_date = df_all['date'].max()
        c_f1, c_f2, c_f3 = st.columns([2,2,3])
        with c_f1:
            date_range = st.date_input('Date range', value=(min_date, max_date))
        with c_f2:
            agents_all = sorted([a for a in df_all['sales_agent'].dropna().unique().tolist()])
            sel_agents = st.multiselect('Agents', options=agents_all, default=agents_all)
        with c_f3:
            status_all = ['new','contacted','qualified','won','lost']
            sel_statuses = st.multiselect('Statuses', options=status_all, default=status_all)

        # Apply filters
        df_f = df_all.copy()
        if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
            start_d, end_d = pd.to_datetime(date_range[0]), pd.to_datetime(date_range[1])
            df_f = df_f[(df_f['uploaded_at'] >= start_d) & (df_f['uploaded_at'] <= (end_d + pd.Timedelta(days=1)))]
        if sel_agents:
            df_f = df_f[df_f['sales_agent'].isin(sel_agents)]
        if sel_statuses:
            df_f = df_f[df_f['status'].isin(sel_statuses)]

        st.subheader('Leads over time')
        daily = df_f.groupby('date').size().reset_index(name='count')
        fig = px.line(daily, x='date', y='count', title='Leads per day')
        st.plotly_chart(fig, use_container_width=True)

        # Lead Assignment (CTO)
        st.markdown('---')
        st.subheader('Lead Assignment')
        # Toggle uploads lock for salesmen
        with get_session() as _db_lock:
            lock_row = _db_lock.query(Setting).filter(Setting.key=='uploads_locked').first()
            cur_locked = (lock_row and (lock_row.value or '').lower() in ('1','true','yes','on'))
        new_locked = st.toggle('Lock salesman uploads', value=cur_locked, help='Prevent salesmen from uploading files')
        if new_locked != cur_locked:
            with get_session() as _db_lock2:
                row = _db_lock2.query(Setting).filter(Setting.key=='uploads_locked').first()
                if not row:
                    row = Setting(key='uploads_locked', value='1' if new_locked else '0')
                    _db_lock2.add(row)
                else:
                    row.value = '1' if new_locked else '0'
                _db_lock2.commit()
            st.success('Upload lock setting updated')
        with get_session() as _db_assign:
            # consider unassigned as sales_agent is null/empty/'unassigned' OR assigned_to is null
            unassigned = (
                _db_assign.query(Lead)
                .filter(
                    sa.or_(
                        Lead.sales_agent.is_(None),
                        func.trim(Lead.sales_agent) == '',
                        func.lower(func.trim(Lead.sales_agent)) == 'unassigned',
                        Lead.assigned_to.is_(None)
                    )
                )
                .order_by(Lead.uploaded_at.desc())
                .all()
            )
            st.write(f"Unassigned leads: {len(unassigned)}")
        with get_session() as _db_users:
            salesman_users = [u.username for u in _db_users.query(User).filter(User.role == 'salesman').order_by(User.username.asc()).all()]
        target_agents = st.multiselect('Select salesmen to distribute to', options=salesman_users, default=salesman_users)
        max_to_assign = st.number_input('Max leads to distribute (0 = all)', min_value=0, value=0, step=1)
        if st.button('Distribute unassigned leads equally'):
            if not target_agents:
                st.warning('Select at least one salesman.')
            else:
                with get_session() as db:
                    leads = (
                        db.query(Lead)
                        .filter(
                            sa.or_(
                                Lead.sales_agent.is_(None),
                                func.trim(Lead.sales_agent) == '',
                                func.lower(func.trim(Lead.sales_agent)) == 'unassigned',
                                Lead.assigned_to.is_(None)
                            )
                        )
                        .order_by(Lead.uploaded_at.asc())
                        .all()
                    )
                    if max_to_assign and max_to_assign > 0:
                        leads = leads[:max_to_assign]
                    if not leads:
                        st.info('No unassigned leads to distribute.')
                    else:
                        random.shuffle(leads)
                        assigned_count = {a: 0 for a in target_agents}
                        for idx, lead in enumerate(leads):
                            agent = target_agents[idx % len(target_agents)]
                            lead.sales_agent = agent
                            lead.assigned_to = agent
                            db.add(lead)
                            db.flush()
                            log_activity(db, lead.id, current_user.username, 'assign', detail=f'Assigned to {agent} by CTO')
                            assigned_count[agent] += 1
                        db.commit()
                        st.success('Distribution completed: ' + ', '.join([f"{a}: {n}" for a, n in assigned_count.items()]))
                        st.rerun()

        # Demo: randomize statuses to see distribution (dev use)
        with st.expander('Demo tools (dev)'):
            if st.button('Randomize lead statuses for demo'):
                with get_session() as db:
                    statuses = ['new','contacted','qualified','lost','won']
                    all_leads = db.query(Lead).all()
                    for lead in all_leads:
                        lead.status = random.choice(statuses)
                    db.commit()
                st.success('Statuses randomized. Refresh charts above.')

        st.subheader('Leads by agent')
        agent_counts = df_f['sales_agent'].value_counts().reset_index()
        agent_counts.columns = ['sales_agent','count']
        fig2 = px.bar(agent_counts, x='sales_agent', y='count', title='Leads by Agent')
        st.plotly_chart(fig2, use_container_width=True)

        # Login events (visibility for CTO)
        st.markdown('---')
        st.subheader('User Login Activity (last 7 days)')
        try:
            with get_session() as db:
                seven_days_ago = datetime.utcnow() - pd.Timedelta(days=7)
                logins = db.query(LoginEvent).filter(LoginEvent.logged_in_at >= seven_days_ago).order_by(LoginEvent.logged_in_at.desc()).all()
                if logins:
                    df_login = pd.DataFrame([
                        {
                            'username': e.username,
                            'role': e.role,
                            'logged_in_at': e.logged_in_at
                        } for e in logins
                    ])
                    st.dataframe(df_login)
                    agg = df_login.groupby('username').size().reset_index(name='logins')
                    fig_login = px.bar(agg, x='username', y='logins', title='Logins per user (7 days)')
                    st.plotly_chart(fig_login, use_container_width=True)
                else:
                    st.info('No login events yet.')
        except Exception:
            st.info('Login activity not available yet.')

        st.subheader('Status breakdown')
        status_counts = df_f['status'].value_counts().reset_index()
        status_counts.columns = ['status','count']
        fig3 = px.pie(status_counts, names='status', values='count', title='Leads by Status')
        st.plotly_chart(fig3, use_container_width=True)

        # Leads by status per salesman (stacked)
        st.subheader('Leads by Status per Salesman')
        status_per_sales = (
            df_f.dropna(subset=['sales_agent','status'])
                  .groupby(['sales_agent','status'])
                  .size()
                  .reset_index(name='count')
        )
        if not status_per_sales.empty:
            fig_status_sales = px.bar(
                status_per_sales,
                x='sales_agent', y='count', color='status',
                barmode='stack', title='Leads by Status per Salesman'
            )
            st.plotly_chart(fig_status_sales, use_container_width=True)
        else:
            st.info('No data for status per salesman yet.')

        st.subheader('Feedback word cloud-ish (top words)')
        feedbacks = df_f['feedback'].dropna().astype(str)
        if not feedbacks.empty:
            all_text = ' '.join(feedbacks.tolist()).lower()
            words = [w.strip('.,!?:;()"\'') for w in all_text.split() if len(w) > 3]
            if words:
                wc = pd.Series(words).value_counts().head(30).reset_index()
                wc.columns = ['word','count']
                fig4 = px.bar(wc, x='word', y='count', title='Top feedback words')
                st.plotly_chart(fig4, use_container_width=True)
            else:
                st.info('No keywords found in feedback yet.')
        else:
            st.info('No feedback text yet to analyze.')

        st.subheader('HOW TO CONTACT — top categories')
        try:
            _plot_top_categories(df_f, 'contact', 'Contact method — top categories')
        except Exception:
            st.info('No values in contact to plot.')

        st.subheader('CASE — top categories')
        try:
            _plot_top_categories(df_f, 'case_desc', 'Case — top categories')
        except Exception:
            st.info('No values in case_desc to plot.')

        st.subheader('FEED BACK — top categories')
        try:
            _plot_top_categories(df_f, 'feedback', 'Feedback — top categories')
        except Exception:
            st.info('No values in feedback to plot.')

        # Additional charts
        st.subheader('Sales Pipeline — Funnel')
        try:
            ordered_status = ['new','contacted','qualified','won','lost']
            counts_map = {s: int((df_f['status'] == s).sum()) for s in ordered_status}
            funnel_df = pd.DataFrame({'status': list(counts_map.keys()), 'count': list(counts_map.values())})
            fig_funnel = px.funnel(funnel_df, x='count', y='status', title='Lead Funnel')
            st.plotly_chart(fig_funnel, use_container_width=True)
        except Exception:
            pass

        st.subheader('Contact method by agent')
        cm = df_f.dropna(subset=['sales_agent','contact'])
        if not cm.empty:
            cm_counts = cm.groupby(['sales_agent','contact']).size().reset_index(name='count')
            fig_cm = px.bar(cm_counts, x='sales_agent', y='count', color='contact', barmode='stack', title='Contact methods by agent')
            st.plotly_chart(fig_cm, use_container_width=True)

        st.subheader('Lead uploads heatmap (weekday x hour)')
        if not df_f.empty:
            tmp = df_f.copy()
            tmp['weekday'] = tmp['uploaded_at'].dt.day_name()
            tmp['hour'] = tmp['uploaded_at'].dt.hour
            heat = tmp.groupby(['weekday','hour']).size().reset_index(name='count')
            if not heat.empty:
                fig_heat = px.density_heatmap(heat, x='hour', y='weekday', z='count', histfunc='avg', title='Uploads heatmap')
                st.plotly_chart(fig_heat, use_container_width=True)

        st.subheader('Leads per day (7-day rolling avg)')
        if not daily.empty:
            daily_ra = daily.sort_values('date').copy()
            daily_ra['rolling_7d'] = daily_ra['count'].rolling(window=7, min_periods=1).mean()
            fig_ra = px.line(daily_ra, x='date', y=['count','rolling_7d'], labels={'value':'leads','variable':'series'}, title='Daily leads and 7d avg')
            st.plotly_chart(fig_ra, use_container_width=True)

        st.subheader('Export filtered leads')
        csv_bytes = df_f.to_csv(index=False).encode('utf-8')
        st.download_button('Download CSV (filtered)', data=csv_bytes, file_name='leads_filtered.csv', mime='text/csv')
        xls_buf = io.BytesIO()
        with pd.ExcelWriter(xls_buf, engine='openpyxl') as writer:
            df_f.to_excel(writer, index=False, sheet_name='leads_filtered')
        st.download_button('Download Excel (filtered)', data=xls_buf.getvalue(), file_name='leads_filtered.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # CTO Analytics Package Download
        st.markdown('---')
        st.subheader('📦 Download Complete Analytics Package')
        
        from datetime import datetime
        import zipfile
        
        # Generate filename with current day and date
        now = datetime.now()
        day_name = now.strftime('%A')  # Full day name (e.g., 'Monday')
        date_str = now.strftime('%Y-%m-%d')  # Date format (e.g., '2024-08-09')
        zip_filename = f"CTO_Analytics_{day_name}_{date_str}.zip"
        
        if st.button('📊 Generate & Download CTO Analytics Package'):
            try:
                # Create ZIP file in memory
                zip_buffer = io.BytesIO()
                
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    # Add all chart data as Excel sheets
                    excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                        # Main filtered data
                        df_f.to_excel(writer, sheet_name='Filtered_Leads', index=False)
                        
                        # Chart data from all the charts above
                        charts_data = {}
                        
                        # Daily leads
                        if not daily.empty:
                            daily.to_excel(writer, sheet_name='Daily_Leads', index=False)
                            charts_data['daily_leads'] = daily
                        
                        # Agent breakdown
                        if not agent_counts.empty:
                            agent_counts.to_excel(writer, sheet_name='Agent_Breakdown', index=False)
                            charts_data['agent_breakdown'] = agent_counts
                        
                        # Status breakdown
                        if not status_counts.empty:
                            status_counts.to_excel(writer, sheet_name='Status_Breakdown', index=False)
                            charts_data['status_breakdown'] = status_counts
                        
                        # Sales funnel
                        if 'funnel_df' in locals():
                            funnel_df.to_excel(writer, sheet_name='Sales_Funnel', index=False)
                            charts_data['sales_funnel'] = funnel_df
                        
                        # Contact methods
                        if 'cm_counts' in locals() and not cm_counts.empty:
                            cm_counts.to_excel(writer, sheet_name='Contact_Methods', index=False)
                            charts_data['contact_methods'] = cm_counts
                        
                        # Activity heatmap
                        if 'heat' in locals() and not heat.empty:
                            heat.to_excel(writer, sheet_name='Activity_Heatmap', index=False)
                            charts_data['activity_heatmap'] = heat
                        
                        # Trends analysis
                        if 'daily_ra' in locals():
                            daily_ra.to_excel(writer, sheet_name='Trends_Analysis', index=False)
                            charts_data['trends'] = daily_ra
                        
                        # CTO Summary statistics
                        summary_data = {
                            'Metric': ['Total Leads', 'Filtered Leads', 'Active Agents', 'Date Range', 'Generated By'],
                            'Value': [
                                len(df_all),
                                len(df_f),
                                len(agent_counts) if 'agent_breakdown' in charts_data else 0,
                                f"{date_range[0] if isinstance(date_range, (list, tuple)) else 'All'} to {date_range[1] if isinstance(date_range, (list, tuple)) else 'All'}",
                                'CTO Dashboard'
                            ]
                        }
                        pd.DataFrame(summary_data).to_excel(writer, sheet_name='CTO_Summary', index=False)
                    
                    # Add Excel file to ZIP
                    zip_file.writestr(f'CTO_Analytics_{date_str}.xlsx', excel_buffer.getvalue())
                    
                    # Add deals data if available
                    deals_df, _ = read_deals_df(limit=100000)
                    if not deals_df.empty:
                        deals_buffer = io.BytesIO()
                        with pd.ExcelWriter(deals_buffer, engine='openpyxl') as deals_writer:
                            deals_df.to_excel(deals_writer, sheet_name='All_Deals', index=False)
                            
                            # Deals summary
                            deals_summary = deals_df.groupby('uploaded_by').size().reset_index(name='deals_count')
                            deals_summary.to_excel(deals_writer, sheet_name='Deals_by_Agent', index=False)
                        
                        zip_file.writestr(f'CTO_Deals_Report_{date_str}.xlsx', deals_buffer.getvalue())
                    
                    # Add a README file
                    readme_content = f"""CTO Analytics Package
Generated on: {now.strftime('%A, %B %d, %Y at %H:%M')}
Generated for: CTO Dashboard

Contents:
- CTO_Analytics_{date_str}.xlsx: Complete analytics with all charts data
- CTO_Deals_Report_{date_str}.xlsx: Deals tracking and performance data

Chart Data Included:
- Daily leads trends
- Agent performance breakdown
- Lead status distribution
- Sales funnel analysis
- Contact method analysis
- Activity heatmaps
- Rolling averages and trends
- Filtered data based on current selections

CTO Summary:
- Total Leads: {len(df_all)}
- Filtered Leads: {len(df_f)}
- Active Agents: {len(agent_counts) if 'agent_breakdown' in charts_data else 0}

This package contains comprehensive technical analytics for system management.
"""
                    zip_file.writestr('README.txt', readme_content)
                
                # Offer download
                st.download_button(
                    label=f'📥 Download {zip_filename}',
                    data=zip_buffer.getvalue(),
                    file_name=zip_filename,
                    mime='application/zip',
                    help=f'Download complete CTO analytics package for {day_name}, {date_str}'
                )
                
                st.success(f'✅ CTO Analytics package ready! Contains all dashboard charts and technical reports for {day_name}, {date_str}')
                
            except Exception as e:
                st.error(f'Error creating CTO analytics package: {str(e)}')

        # Done deals charts (handles empty safely)
        st.markdown('---')
        st.subheader('Done Deals — Analytics')
        deals_df, _ = read_deals_df(limit=100000)
        if deals_df.empty:
            st.info('No deals yet')
        else:
            deals_df['created_at'] = pd.to_datetime(deals_df['created_at'])
            deals_df['date'] = deals_df['created_at'].dt.date
            if not deals_df.empty:
                dd_daily = deals_df.groupby('date').size().reset_index(name='deals')
                if not dd_daily.empty:
                    dd_fig = px.line(dd_daily, x='date', y='deals', title='Deals per day')
                    st.plotly_chart(dd_fig, use_container_width=True)
                dd_sales = deals_df['uploaded_by'].value_counts().reset_index()
                dd_sales.columns = ['salesman','deals']
                if not dd_sales.empty:
                    dd_fig2 = px.bar(dd_sales, x='salesman', y='deals', title='Deals by salesman')
                    st.plotly_chart(dd_fig2, use_container_width=True)

elif role == 'ceo':
    st.header('CEO — Executive Dashboard & Reports')
    df_all, total = read_leads_df(limit=10000, offset=0)
    if df_all.empty:
        st.info('No data yet')
    else:
        st.subheader('Executive KPIs')
        total_leads = total
        unique_contacts = int(df_all['contact'].nunique())
        vc_all = df_all['sales_agent'].value_counts()
        top_agent = (vc_all.idxmax() if not vc_all.empty else '—')
        c1, c2, c3 = st.columns(3)
        c1.metric('Total Leads', total_leads)
        c2.metric('Unique Contacts', unique_contacts)
        c3.metric('Top Agent', top_agent)

        # CTO Analytics Section for CEO
        st.markdown('---')
        st.subheader('📊 CTO Analytics Dashboard')
        
        # Create all the charts that CTO sees
        df_f = df_all.copy()
        charts_data = {}
        
        try:
            # 1. Time series chart
            if not df_f.empty:
                df_f['uploaded_at'] = pd.to_datetime(df_f['uploaded_at'])
                daily = df_f.groupby(df_f['uploaded_at'].dt.date).size().reset_index(name='count')
                daily.columns = ['date', 'count']
                if not daily.empty:
                    fig_time = px.line(daily, x='date', y='count', title='Leads per day')
                    st.plotly_chart(fig_time, use_container_width=True)
                    charts_data['daily_leads'] = daily
            
            # 2. Agent breakdown
            agent_counts = df_f['sales_agent'].value_counts().reset_index()
            agent_counts.columns = ['agent', 'leads']
            if not agent_counts.empty:
                fig_agents = px.bar(agent_counts, x='agent', y='leads', title='Leads by agent')
                st.plotly_chart(fig_agents, use_container_width=True)
                charts_data['agent_breakdown'] = agent_counts
            
            # 3. Status breakdown
            status_counts = df_f['status'].value_counts().reset_index()
            status_counts.columns = ['status', 'count']
            if not status_counts.empty:
                fig_status = px.pie(status_counts, values='count', names='status', title='Lead status distribution')
                st.plotly_chart(fig_status, use_container_width=True)
                charts_data['status_breakdown'] = status_counts
            
            # 4. Sales Pipeline Funnel
            st.subheader('Sales Pipeline — Funnel Analysis')
            ordered_status = ['new','contacted','qualified','won','lost']
            counts_map = {s: int((df_f['status'] == s).sum()) for s in ordered_status}
            funnel_df = pd.DataFrame({'status': list(counts_map.keys()), 'count': list(counts_map.values())})
            fig_funnel = px.funnel(funnel_df, x='count', y='status', title='Lead Conversion Funnel')
            st.plotly_chart(fig_funnel, use_container_width=True)
            charts_data['sales_funnel'] = funnel_df
            
        except Exception as e:
            st.error(f'Error generating charts: {str(e)}')
        
        # Download all charts and data as ZIP
        st.markdown('---')
        st.subheader('📦 Download Complete Analytics Package')
        
        from datetime import datetime
        import zipfile
        
        # Generate filename with current day and date
        now = datetime.now()
        day_name = now.strftime('%A')  # Full day name (e.g., 'Monday')
        date_str = now.strftime('%Y-%m-%d')  # Date format (e.g., '2024-08-09')
        zip_filename = f"IQ_Stats_CRM_Analytics_{day_name}_{date_str}.zip"
        
        if st.button('📊 Generate & Download Analytics Package'):
            try:
                # Create ZIP file in memory
                zip_buffer = io.BytesIO()
                
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    # Add all chart data as Excel sheets
                    excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                        # Main data
                        df_all.to_excel(writer, sheet_name='All_Leads', index=False)
                        
                        # Chart data
                        for chart_name, chart_df in charts_data.items():
                            if isinstance(chart_df, pd.DataFrame) and not chart_df.empty:
                                chart_df.to_excel(writer, sheet_name=chart_name.replace('_', ' ').title()[:31], index=False)
                        
                        # Summary statistics
                        summary_data = {
                            'Metric': ['Total Leads', 'Unique Contacts', 'Top Agent', 'Active Salesmen', 'Conversion Rate'],
                            'Value': [
                                total_leads,
                                unique_contacts,
                                top_agent,
                                len(agent_counts) if 'agent_breakdown' in charts_data else 0,
                                f"{(df_f['status'] == 'won').sum() / len(df_f) * 100:.1f}%" if not df_f.empty else "0%"
                            ]
                        }
                        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Executive_Summary', index=False)
                    
                    # Add Excel file to ZIP
                    zip_file.writestr(f'CRM_Analytics_{date_str}.xlsx', excel_buffer.getvalue())
                    
                    # Add deals data if available
                    deals_df, _ = read_deals_df(limit=100000)
                    if not deals_df.empty:
                        deals_buffer = io.BytesIO()
                        with pd.ExcelWriter(deals_buffer, engine='openpyxl') as deals_writer:
                            deals_df.to_excel(deals_writer, sheet_name='All_Deals', index=False)
                            
                            # Deals summary
                            deals_summary = deals_df.groupby('uploaded_by').size().reset_index(name='deals_count')
                            deals_summary.to_excel(deals_writer, sheet_name='Deals_by_Agent', index=False)
                        
                        zip_file.writestr(f'Deals_Report_{date_str}.xlsx', deals_buffer.getvalue())
                    
                    # Add a README file
                    readme_content = f"""IQ Stats CRM Analytics Package
Generated on: {now.strftime('%A, %B %d, %Y at %H:%M')}
Generated for: CEO Dashboard

Contents:
- CRM_Analytics_{date_str}.xlsx: Complete leads analysis with all charts data
- Deals_Report_{date_str}.xlsx: Deals tracking and performance data

Chart Data Included:
- Daily leads trends
- Agent performance breakdown
- Lead status distribution
- Sales funnel analysis

Executive Summary:
- Total Leads: {total_leads}
- Unique Contacts: {unique_contacts}
- Top Performing Agent: {top_agent}

This package contains comprehensive analytics for strategic decision making.
"""
                    zip_file.writestr('README.txt', readme_content)
                
                # Offer download
                st.download_button(
                    label=f'📥 Download {zip_filename}',
                    data=zip_buffer.getvalue(),
                    file_name=zip_filename,
                    mime='application/zip',
                    help=f'Download complete analytics package for {day_name}, {date_str}'
                )
                
                st.success(f'✅ Analytics package ready! Contains all CTO dashboard charts and executive reports for {day_name}, {date_str}')
                
            except Exception as e:
                st.error(f'Error creating analytics package: {str(e)}')

        st.markdown('---')
        st.subheader('Download reports')
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            df_all.to_excel(writer, index=False, sheet_name='leads')
        st.download_button('Download Excel report', data=buf.getvalue(), file_name='crm_report.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        st.subheader('Summary by agent')
        summary = df_all.groupby('sales_agent').agg(leads=('id','count'), unique_contacts=('contact','nunique')).reset_index()
        st.dataframe(summary)

        # Done Deals Reports
        st.markdown('---')
        st.subheader('Done Deals — Reports & Export')
        deals_df, deals_total = read_deals_df(limit=100000, offset=0)
        if deals_df.empty:
            st.info('No deals submitted yet')
        else:
            deals_df['created_at'] = pd.to_datetime(deals_df['created_at'])
            c1, c2, c3 = st.columns(3)
            c1.metric('Total Deals', deals_total)
            c2.metric('Unique Customers', int(deals_df['customer_name'].nunique()))
            c3.metric('Salesmen Submitting', int(deals_df['uploaded_by'].nunique()))

            # Summaries
            by_salesman = deals_df.groupby('uploaded_by').size().reset_index(name='deals')
            by_day = deals_df.groupby(deals_df['created_at'].dt.date).size().reset_index(name='deals')
            st.write('Deals by Salesman')
            st.dataframe(by_salesman)
            st.write('Deals per Day')
            st.dataframe(by_day)

            # Export comprehensive Excel for advanced analysis
            deals_buf = io.BytesIO()
            with pd.ExcelWriter(deals_buf, engine='openpyxl') as writer:
                deals_df.to_excel(writer, index=False, sheet_name='deals')
                by_salesman.to_excel(writer, index=False, sheet_name='by_salesman')
                by_day.to_excel(writer, index=False, sheet_name='by_day')
            st.download_button('Download Done Deals Excel', data=deals_buf.getvalue(), file_name='done_deals_report.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            # Optional: Excel with embedded images for CEO (may be large)
            try:
                db = get_session()
                all_deals = db.query(Deal).order_by(Deal.created_at.desc()).all()
                excel_with_images = build_deals_excel_with_images(all_deals)
                if excel_with_images:
                    st.download_button('Download Done Deals (Excel with images)', data=excel_with_images, file_name='done_deals_with_images.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                zip_all = build_deals_images_zip(all_deals)
                st.download_button('Download all deal screenshots (ZIP)', data=zip_all, file_name='done_deals_screenshots.zip', mime='application/zip')
            finally:
                try:
                    db.close()
                except Exception:
                    pass

        # Login activity (visibility for CEO)
        st.markdown('---')
        st.subheader('User Login Activity (last 30 days)')
        try:
            with get_session() as db:
                thirty_days_ago = datetime.utcnow() - pd.Timedelta(days=30)
                logins = db.query(LoginEvent).filter(LoginEvent.logged_in_at >= thirty_days_ago).order_by(LoginEvent.logged_in_at.desc()).all()
                if logins:
                    df_login = pd.DataFrame([
                        {
                            'username': e.username,
                            'role': e.role,
                            'logged_in_at': e.logged_in_at
                        } for e in logins
                    ])
                    st.dataframe(df_login)
                    agg = df_login.groupby(['role','username']).size().reset_index(name='logins')
                    fig_login = px.bar(agg, x='username', y='logins', color='role', title='Logins per user (30 days)')
                    st.plotly_chart(fig_login, use_container_width=True)
                else:
                    st.info('No login events yet.')
        except Exception:
            st.info('Login activity not available yet.')

else:
    st.write('Role not supported in UI')

# Recent Activity Section - Only for CEO and CTO
if role in ['ceo', 'cto']:
    st.sidebar.markdown('---')
    st.sidebar.markdown('### 📊 Recent Activity')
    if st.sidebar.checkbox('Show recent activity', value=True):
        db = get_session()
        try:
            # Get recent activities
            acts = db.query(Activity).order_by(Activity.timestamp.desc()).limit(20).all()
            if acts:
                st.sidebar.markdown('**Latest System Activities:**')
                for i, a in enumerate(acts, 1):
                    time_str = a.timestamp.strftime('%m/%d %H:%M')
                    st.sidebar.markdown(f"**{i}.** `{time_str}` - **{a.actor}** {a.action} Lead #{a.lead_id}")
            else:
                st.sidebar.info('No recent activities found.')
            
            # Get recent login events
            st.sidebar.markdown('---')
            st.sidebar.markdown('**Recent Logins:**')
            recent_logins = db.query(LoginEvent).order_by(LoginEvent.logged_in_at.desc()).limit(10).all()
            if recent_logins:
                for i, login in enumerate(recent_logins, 1):
                    time_str = login.logged_in_at.strftime('%m/%d %H:%M')
                    st.sidebar.markdown(f"**{i}.** `{time_str}` - **{login.username}** ({login.role})")
            else:
                st.sidebar.info('No recent logins found.')
                
        except Exception as e:
            st.sidebar.error(f'Error loading activities: {str(e)}')
        finally:
            db.close()


