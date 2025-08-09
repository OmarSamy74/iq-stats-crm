"""
Full-featured Streamlit CRM — single-file example

Features:
- SQLAlchemy ORM models: Users, Leads, Activities (audit log), Comments
- Secure password hashing (passlib) and simple user management (admin can add users)
- Role-based permissions: salesman, head_of_sales, cto, ceo, admin
- Lead lifecycle: statuses (new, contacted, qualified, lost, won), assign to agent
- CRUD for leads: create, read, update, delete (with activity log)
- Upload XLSX/CSV and map columns; bulk-create leads
- Lead detail view with comments (notes) and history
- Filtering, searching, sorting, and pagination
- Dashboards for CTO (time-series, agent breakdown, status breakdown)
- Reporting and export (CSV / Excel) for CEO and admin
- Dockerfile and requirements notes in header

Run (dev):
1. pip install -r requirements.txt
   requirements.txt content below in comment
2. streamlit run streamlit_crm_full.py

Security note: This example uses an in-app user management for demo only.
For production use central auth (OAuth2 / SSO) and secure DB credentials.
"""

import streamlit as st
import pandas as pd
import sqlalchemy as sa
from sqlalchemy.orm import declarative_base, sessionmaker, relationship
from sqlalchemy import Column, Integer, String, DateTime, Text, ForeignKey, LargeBinary, func
from datetime import datetime, date
import io
import plotly.express as px
from passlib.context import CryptContext
import os
import math
import zipfile
import random
import numpy as np

# ----------------- Requirements (put in requirements.txt) -----------------
# streamlit
# pandas
# sqlalchemy
# openpyxl
# plotly
# passlib
# python-dateutil
# --------------------------------------------------------------------------

BASE_DIR = os.path.dirname(__file__) if '__file__' in globals() else '.'
DB_FILE = os.path.join(BASE_DIR, 'crm_full.db')
DATABASE_URL = f"sqlite:///{DB_FILE}"

# ----------------- DB setup -----------------
engine = sa.create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# ----------------- Models -----------------
class User(Base):
    __tablename__ = 'users'
    id = Column(Integer, primary_key=True)
    username = Column(String, unique=True, nullable=False)
    name = Column(String)
    password_hash = Column(String, nullable=False)
    role = Column(String, nullable=False)  # salesman, head_of_sales, cto, ceo, admin
    created_at = Column(DateTime, default=datetime.utcnow)

    # relationship to leads via uploaded_by_id
    uploads = relationship('Lead', back_populates='uploader', cascade="all, delete-orphan")
    # relationship to deals via uploaded_by_id
    deals = relationship('Deal', back_populates='uploader', cascade="all, delete-orphan")

    def verify_password(self, password):
        return pwd_context.verify(password, self.password_hash)

    @classmethod
    def hash_password(cls, password):
        return pwd_context.hash(password)

class Lead(Base):
    __tablename__ = 'leads'
    id = Column(Integer, primary_key=True)
    number = Column(String, index=True)
    name = Column(String)
    sales_agent = Column(String, index=True)
    contact = Column(String)
    case_desc = Column(Text)
    feedback = Column(Text)
    status = Column(String, default='new')  # new, contacted, qualified, lost, won
    assigned_to = Column(String, nullable=True)

    # Keep a readable username string and also a foreign-key to users.id
    uploaded_by = Column(String)                     # username (readable)
    uploaded_by_id = Column(Integer, ForeignKey('users.id'))  # FK
    uploaded_at = Column(DateTime, default=datetime.utcnow)

    # Archiving fields for CTO
    is_archived = Column(String, default='no')  # yes, no
    archived_by = Column(String, nullable=True)  # username who archived
    archived_at = Column(DateTime, nullable=True)
    archive_reason = Column(Text, nullable=True)  # reason for archiving
    archive_date = Column(DateTime, nullable=True)  # date when lead should be archived

    # relationships
    uploader = relationship('User', back_populates='uploads')
    activities = relationship('Activity', back_populates='lead', cascade="all, delete-orphan")
    comments = relationship('Comment', back_populates='lead', cascade="all, delete-orphan")

class Activity(Base):
    __tablename__ = 'activities'
    id = Column(Integer, primary_key=True)
    lead_id = Column(Integer, ForeignKey('leads.id'))
    actor = Column(String)
    action = Column(String)
    timestamp = Column(DateTime, default=datetime.utcnow)
    detail = Column(Text)
    lead = relationship('Lead', back_populates='activities')

class Comment(Base):
    __tablename__ = 'comments'
    id = Column(Integer, primary_key=True)
    lead_id = Column(Integer, ForeignKey('leads.id'))
    author = Column(String)
    text = Column(Text)
    created_at = Column(DateTime, default=datetime.utcnow)
    lead = relationship('Lead', back_populates='comments')

class Deal(Base):
    __tablename__ = 'deals'
    id = Column(Integer, primary_key=True)
    customer_name = Column(String, nullable=False)
    phone = Column(String, nullable=False)
    payment_screenshot = Column(LargeBinary, nullable=False)
    uploaded_by = Column(String)                     # username (readable)
    uploaded_by_id = Column(Integer, ForeignKey('users.id'))  # FK
    created_at = Column(DateTime, default=datetime.utcnow)
    uploader = relationship('User', back_populates='deals')

class LoginEvent(Base):
    __tablename__ = 'login_events'
    id = Column(Integer, primary_key=True)
    username = Column(String, nullable=False)
    role = Column(String, nullable=False)
    logged_in_at = Column(DateTime, default=datetime.utcnow, index=True)

class Setting(Base):
    __tablename__ = 'settings'
    key = Column(String, primary_key=True)
    value = Column(String)

Base.metadata.create_all(bind=engine)

# Ensure DB schema matches current models (adds new columns if missing)
def ensure_schema():
    """Lightweight auto-migration for SQLite.
    Adds missing columns on existing tables to avoid OperationalError when models evolve.
    """
    with engine.begin() as conn:
        # Leads table columns
        try:
            rows = conn.exec_driver_sql("PRAGMA table_info('leads')").fetchall()
            existing_columns = {row[1] for row in rows}
        except Exception:
            existing_columns = set()

        # Add columns if they don't exist
        if 'uploaded_by' not in existing_columns:
            conn.exec_driver_sql("ALTER TABLE leads ADD COLUMN uploaded_by VARCHAR")
        if 'uploaded_by_id' not in existing_columns:
            conn.exec_driver_sql("ALTER TABLE leads ADD COLUMN uploaded_by_id INTEGER")
        if 'uploaded_at' not in existing_columns:
            conn.exec_driver_sql("ALTER TABLE leads ADD COLUMN uploaded_at DATETIME")
        if 'assigned_to' not in existing_columns:
            conn.exec_driver_sql("ALTER TABLE leads ADD COLUMN assigned_to VARCHAR")
        if 'status' not in existing_columns:
            conn.exec_driver_sql("ALTER TABLE leads ADD COLUMN status VARCHAR DEFAULT 'new'")
        
        # Add archiving columns if they don't exist
        if 'is_archived' not in existing_columns:
            conn.exec_driver_sql("ALTER TABLE leads ADD COLUMN is_archived VARCHAR DEFAULT 'no'")
        if 'archived_by' not in existing_columns:
            conn.exec_driver_sql("ALTER TABLE leads ADD COLUMN archived_by VARCHAR")
        if 'archived_at' not in existing_columns:
            conn.exec_driver_sql("ALTER TABLE leads ADD COLUMN archived_at DATETIME")
        if 'archive_reason' not in existing_columns:
            conn.exec_driver_sql("ALTER TABLE leads ADD COLUMN archive_reason TEXT")
        if 'archive_date' not in existing_columns:
            conn.exec_driver_sql("ALTER TABLE leads ADD COLUMN archive_date DATETIME")

        # Optional: ensure indexes exist (SQLite auto-creates pk index; skip others for simplicity)

        # Deals table columns (create table if not exists then add new columns if missing)
        conn.exec_driver_sql("""
        CREATE TABLE IF NOT EXISTS deals (
            id INTEGER PRIMARY KEY,
            customer_name VARCHAR NOT NULL,
            phone VARCHAR NOT NULL,
            payment_screenshot BLOB NOT NULL,
            uploaded_by VARCHAR,
            uploaded_by_id INTEGER,
            created_at DATETIME
        )
        """)
        try:
            rows_deals = conn.exec_driver_sql("PRAGMA table_info('deals')").fetchall()
            existing_deals_cols = {row[1] for row in rows_deals}
        except Exception:
            existing_deals_cols = set()
        if 'uploaded_by' not in existing_deals_cols:
            conn.exec_driver_sql("ALTER TABLE deals ADD COLUMN uploaded_by VARCHAR")
        if 'uploaded_by_id' not in existing_deals_cols:
            conn.exec_driver_sql("ALTER TABLE deals ADD COLUMN uploaded_by_id INTEGER")
        if 'created_at' not in existing_deals_cols:
            conn.exec_driver_sql("ALTER TABLE deals ADD COLUMN created_at DATETIME")

        # Login events table
        conn.exec_driver_sql("""
        CREATE TABLE IF NOT EXISTS login_events (
            id INTEGER PRIMARY KEY,
            username VARCHAR NOT NULL,
            role VARCHAR NOT NULL,
            logged_in_at DATETIME
        )
        """)

        # Settings table
        conn.exec_driver_sql("""
        CREATE TABLE IF NOT EXISTS settings (
            key VARCHAR PRIMARY KEY,
            value VARCHAR
        )
        """)

ensure_schema()

# ----------------- Utility helpers -----------------

def get_session():
    return SessionLocal()

def get_user_by_username(db, username):
    return db.query(User).filter(User.username == username).first()

def create_user(db, username, password, role='salesman', name=None):
    if get_user_by_username(db, username):
        return None
    user = User(username=username, password_hash=User.hash_password(password), role=role, name=name or username)
    db.add(user)
    db.commit()
    db.refresh(user)
    return user

def update_or_create_user(db, username, password, role='salesman', name=None):
    """Update existing user or create new one"""
    user = get_user_by_username(db, username)
    if user:
        # Update existing user
        user.password_hash = User.hash_password(password)
        user.name = name or user.name
        user.role = role
        db.commit()
        db.refresh(user)
        return user
    else:
        # Create new user
        return create_user(db, username, password, role, name)

def ensure_demo_users():
    db = get_session()
    try:
        # Management roles
        update_or_create_user(db, 'head', 'IQstats@iq-2024', role='head_of_sales', name='Mohamed Akmal')
        update_or_create_user(db, 'cto', 'IQstats@iq-2025', role='cto', name='Omar Samy')
        update_or_create_user(db, 'ceo', 'IQstats@iq-2026', role='ceo', name='ENG Ahmed Essam')
        # Sales team
        update_or_create_user(db, 'toqa', 'IQstats@iq-2027', role='salesman', name='Toqa Amin')
        update_or_create_user(db, 'mahmoud', 'IQstats@iq-2028', role='salesman', name='Mahmoud Fathalla')
        update_or_create_user(db, 'mazen', 'IQstats@iq-2029', role='salesman', name='Mazen Ashraf')
        update_or_create_user(db, 'ahmed_malek', 'IQstats@iq-2030', role='salesman', name='Ahmed Malek')
        update_or_create_user(db, 'youssry', 'IQstats@iq-2031', role='salesman', name='Youssry Hassan')
    finally:
        db.close()

ensure_demo_users()

# ----------------- Activity logger -----------------

def log_activity(db, lead_id, actor, action, detail=None):
    act = Activity(lead_id=lead_id, actor=actor, action=action, detail=detail)
    db.add(act)
    db.commit()

# ----------------- Archiving helpers -----------------

def archive_lead(db, lead_id, archived_by, reason=None, archive_date=None):
    """Archive a lead by CTO"""
    lead = db.query(Lead).get(lead_id)
    if lead:
        lead.is_archived = 'yes'
        lead.archived_by = archived_by
        lead.archived_at = datetime.utcnow()
        lead.archive_reason = reason
        lead.archive_date = archive_date
        db.add(lead)
        db.commit()
        log_activity(db, lead_id, archived_by, 'archive', detail=f'Archived: {reason}')
        return True
    return False

def unarchive_lead(db, lead_id, unarchived_by):
    """Unarchive a lead by CTO"""
    lead = db.query(Lead).get(lead_id)
    if lead:
        lead.is_archived = 'no'
        lead.archived_by = None
        lead.archived_at = None
        lead.archive_reason = None
        lead.archive_date = None
        db.add(lead)
        db.commit()
        log_activity(db, lead_id, unarchived_by, 'unarchive', detail='Lead unarchived')
        return True
    return False

def delete_lead_from_db(db, lead_id, deleted_by, reason=None):
    """Permanently delete a lead from database"""
    lead = db.query(Lead).get(lead_id)
    if lead:
        # Log the deletion before removing
        log_activity(db, lead_id, deleted_by, 'delete', detail=f'Lead permanently deleted: {reason}')
        
        # Delete related activities and comments first (due to foreign key constraints)
        db.query(Activity).filter(Activity.lead_id == lead_id).delete()
        db.query(Comment).filter(Comment.lead_id == lead_id).delete()
        
        # Delete the lead
        db.delete(lead)
        db.commit()
        return True
    return False

def bulk_delete_leads_from_db(db, lead_ids, deleted_by, reason=None):
    """Bulk delete multiple leads from database"""
    deleted_count = 0
    for lead_id in lead_ids:
        if delete_lead_from_db(db, lead_id, deleted_by, reason):
            deleted_count += 1
    return deleted_count

def bulk_archive_leads(db, lead_ids, archived_by, reason=None, archive_date=None):
    """Bulk archive multiple leads"""
    archived_count = 0
    for lead_id in lead_ids:
        if archive_lead(db, lead_id, archived_by, reason, archive_date):
            archived_count += 1
    return archived_count

def get_archived_leads_by_date(db, date_filter=None):
    """Get archived leads filtered by date"""
    q = db.query(Lead).filter(Lead.is_archived == 'yes')
    if date_filter:
        if isinstance(date_filter, str):
            # Filter by specific date
            q = q.filter(func.date(Lead.archived_at) == date_filter)
        elif isinstance(date_filter, (list, tuple)) and len(date_filter) == 2:
            # Filter by date range
            start_date, end_date = date_filter
            q = q.filter(
                Lead.archived_at >= start_date,
                Lead.archived_at <= end_date
            )
    return q.order_by(Lead.archived_at.desc()).all()

def generate_archived_leads_analytics(df):
    """Generate analytics data for archived leads"""
    charts_data = {}
    
    if df.empty:
        return charts_data
    
    # Convert date columns
    df['archive_date'] = pd.to_datetime(df['Archive Date'])
    df['upload_date'] = pd.to_datetime(df['Original Upload Date'])
    
    # 1. Daily archive trends
    daily_archives = df.groupby(df['archive_date'].dt.date).size().reset_index(name='count')
    daily_archives.columns = ['date', 'count']
    charts_data['daily_archives'] = daily_archives
    
    # 2. Archive reasons breakdown
    reason_counts = df['Archive Reason'].value_counts().reset_index()
    reason_counts.columns = ['reason', 'count']
    charts_data['archive_reasons'] = reason_counts
    
    # 3. Archived by breakdown
    archiver_counts = df['Archived By'].value_counts().reset_index()
    archiver_counts.columns = ['archiver', 'count']
    charts_data['archivers'] = archiver_counts
    
    # 4. Original status distribution
    status_counts = df['Status'].value_counts().reset_index()
    status_counts.columns = ['status', 'count']
    charts_data['original_status'] = status_counts
    
    # 5. Sales agent breakdown (original)
    agent_counts = df['Sales Agent'].value_counts().reset_index()
    agent_counts.columns = ['sales_agent', 'count']
    charts_data['original_agents'] = agent_counts
    
    # 6. Contact methods breakdown
    contact_counts = df['Contact Method'].value_counts().reset_index()
    contact_counts.columns = ['contact', 'count']
    charts_data['contact_methods'] = contact_counts
    
    # 7. Archive timing analysis (days between upload and archive)
    df['days_to_archive'] = (df['archive_date'] - df['upload_date']).dt.days
    timing_analysis = df['days_to_archive'].describe().reset_index()
    timing_analysis.columns = ['metric', 'value']
    charts_data['archive_timing'] = timing_analysis
    
    # 8. Monthly archive trends
    monthly_archives = df.groupby(df['archive_date'].dt.to_period('M')).size().reset_index(name='count')
    monthly_archives['month'] = monthly_archives['archive_date'].astype(str)
    monthly_archives = monthly_archives[['month', 'count']]
    charts_data['monthly_archives'] = monthly_archives
    
    return charts_data

def export_archived_leads_report(db, date_range=None, format='excel', include_graphs=False):
    """Export archived leads report with date information and optional graphs"""
    q = db.query(Lead).filter(Lead.is_archived == 'yes')
    
    if date_range:
        if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
            start_date, end_date = date_range
            q = q.filter(
                Lead.archived_at >= start_date,
                Lead.archived_at <= end_date
            )
    
    archived_leads = q.order_by(Lead.archived_at.desc()).all()
    
    # Create DataFrame with all relevant information
    data = []
    for lead in archived_leads:
        data.append({
            'Lead ID': lead.id,
            'Lead Number': lead.number,
            'Customer Name': lead.name,
            'Sales Agent': lead.sales_agent,
            'Contact Method': lead.contact,
            'Case Description': lead.case_desc,
            'Feedback': lead.feedback,
            'Status': lead.status,
            'Assigned To': lead.assigned_to,
            'Original Upload Date': lead.uploaded_at.strftime('%Y-%m-%d %H:%M') if lead.uploaded_at else '',
            'Archived By': lead.archived_by,
            'Archive Date': lead.archived_at.strftime('%Y-%m-%d %H:%M') if lead.archived_at else '',
            'Archive Reason': lead.archive_reason,
            'Scheduled Archive Date': lead.archive_date.strftime('%Y-%m-%d %H:%M') if lead.archive_date else ''
        })
    
    df = pd.DataFrame(data)
    
    if format == 'excel':
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Archived_Leads', index=False)
            
            # Add analytics sheets if graphs are requested
            if include_graphs and not df.empty:
                charts_data = generate_archived_leads_analytics(df)
                
                # Add analytics sheets
                for chart_name, chart_df in charts_data.items():
                    if isinstance(chart_df, pd.DataFrame) and not chart_df.empty:
                        sheet_name = chart_name.replace('_', ' ').title()[:31]
                        chart_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Add summary sheet
            summary_data = {
                'Metric': ['Total Archived Leads', 'Date Range', 'Generated By', 'Generated On'],
                'Value': [
                    len(archived_leads),
                    f"{date_range[0] if date_range else 'All'} to {date_range[1] if date_range else 'All'}",
                    'CTO Dashboard',
                    datetime.now().strftime('%Y-%m-%d %H:%M')
                ]
            }
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
        
        return buffer.getvalue()
    else:
        return df.to_csv(index=False)

# ----------------- Streamlit UI -----------------

st.set_page_config(page_title='IQ Stats CRM — Full', layout='wide')

# --- Authentication ---
if 'user' not in st.session_state:
    st.session_state['user'] = None

if st.session_state['user'] is None:
    st.title('IQ Stats CRM — Login')
    with st.form('login_form'):
        username = st.text_input('Username')
        password = st.text_input('Password', type='password')
        submitted = st.form_submit_button('Login')
        if submitted:
            db = get_session()
            user = get_user_by_username(db, username)
            if user and user.verify_password(password):
                st.session_state['user'] = user.username
                st.session_state['role'] = user.role
                # log login event
                try:
                    evt = LoginEvent(username=user.username, role=user.role, logged_in_at=datetime.utcnow())
                    db.add(evt)
                    db.commit()
                except Exception:
                    pass
                st.success(f'Welcome, {user.name} ({user.role})')
                st.rerun()
            else:
                st.error('Invalid username or password')


    st.stop()

# Load current user
current_user = None
with get_session() as db:
    current_user = get_user_by_username(db, st.session_state['user'])

if current_user is None:
    st.error('User not found. Please login again.')
    st.session_state.clear()
    st.stop()

role = current_user.role

# Top bar
st.sidebar.title('IQ Stats CRM')
st.sidebar.write(f'Logged in as **{current_user.name}** — *{role}*')
if st.sidebar.button('Logout'):
    for k in list(st.session_state.keys()):
        del st.session_state[k]
    st.rerun()

# Admin: manage users
if role == 'admin':
    st.header('Admin — System Management')
    
    # User Management Section
    st.subheader('👥 User Management')
    db = get_session()
    users = db.query(User).all()
    
    # Display current users
    st.write('**Current Users:**')
    cols = st.columns([1,2,2,1])
    cols[0].write('ID')
    cols[1].write('Username')
    cols[2].write('Name')
    cols[3].write('Role')
    for u in users:
        c0, c1, c2, c3 = st.columns([1,2,2,1])
        c0.write(u.id)
        c1.write(u.username)
        c2.write(u.name)
        c3.write(u.role)

    # Create new user form
    st.markdown('---')
    st.subheader('➕ Create New User')
    with st.form('create_user'):
        col1, col2 = st.columns(2)
        with col1:
            new_username = st.text_input('Username', placeholder='e.g., john_doe')
            new_name = st.text_input('Full Name', placeholder='e.g., John Doe')
            new_role = st.selectbox('Role', ['salesman','head_of_sales','cto','ceo','admin'])
        with col2:
            new_pass = st.text_input('Password', type='password', placeholder='Enter password')
            confirm_pass = st.text_input('Confirm Password', type='password', placeholder='Confirm password')
            st.write('**Password Format:** IQstats@iq-XXXX (recommended)')
        
        create = st.form_submit_button('✅ Create User')
        if create:
            if new_username and new_pass and new_name:
                if new_pass == confirm_pass:
                    res = create_user(db, new_username, new_pass, role=new_role, name=new_name)
                    if res:
                        st.success(f'✅ User "{new_username}" created successfully!')
                        st.rerun()
                    else:
                        st.error('❌ Username already exists')
                else:
                    st.error('❌ Passwords do not match')
            else:
                st.error('❌ Please fill all required fields')
    
    # User Actions
    st.markdown('---')
    st.subheader('🔧 User Actions')
    col1, col2 = st.columns(2)
    
    with col1:
        st.write('**Delete User:**')
        if users:
            user_to_delete = st.selectbox('Select user to delete', 
                                        options=[f"{u.username} ({u.name})" for u in users if u.username != 'admin'],
                                        key='delete_user')
            if st.button('🗑️ Delete User', type='secondary'):
                if user_to_delete:
                    username = user_to_delete.split(' (')[0]
                    user = db.query(User).filter(User.username == username).first()
                    if user:
                        db.delete(user)
                        db.commit()
                        st.success(f'✅ User "{username}" deleted successfully!')
                        st.rerun()
    
    with col2:
        st.write('**Update User Password:**')
        if users:
            user_to_update = st.selectbox('Select user to update', 
                                        options=[f"{u.username} ({u.name})" for u in users],
                                        key='update_user')
            new_password = st.text_input('New Password', type='password', key='new_pwd')
            if st.button('🔐 Update Password', type='secondary'):
                if user_to_update and new_password:
                    username = user_to_update.split(' (')[0]
                    user = db.query(User).filter(User.username == username).first()
                    if user:
                        user.password_hash = User.hash_password(new_password)
                        db.commit()
                        st.success(f'✅ Password updated for "{username}"!')
                        st.rerun()
    
    db.close()
    
    # System Maintenance Section
    st.markdown('---')
    st.subheader('⚙️ System Maintenance')
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write('**Database Management:**')
        if st.button('🔄 Reset Database (Drop & Recreate)', type='primary'):
            try:
                # Close sessions and remove DB file
                try:
                    del st.session_state['user']
                except Exception:
                    pass
                if os.path.exists(DB_FILE):
                    os.remove(DB_FILE)
                # Recreate
                Base.metadata.create_all(bind=engine)
                ensure_schema()
                ensure_demo_users()
                st.success('✅ Database reset successfully! Please reload the app.')
                st.rerun()
            except Exception as e:
                st.error(f'❌ Failed to reset database: {e}')
        
        if st.button('📊 Export All Data', type='secondary'):
            try:
                from datetime import datetime
                import zipfile
                
                now = datetime.now()
                date_str = now.strftime('%Y-%m-%d_%H-%M')
                zip_filename = f"Admin_Data_Export_{date_str}.zip"
                
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    # Export users
                    users_df = pd.DataFrame([{
                        'id': u.id, 'username': u.username, 'name': u.name, 
                        'role': u.role, 'created_at': u.created_at
                    } for u in users])
                    users_buffer = io.BytesIO()
                    with pd.ExcelWriter(users_buffer, engine='openpyxl') as writer:
                        users_df.to_excel(writer, sheet_name='Users', index=False)
                    zip_file.writestr('Users.xlsx', users_buffer.getvalue())
                    
                    # Export leads
                    leads_df, _ = read_leads_df(limit=100000)
                    if not leads_df.empty:
                        leads_buffer = io.BytesIO()
                        with pd.ExcelWriter(leads_buffer, engine='openpyxl') as writer:
                            leads_df.to_excel(writer, sheet_name='Leads', index=False)
                        zip_file.writestr('Leads.xlsx', leads_buffer.getvalue())
                    
                    # Export deals
                    deals_df, _ = read_deals_df(limit=100000)
                    if not deals_df.empty:
                        deals_buffer = io.BytesIO()
                        with pd.ExcelWriter(deals_buffer, engine='openpyxl') as writer:
                            deals_df.to_excel(writer, sheet_name='Deals', index=False)
                        zip_file.writestr('Deals.xlsx', deals_buffer.getvalue())
                    
                    # README
                    readme_content = f"""Admin Data Export
Generated on: {now.strftime('%A, %B %d, %Y at %H:%M')}
Generated by: Admin

Contents:
- Users.xlsx: All user accounts and roles
- Leads.xlsx: All leads data
- Deals.xlsx: All deals data

This is a complete system backup for administrative purposes.
"""
                    zip_file.writestr('README.txt', readme_content)
                
                st.download_button(
                    label=f'📥 Download {zip_filename}',
                    data=zip_buffer.getvalue(),
                    file_name=zip_filename,
                    mime='application/zip'
                )
                st.success('✅ Data export ready!')
                
            except Exception as e:
                st.error(f'❌ Error creating data export: {str(e)}')
    
    with col2:
        st.write('**System Information:**')
        st.info(f"""
        **Database File:** {DB_FILE}
        **Total Users:** {len(users)}
        **Database Size:** {os.path.getsize(DB_FILE) / 1024:.1f} KB
        **Last Modified:** {datetime.fromtimestamp(os.path.getmtime(DB_FILE)).strftime('%Y-%m-%d %H:%M')}
        """)
        
        st.write('**Quick Actions:**')
        if st.button('🔄 Refresh Page', type='secondary'):
            st.rerun()
        
        if st.button('📋 Copy System Info', type='secondary'):
            st.code(f"""
Database: {DB_FILE}
Users: {len(users)}
Size: {os.path.getsize(DB_FILE) / 1024:.1f} KB
Modified: {datetime.fromtimestamp(os.path.getmtime(DB_FILE)).strftime('%Y-%m-%d %H:%M')}
            """)
    
    st.stop()

# Common functions for leads

def read_leads_df(filters=None, search=None, order_by='uploaded_at', desc=True, limit=100, offset=0, include_archived=False):
    db = get_session()
    q = db.query(Lead)
    
    # Default filter: exclude archived leads unless specifically requested
    if not include_archived:
        q = q.filter(sa.or_(Lead.is_archived.is_(None), Lead.is_archived == 'no'))
    
    if filters:
        for k, v in filters.items():
            if k == 'sales_agent' and v != 'All':
                q = q.filter(Lead.sales_agent == v)
            if k == 'status' and v != 'All':
                q = q.filter(Lead.status == v)
            if k == 'is_archived' and v != 'All':
                q = q.filter(Lead.is_archived == v)
            if k == 'archived_by' and v != 'All':
                q = q.filter(Lead.archived_by == v)
    
    if search:
        like = f"%{search}%"
        q = q.filter(sa.or_(Lead.name.ilike(like), Lead.number.ilike(like), Lead.contact.ilike(like)))
    
    total = q.count()
    if order_by:
        col = getattr(Lead, order_by)
        q = q.order_by(col.desc() if desc else col)
    q = q.offset(offset).limit(limit)
    df = pd.read_sql(q.statement, q.session.bind)
    db.close()
    return df, total

def read_deals_df(filters=None, search=None, order_by='created_at', desc=True, limit=1000, offset=0):
    db = get_session()
    q = db.query(Deal)
    if filters:
        for k, v in filters.items():
            if k == 'uploaded_by' and v != 'All':
                q = q.filter(Deal.uploaded_by == v)
    if search:
        like = f"%{search}%"
        q = q.filter(sa.or_(Deal.customer_name.ilike(like), Deal.phone.ilike(like)))
    total = q.count()
    if order_by:
        col = getattr(Deal, order_by)
        q = q.order_by(col.desc() if desc else col)
    q = q.offset(offset).limit(limit)
    # Exclude large binary column when reading to DataFrame
    cols = [Deal.id, Deal.customer_name, Deal.phone, Deal.uploaded_by, Deal.uploaded_by_id, Deal.created_at]
    q = db.query(*cols).filter(Deal.id.in_(db.query(Deal.id).subquery())) if True else q
    df = pd.read_sql(q.statement, q.session.bind)
    db.close()
    return df, total

# ----------------- Export helpers -----------------
def build_deals_excel_with_images(deals):
    """Create an Excel workbook with deals and embedded screenshots if possible.
    Returns bytes, or None if image embedding is unavailable.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
    except Exception:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = 'deals'
    headers = ['id', 'customer_name', 'phone', 'uploaded_by', 'created_at', 'screenshot']
    ws.append(headers)

    # Wider column for screenshot
    try:
        ws.column_dimensions['F'].width = 40
    except Exception:
        pass

    for idx, d in enumerate(deals, start=2):
        ws.cell(row=idx, column=1, value=d.id)
        ws.cell(row=idx, column=2, value=d.customer_name)
        ws.cell(row=idx, column=3, value=d.phone)
        ws.cell(row=idx, column=4, value=d.uploaded_by)
        ws.cell(row=idx, column=5, value=(d.created_at.strftime('%Y-%m-%d %H:%M') if d.created_at else None))
        # Try embedding image
        if getattr(d, 'payment_screenshot', None):
            try:
                img_stream = io.BytesIO(d.payment_screenshot)
                img = XLImage(img_stream)
                cell = f'F{idx}'
                ws.add_image(img, cell)
                # Increase row height for visibility
                try:
                    ws.row_dimensions[idx].height = 120
                except Exception:
                    pass
            except Exception:
                ws.cell(row=idx, column=6, value='[image available]')
        else:
            ws.cell(row=idx, column=6, value='')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def build_deals_images_zip(deals):
    """Create a zip containing deals' screenshots. Returns bytes."""
    def _detect_ext(content: bytes) -> str:
        # PNG
        if content[:8] == b'\x89PNG\r\n\x1a\n':
            return 'png'
        # JPEG
        if content[:3] == b'\xff\xd8\xff':
            return 'jpg'
        # WEBP (RIFF....WEBP)
        if len(content) >= 12 and content[:4] == b'RIFF' and content[8:12] == b'WEBP':
            return 'webp'
        return 'png'
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        for d in deals:
            content = getattr(d, 'payment_screenshot', None)
            if not content:
                continue
            ext = _detect_ext(content)
            zf.writestr(f'deal_{d.id}_payment.{ext}', content)
    return buf.getvalue()

def generate_analytics_graphs(df, charts_data, date_str, title_prefix="Analytics"):
    """Generate comprehensive analytics graphs and return as bytes for zip inclusion."""
    try:
        import plotly.graph_objects as go
        import plotly.express as px
        from plotly.subplots import make_subplots
        import matplotlib.pyplot as plt
        import seaborn as sns
        from matplotlib.backends.backend_pdf import PdfPages
        # Removed local 'import io' to avoid shadowing the module-level import
        
        # Create a PDF with multiple graphs
        pdf_buffer = io.BytesIO()
        
        with PdfPages(pdf_buffer) as pdf:
            # Set style
            plt.style.use('seaborn-v0_8')
            fig_width, fig_height = 12, 8
            
            # 1. Daily Leads Trend
            if 'daily_leads' in charts_data and not charts_data['daily_leads'].empty:
                fig, ax = plt.subplots(figsize=(fig_width, fig_height))
                daily_data = charts_data['daily_leads']
                ax.plot(daily_data['date'], daily_data['count'], marker='o', linewidth=2, markersize=6)
                ax.set_title(f'{title_prefix} - Daily Leads Trend', fontsize=16, fontweight='bold')
                ax.set_xlabel('Date', fontsize=12)
                ax.set_ylabel('Number of Leads', fontsize=12)
                ax.grid(True, alpha=0.3)
                plt.xticks(rotation=45)
                plt.tight_layout()
                pdf.savefig(fig)
                plt.close()
            
            # 2. Agent Performance Breakdown
            if 'agent_breakdown' in charts_data and not charts_data['agent_breakdown'].empty:
                fig, ax = plt.subplots(figsize=(fig_width, fig_height))
                agent_data = charts_data['agent_breakdown']
                bars = ax.bar(agent_data['sales_agent'], agent_data['count'], 
                             color=plt.cm.Set3(np.linspace(0, 1, len(agent_data))))
                ax.set_title(f'{title_prefix} - Agent Performance Breakdown', fontsize=16, fontweight='bold')
                ax.set_xlabel('Sales Agent', fontsize=12)
                ax.set_ylabel('Number of Leads', fontsize=12)
                ax.grid(True, alpha=0.3, axis='y')
                
                # Add value labels on bars
                for bar in bars:
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width()/2., height + 0.01,
                           f'{int(height)}', ha='center', va='bottom', fontweight='bold')
                
                plt.xticks(rotation=45)
                plt.tight_layout()
                pdf.savefig(fig)
                plt.close()
            
            # 3. Lead Status Distribution
            if 'status_breakdown' in charts_data and not charts_data['status_breakdown'].empty:
                fig, ax = plt.subplots(figsize=(fig_width, fig_height))
                status_data = charts_data['status_breakdown']
                colors = ['#2E8B57', '#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4']
                wedges, texts, autotexts = ax.pie(status_data['count'], labels=status_data['status'], 
                                                 autopct='%1.1f%%', colors=colors, startangle=90)
                ax.set_title(f'{title_prefix} - Lead Status Distribution', fontsize=16, fontweight='bold')
                
                # Make percentage labels bold
                for autotext in autotexts:
                    autotext.set_color('white')
                    autotext.set_fontweight('bold')
                
                plt.tight_layout()
                pdf.savefig(fig)
                plt.close()
            
            # 4. Sales Funnel
            if 'sales_funnel' in charts_data and not charts_data['sales_funnel'].empty:
                fig, ax = plt.subplots(figsize=(fig_width, fig_height))
                funnel_data = charts_data['sales_funnel']
                stages = funnel_data['stage']
                counts = funnel_data['count']
                
                # Create funnel chart
                y_pos = np.arange(len(stages))
                bars = ax.barh(y_pos, counts, color=plt.cm.viridis(np.linspace(0, 1, len(stages))))
                ax.set_yticks(y_pos)
                ax.set_yticklabels(stages)
                ax.set_xlabel('Number of Leads', fontsize=12)
                ax.set_title(f'{title_prefix} - Sales Funnel', fontsize=16, fontweight='bold')
                ax.grid(True, alpha=0.3, axis='x')
                
                # Add value labels
                for i, (bar, count) in enumerate(zip(bars, counts)):
                    ax.text(bar.get_width() + 0.01, bar.get_y() + bar.get_height()/2,
                           f'{int(count)}', ha='left', va='center', fontweight='bold')
                
                plt.tight_layout()
                pdf.savefig(fig)
                plt.close()
            
            # 5. Contact Methods Analysis
            if 'contact_methods' in charts_data and not charts_data['contact_methods'].empty:
                fig, ax = plt.subplots(figsize=(fig_width, fig_height))
                contact_data = charts_data['contact_methods']
                bars = ax.bar(contact_data['contact'], contact_data['count'], 
                             color=plt.cm.Pastel1(np.linspace(0, 1, len(contact_data))))
                ax.set_title(f'{title_prefix} - Contact Methods Analysis', fontsize=16, fontweight='bold')
                ax.set_xlabel('Contact Method', fontsize=12)
                ax.set_ylabel('Number of Leads', fontsize=12)
                ax.grid(True, alpha=0.3, axis='y')
                
                # Add value labels
                for bar in bars:
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width()/2., height + 0.01,
                           f'{int(height)}', ha='center', va='bottom', fontweight='bold')
                
                plt.xticks(rotation=45)
                plt.tight_layout()
                pdf.savefig(fig)
                plt.close()
            
            # 6. Activity Heatmap
            if 'activity_heatmap' in charts_data and not charts_data['activity_heatmap'].empty:
                fig, ax = plt.subplots(figsize=(fig_width, fig_height))
                heat_data = charts_data['activity_heatmap']
                
                # Create heatmap
                sns.heatmap(heat_data.pivot_table(index='hour', columns='day_of_week', values='count', fill_value=0),
                           annot=True, fmt='d', cmap='YlOrRd', ax=ax)
                ax.set_title(f'{title_prefix} - Activity Heatmap', fontsize=16, fontweight='bold')
                ax.set_xlabel('Day of Week', fontsize=12)
                ax.set_ylabel('Hour of Day', fontsize=12)
                
                plt.tight_layout()
                pdf.savefig(fig)
                plt.close()
            
            # 7. Trends Analysis (if available)
            if 'trends' in charts_data and not charts_data['trends'].empty:
                fig, ax = plt.subplots(figsize=(fig_width, fig_height))
                trends_data = charts_data['trends']
                ax.plot(trends_data['date'], trends_data['count'], label='Daily Count', marker='o')
                if 'rolling_avg' in trends_data.columns:
                    ax.plot(trends_data['date'], trends_data['rolling_avg'], 
                           label='7-Day Rolling Average', linewidth=2, color='red')
                ax.set_title(f'{title_prefix} - Trends Analysis', fontsize=16, fontweight='bold')
                ax.set_xlabel('Date', fontsize=12)
                ax.set_ylabel('Number of Leads', fontsize=12)
                ax.legend()
                ax.grid(True, alpha=0.3)
                plt.xticks(rotation=45)
                plt.tight_layout()
                pdf.savefig(fig)
                plt.close()
            
            # 8. Summary Dashboard (Combined view)
            fig = plt.figure(figsize=(16, 12))
            fig.suptitle(f'{title_prefix} - Summary Dashboard', fontsize=20, fontweight='bold')
            
            # Create subplots
            gs = fig.add_gridspec(3, 3, hspace=0.3, wspace=0.3)
            
            # Key metrics
            ax1 = fig.add_subplot(gs[0, :])
            ax1.axis('off')
            summary_text = f"""
            {title_prefix} Summary Report - Generated on {date_str}
            
            Total Leads: {len(df)}
            Date Range: {df['uploaded_at'].min().strftime('%Y-%m-%d') if not df.empty else 'N/A'} to {df['uploaded_at'].max().strftime('%Y-%m-%d') if not df.empty else 'N/A'}
            Active Agents: {len(df['sales_agent'].unique()) if not df.empty else 0}
            Conversion Rate: {((df['status'] == 'won').sum() / len(df) * 100):.1f}% if not df.empty else 0%
            """
            ax1.text(0.1, 0.5, summary_text, fontsize=14, verticalalignment='center',
                    bbox=dict(boxstyle="round,pad=0.3", facecolor="lightblue", alpha=0.5))
            
            # Status distribution (pie chart)
            if 'status_breakdown' in charts_data and not charts_data['status_breakdown'].empty:
                ax2 = fig.add_subplot(gs[1, 0])
                status_data = charts_data['status_breakdown']
                ax2.pie(status_data['count'], labels=status_data['status'], autopct='%1.1f%%', startangle=90)
                ax2.set_title('Status Distribution')
            
            # Agent performance (bar chart)
            if 'agent_breakdown' in charts_data and not charts_data['agent_breakdown'].empty:
                ax3 = fig.add_subplot(gs[1, 1])
                agent_data = charts_data['agent_breakdown'].head(5)  # Top 5 agents
                ax3.bar(agent_data['sales_agent'], agent_data['count'])
                ax3.set_title('Top 5 Agents')
                plt.setp(ax3.xaxis.get_majorticklabels(), rotation=45)
            
            # Daily trend (line chart)
            if 'daily_leads' in charts_data and not charts_data['daily_leads'].empty:
                ax4 = fig.add_subplot(gs[1, 2])
                daily_data = charts_data['daily_leads']
                ax4.plot(daily_data['date'], daily_data['count'], marker='o')
                ax4.set_title('Daily Trend')
                plt.setp(ax4.xaxis.get_majorticklabels(), rotation=45)
            
            # Contact methods (horizontal bar)
            if 'contact_methods' in charts_data and not charts_data['contact_methods'].empty:
                ax5 = fig.add_subplot(gs[2, :])
                contact_data = charts_data['contact_methods']
                ax5.barh(contact_data['contact'], contact_data['count'])
                ax5.set_title('Contact Methods')
                ax5.set_xlabel('Number of Leads')
            
            plt.tight_layout()
            pdf.savefig(fig)
            plt.close()
        
        return pdf_buffer.getvalue()
        
    except Exception as e:
        # Fallback: create simple text report
        report_buffer = io.BytesIO()
        report_content = f"""
{title_prefix} Report - Generated on {date_str}

Summary Statistics:
- Total Leads: {len(df)}
- Date Range: {df['uploaded_at'].min().strftime('%Y-%m-%d') if not df.empty else 'N/A'} to {df['uploaded_at'].max().strftime('%Y-%m-%d') if not df.empty else 'N/A'}
- Active Agents: {len(df['sales_agent'].unique()) if not df.empty else 0}
- Conversion Rate: {((df['status'] == 'won').sum() / len(df) * 100):.1f}% if not df.empty else 0%

Chart Data Available:
{chr(10).join([f"- {chart_name}: {len(chart_df)} records" for chart_name, chart_df in charts_data.items() if isinstance(chart_df, pd.DataFrame)])}

Note: Graphs could not be generated due to missing dependencies.
Please ensure matplotlib, seaborn, and plotly are installed for full graph functionality.
"""
        report_buffer.write(report_content.encode('utf-8'))
        return report_buffer.getvalue()

def generate_plotly_graphs(df, charts_data, date_str, title_prefix="Analytics"):
    """Generate interactive Plotly graphs and return as HTML files for zip inclusion."""
    try:
        import plotly.graph_objects as go
        import plotly.express as px
        from plotly.subplots import make_subplots
        
        graphs = {}
        
        # 1. Daily Leads Trend
        if 'daily_leads' in charts_data and not charts_data['daily_leads'].empty:
            daily_data = charts_data['daily_leads']
            fig = px.line(daily_data, x='date', y='count', 
                         title=f'{title_prefix} - Daily Leads Trend',
                         markers=True)
            fig.update_layout(
                xaxis_title="Date",
                yaxis_title="Number of Leads",
                template="plotly_white"
            )
            graphs['daily_leads_trend.html'] = fig.to_html(include_plotlyjs='cdn')
        
        # 2. Agent Performance Breakdown
        if 'agent_breakdown' in charts_data and not charts_data['agent_breakdown'].empty:
            agent_data = charts_data['agent_breakdown']
            fig = px.bar(agent_data, x='sales_agent', y='count',
                        title=f'{title_prefix} - Agent Performance Breakdown',
                        color='count', color_continuous_scale='viridis')
            fig.update_layout(
                xaxis_title="Sales Agent",
                yaxis_title="Number of Leads",
                template="plotly_white"
            )
            graphs['agent_performance.html'] = fig.to_html(include_plotlyjs='cdn')
        
        # 3. Lead Status Distribution
        if 'status_breakdown' in charts_data and not charts_data['status_breakdown'].empty:
            status_data = charts_data['status_breakdown']
            fig = px.pie(status_data, values='count', names='status',
                        title=f'{title_prefix} - Lead Status Distribution')
            fig.update_layout(template="plotly_white")
            graphs['status_distribution.html'] = fig.to_html(include_plotlyjs='cdn')
        
        # 4. Sales Funnel
        if 'sales_funnel' in charts_data and not charts_data['sales_funnel'].empty:
            funnel_data = charts_data['sales_funnel']
            fig = go.Figure(go.Funnel(
                y=funnel_data['stage'],
                x=funnel_data['count'],
                textinfo="value+percent initial"
            ))
            fig.update_layout(
                title=f'{title_prefix} - Sales Funnel',
                template="plotly_white"
            )
            graphs['sales_funnel.html'] = fig.to_html(include_plotlyjs='cdn')
        
        # 5. Contact Methods Analysis
        if 'contact_methods' in charts_data and not charts_data['contact_methods'].empty:
            contact_data = charts_data['contact_methods']
            fig = px.bar(contact_data, x='contact', y='count',
                        title=f'{title_prefix} - Contact Methods Analysis',
                        color='count', color_continuous_scale='plasma')
            fig.update_layout(
                xaxis_title="Contact Method",
                yaxis_title="Number of Leads",
                template="plotly_white"
            )
            graphs['contact_methods.html'] = fig.to_html(include_plotlyjs='cdn')
        
        # 6. Activity Heatmap
        if 'activity_heatmap' in charts_data and not charts_data['activity_heatmap'].empty:
            heat_data = charts_data['activity_heatmap']
            pivot_data = heat_data.pivot_table(index='hour', columns='day_of_week', values='count', fill_value=0)
            
            fig = px.imshow(pivot_data.values,
                           x=pivot_data.columns,
                           y=pivot_data.index,
                           title=f'{title_prefix} - Activity Heatmap',
                           color_continuous_scale='YlOrRd',
                           aspect="auto")
            fig.update_layout(
                xaxis_title="Day of Week",
                yaxis_title="Hour of Day",
                template="plotly_white"
            )
            graphs['activity_heatmap.html'] = fig.to_html(include_plotlyjs='cdn')
        
        # 7. Dashboard (Combined view)
        if charts_data:
            fig = make_subplots(
                rows=2, cols=2,
                subplot_titles=('Daily Trend', 'Agent Performance', 'Status Distribution', 'Contact Methods'),
                specs=[[{"type": "scatter"}, {"type": "bar"}],
                       [{"type": "pie"}, {"type": "bar"}]]
            )
            
            # Add traces
            if 'daily_leads' in charts_data and not charts_data['daily_leads'].empty:
                daily_data = charts_data['daily_leads']
                fig.add_trace(
                    go.Scatter(x=daily_data['date'], y=daily_data['count'], mode='lines+markers'),
                    row=1, col=1
                )
            
            if 'agent_breakdown' in charts_data and not charts_data['agent_breakdown'].empty:
                agent_data = charts_data['agent_breakdown']
                fig.add_trace(
                    go.Bar(x=agent_data['sales_agent'], y=agent_data['count']),
                    row=1, col=2
                )
            
            if 'status_breakdown' in charts_data and not charts_data['status_breakdown'].empty:
                status_data = charts_data['status_breakdown']
                fig.add_trace(
                    go.Pie(values=status_data['count'], labels=status_data['status']),
                    row=2, col=1
                )
            
            if 'contact_methods' in charts_data and not charts_data['contact_methods'].empty:
                contact_data = charts_data['contact_methods']
                fig.add_trace(
                    go.Bar(x=contact_data['contact'], y=contact_data['count']),
                    row=2, col=2
                )
            
            fig.update_layout(height=800, title_text=f"{title_prefix} - Interactive Dashboard")
            graphs['interactive_dashboard.html'] = fig.to_html(include_plotlyjs='cdn')
        
        return graphs
        
    except Exception as e:
        return {}

# Layout by role
if role == 'salesman':
    st.header('Sales')
    tab_leads, tab_deals = st.tabs(['Upload & Manage Leads', 'Done Deals'])
    with tab_leads:
        # Check central lock for uploads
        with get_session() as _db_set:
            lock_row = _db_set.query(Setting).filter(Setting.key == 'uploads_locked').first()
            uploads_locked = (lock_row and (lock_row.value or '').lower() in ('1','true','yes','on'))
        uploaded_file = None
        if uploads_locked:
            st.warning('Uploads are temporarily disabled by CTO/Admin. You can still manage existing leads below.')
        else:
            uploaded_file = st.file_uploader('Upload XLSX/CSV with headers: number, name, sales agent, CONTACT, CASE, FEED BACK', type=['xlsx','xls','csv'])
        # Provide a ready-to-use template
        if st.button('Download Excel template'):
            template = pd.DataFrame({
                'number': [],
                'name': [],
                'sales agent': [],
                'contact': [],
                'case': [],
                'feed back': [],
                'how to contact': [],
            })
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                template.to_excel(writer, index=False, sheet_name='template')
            st.download_button('Download template.xlsx', data=buf.getvalue(), file_name='crm_leads_template.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        if uploaded_file is not None:
            try:
                if uploaded_file.name.endswith('.csv'):
                    df = pd.read_csv(uploaded_file)
                else:
                    df = pd.read_excel(uploaded_file)
            except Exception as e:
                st.error(f'Failed to read file: {e}')
                st.stop()

            st.subheader('Preview')
            st.dataframe(df.head())

        if not uploads_locked and st.button('Save to CRM'):
            # map columns
            col_map = {c.lower().strip(): c for c in df.columns}
            mapping = {}
            expected = ['number','name','sales agent','contact','case','feed back']
            for k in expected:
                if k in col_map:
                    mapping[col_map[k]] = k
            alt_map = {
                'sales_agent':'sales agent','salesagent':'sales agent',
                'feedback':'feed back','case_desc':'case',
                'contact_number':'contact','phone':'contact',
                'how to contact':'contact','how_to_contact':'contact','contact method':'contact'
            }
            for k,v in alt_map.items():
                if k in col_map and v not in mapping.values():
                    mapping[col_map[k]] = v
            df_ren = df.rename(columns=mapping)
            for tgt in ['number','name','sales agent','contact','case','feed back']:
                if tgt not in df_ren.columns:
                    df_ren[tgt] = None
            leads = []
            db = get_session()
            for _, row in df_ren.iterrows():
                lead = Lead(number=str(row['number']) if row['number'] is not None else None,
                            name=str(row['name']) if row['name'] is not None else None,
                            sales_agent=str(row['sales agent']) if row['sales agent'] is not None else current_user.username,
                            contact=str(row['contact']) if row['contact'] is not None else None,
                            case_desc=str(row['case']) if row['case'] is not None else None,
                            feedback=str(row['feed back']) if row['feed back'] is not None else None,
                            uploaded_by=current_user.username,
                            uploaded_by_id=current_user.id,
                            uploaded_at=datetime.utcnow())
                db.add(lead)
                db.flush()
                log_activity(db, lead.id, current_user.username, 'upload', detail='Bulk upload')
                leads.append(lead)
            db.commit()
            db.close()
            st.success(f'Saved {len(leads)} leads')

        st.markdown('---')
        st.subheader('My Leads')
        page = st.number_input('Page', min_value=1, value=1, key='leads_page')
        page_size = st.selectbox('Page size', [10,25,50], index=0, key='leads_page_size')
        offset = (page-1)*page_size
        df, total = read_leads_df(filters={'sales_agent': current_user.username}, limit=page_size, offset=offset)
        st.write(f'Total: {total}')
        if not df.empty:
            # Prepare editable table
            contact_options = ['call', 'call and whatsapp', "didn't reach"]
            status_options = ['new','contacted','qualified','lost','won']
            case_options = ['general', 'pricing', 'technical', 'support', 'complaint', 'other']
            feedback_options = ['positive', 'neutral', 'negative', 'not interested', 'call later', 'wrong number', 'closed won', 'closed lost', 'other']
            # pull sales users for dropdowns
            with get_session() as _db:
                sales_users = [u.username for u in _db.query(User).filter(User.role=='salesman').all()]
            editable_cols = ['number','name','sales_agent','contact','case_desc','feedback','status','assigned_to','comment_text']
            base_cols = ['id','number','name','sales_agent','contact','case_desc','feedback','status','assigned_to']
            present_cols = [c for c in base_cols if c in df.columns]
            df_edit = df[present_cols].copy()
            # Ensure all expected columns exist in editor
            for c in base_cols:
                if c not in df_edit.columns:
                    df_edit[c] = None
            # UI-only column for creating new comments
            df_edit['comment_text'] = ''
            df_edit['sales_agent'] = df_edit['sales_agent'].fillna(current_user.username)
            df_edit['assigned_to'] = df_edit['assigned_to'].fillna(current_user.username)
            edited = st.data_editor(
                df_edit,
                column_config={
                    'contact': st.column_config.SelectboxColumn('HOW TO CONTACT', options=contact_options, help='How did you contact?'),
                    'status': st.column_config.SelectboxColumn('Status', options=status_options),
                    'sales_agent': st.column_config.SelectboxColumn('Sales Agent', options=sales_users),
                    'assigned_to': st.column_config.SelectboxColumn('Assigned To', options=sales_users),
                    'case_desc': st.column_config.SelectboxColumn('CASE', options=sorted(set(case_options + [c for c in df_edit['case_desc'].dropna().astype(str).tolist()] ))),
                    'feedback': st.column_config.SelectboxColumn('FEED BACK', options=sorted(set(feedback_options + [c for c in df_edit['feedback'].dropna().astype(str).tolist()] ))),
                    'comment_text': st.column_config.TextColumn('Comment (new)'),
                },
                disabled=['id'],
                use_container_width=True,
                key='leads_editor'
            )
            if st.button('Save table changes', key='save_table_changes'):
                import pandas as _pd
                db = get_session()
                try:
                    orig = df_edit.set_index('id')
                    cur = edited.set_index('id', drop=False)
                    # Updates for existing ids
                    intersect_ids = [i for i in cur.index if _pd.notna(i)]
                    for lead_id in intersect_ids:
                        if lead_id not in orig.index:
                            continue
                        row_orig = orig.loc[lead_id]
                        row_new = cur.loc[lead_id]
                        changed = False
                        lead = db.query(Lead).get(int(lead_id))
                        for col in editable_cols:
                            if col == 'comment_text':
                                continue
                            new_val = row_new[col]
                            old_val = row_orig[col]
                            if (isinstance(new_val, float) and _pd.isna(new_val)) or new_val == '':
                                new_val = None
                            if (isinstance(old_val, float) and _pd.isna(old_val)) or old_val == '':
                                old_val = None
                            if new_val != old_val:
                                setattr(lead, col if col != 'case_desc' else 'case_desc', new_val)
                                changed = True
                        if changed:
                            db.add(lead)
                            db.commit()
                            log_activity(db, lead.id, current_user.username, 'edit', detail='Edited via table')
                        # Add a comment if provided
                        comment_text = (row_new.get('comment_text') or '').strip()
                        if comment_text:
                            com = Comment(lead_id=lead.id, author=current_user.username, text=comment_text)
                            db.add(com)
                            db.commit()
                            log_activity(db, lead.id, current_user.username, 'comment', detail=comment_text)
                    # Inserts for new rows (id is NaN)
                    new_rows = edited[_pd.isna(edited['id'])]
                    for _, r in new_rows.iterrows():
                        lead = Lead(
                            number=r.get('number') or None,
                            name=r.get('name') or None,
                            sales_agent=r.get('sales_agent') or current_user.username,
                            contact=r.get('contact') or None,
                            case_desc=r.get('case_desc') or None,
                            feedback=r.get('feedback') or None,
                            status=r.get('status') or 'new',
                            assigned_to=r.get('assigned_to') or None,
                            uploaded_by=current_user.username,
                            uploaded_by_id=current_user.id,
                            uploaded_at=datetime.utcnow()
                        )
                        db.add(lead)
                        db.commit()
                        log_activity(db, lead.id, current_user.username, 'create', detail='Added via table')
                        comment_text = (r.get('comment_text') or '').strip()
                        if comment_text:
                            com = Comment(lead_id=lead.id, author=current_user.username, text=comment_text)
                            db.add(com)
                            db.commit()
                            log_activity(db, lead.id, current_user.username, 'comment', detail=comment_text)
                finally:
                    db.close()
                st.success('Changes saved')
        else:
            st.info('No leads yet')

    with tab_deals:
        st.subheader('Done Deals — Add new')
        with st.form('deal_form'):
            customer_name = st.text_input('Customer name')
            phone = st.text_input('Phone number')
            screenshot_file = st.file_uploader('Payment screenshot (image)', type=['png','jpg','jpeg','webp'])
            submit_deal = st.form_submit_button('Save deal')
            if submit_deal:
                if not customer_name or not phone or not screenshot_file:
                    st.error('Please provide name, phone, and payment screenshot')
                else:
                    try:
                        screenshot_bytes = screenshot_file.read()
                        db = get_session()
                        deal = Deal(
                            customer_name=customer_name,
                            phone=phone,
                            payment_screenshot=screenshot_bytes,
                            uploaded_by=current_user.username,
                            uploaded_by_id=current_user.id,
                        )
                        db.add(deal)
                        db.commit()
                        st.success('Deal saved')
                        db.close()
                    except Exception as e:
                        st.error(f'Failed to save deal: {e}')

        st.subheader('My Deals')
        # Simple list with preview and download
        try:
            db = get_session()
            # Load only metadata first
            deals = db.query(Deal).filter(Deal.uploaded_by==current_user.username).order_by(Deal.created_at.desc()).all()
            for d in deals:
                with st.expander(f"{d.created_at:%Y-%m-%d %H:%M} — {d.customer_name} ({d.phone})"):
                    st.write(f"Salesman: {d.uploaded_by}")
                    # Display image if possible
                    try:
                        st.image(d.payment_screenshot, caption='Payment screenshot', use_container_width=True)
                    except Exception:
                        st.download_button('Download screenshot', data=d.payment_screenshot, file_name=f'deal_{d.id}_payment.png')

            # Downloads: Excel with embedded images (if supported) and ZIP of images
            if deals:
                excel_bytes = build_deals_excel_with_images(deals)
                if excel_bytes:
                    st.download_button('Download my deals (Excel with images)', data=excel_bytes, file_name='my_deals_with_images.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                zip_bytes = build_deals_images_zip(deals)
                st.download_button('Download my deal screenshots (ZIP)', data=zip_bytes, file_name='my_deal_screenshots.zip', mime='application/zip')
        finally:
            try:
                db.close()
            except Exception:
                pass

elif role == 'head_of_sales':
    st.header('Head of Sales — Overview')
    db = get_session()
    agents = [r[0] for r in db.query(Lead.sales_agent).distinct().all() if r[0]]
    db.close()
    sel_agent = st.selectbox('Filter by agent', options=['All'] + agents)
    sel_status = st.selectbox('Filter by status', options=['All','new','contacted','qualified','lost','won'])
    search = st.text_input('Search (name, number, contact)')
    page = st.number_input('Page', min_value=1, value=1)
    page_size = st.selectbox('Page size', [10,25,50], index=1)
    offset = (page-1)*page_size
    filters = {}
    if sel_agent: filters['sales_agent'] = sel_agent
    if sel_status: filters['status'] = sel_status
    df, total = read_leads_df(filters=filters, search=search, limit=page_size, offset=offset)
    st.write(f'Total matches: {total}')
    st.dataframe(df)
    if not df.empty:
        # quick KPIs
        st.subheader('KPIs')
        c1, c2, c3 = st.columns(3)
        c1.metric('Total Leads', total)
        vc = df['sales_agent'].value_counts()
        top_agent = (vc.idxmax() if not vc.empty else '—')
        c2.metric('Top Agent', top_agent)
        c3.metric('Unique Contacts', int(df['contact'].nunique()))

    st.markdown('---')
    st.subheader('Recently Assigned to Salesmen')
    with get_session() as _db_recent:
        recent = (
            _db_recent.query(Lead)
            .filter(Lead.assigned_to.isnot(None))
            .order_by(Lead.uploaded_at.desc())
            .limit(50)
            .all()
        )
        if recent:
            rec_df = pd.DataFrame([
                {
                    'id': l.id,
                    'name': l.name,
                    'number': l.number,
                    'assigned_to': l.assigned_to,
                    'status': l.status,
                    'uploaded_at': l.uploaded_at
                } for l in recent
            ])
            st.dataframe(rec_df)
        else:
            st.info('No assignments yet.')

elif role == 'cto':
    st.header('CTO Dashboard — Analytics')
    # --- CTO uploader (centralized uploads) ---
    with st.expander('Upload Leads (CTO)'):
        # Salesmen list for default assignment
        with get_session() as _db_cto_sales:
            cto_salesmen = [u.username for u in _db_cto_sales.query(User).filter(User.role=='salesman').order_by(User.username.asc()).all()]
        default_agent_opt = ['(keep from file)'] + cto_salesmen
        default_agent = st.selectbox('Default sales agent (optional)', options=default_agent_opt)
        cto_uploaded = st.file_uploader('XLSX/CSV with headers: number, name, sales agent, CONTACT, CASE, FEED BACK', type=['xlsx','xls','csv'], key='cto_uploader')
        col_left, col_right = st.columns(2)
        with col_left:
            if st.button('Download Excel template (CTO)'):
                template = pd.DataFrame({
                    'number': [],
                    'name': [],
                    'sales agent': [],
                    'contact': [],
                    'case': [],
                    'feed back': [],
                    'how to contact': [],
                })
                buf_t = io.BytesIO()
                with pd.ExcelWriter(buf_t, engine='openpyxl') as writer:
                    template.to_excel(writer, index=False, sheet_name='template')
                st.download_button('Download CTO template.xlsx', data=buf_t.getvalue(), file_name='crm_leads_template_cto.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        if cto_uploaded is not None:
            try:
                if cto_uploaded.name.endswith('.csv'):
                    cto_df = pd.read_csv(cto_uploaded)
                else:
                    cto_df = pd.read_excel(cto_uploaded)
            except Exception as e:
                st.error(f'Failed to read file: {e}')
                cto_df = None
            if cto_df is not None:
                st.dataframe(cto_df.head())
                if st.button('Save uploaded leads (CTO)'):
                    col_map = {c.lower().strip(): c for c in cto_df.columns}
                    mapping = {}
                    expected = ['number','name','sales agent','contact','case','feed back']
                    for k in expected:
                        if k in col_map:
                            mapping[col_map[k]] = k
                    alt_map = {
                        'sales_agent':'sales agent','salesagent':'sales agent',
                        'feedback':'feed back','case_desc':'case',
                        'contact_number':'contact','phone':'contact',
                        'how to contact':'contact','how_to_contact':'contact','contact method':'contact'
                    }
                    for k,v in alt_map.items():
                        if k in col_map and v not in mapping.values():
                            mapping[col_map[k]] = v
                    df_ren = cto_df.rename(columns=mapping)
                    for tgt in ['number','name','sales agent','contact','case','feed back']:
                        if tgt not in df_ren.columns:
                            df_ren[tgt] = None
                    saved = 0
                    with get_session() as db:
                        for _, row in df_ren.iterrows():
                            agent_from_file = str(row['sales agent']) if row['sales agent'] is not None else None
                            final_agent = agent_from_file
                            if (not final_agent or final_agent.strip()=='') and default_agent != '(keep from file)':
                                final_agent = default_agent
                            lead = Lead(
                                number=str(row['number']) if row['number'] is not None else None,
                                name=str(row['name']) if row['name'] is not None else None,
                                sales_agent=str(final_agent) if final_agent else None,
                                contact=str(row['contact']) if row['contact'] is not None else None,
                                case_desc=str(row['case']) if row['case'] is not None else None,
                                feedback=str(row['feed back']) if row['feed back'] is not None else None,
                                uploaded_by=current_user.username,
                                uploaded_by_id=getattr(current_user, 'id', None),
                                uploaded_at=datetime.utcnow(),
                                assigned_to=str(final_agent) if final_agent else None,
                            )
                            db.add(lead)
                            db.flush()
                            log_activity(db, lead.id, current_user.username, 'upload', detail='CTO upload')
                            saved += 1
                        db.commit()
                    st.success(f'Saved {saved} leads')
                    st.rerun()

    # --- CTO Add New Leads Manually ---
    with st.expander('Add New Leads Manually (CTO)'):
        st.write('**Add individual leads manually**')
        
        with st.form('cto_add_lead_form'):
            col1, col2 = st.columns(2)
            with col1:
                new_lead_number = st.text_input('Lead Number', placeholder='Enter lead number')
                new_lead_name = st.text_input('Lead Name', placeholder='Enter customer name')
                new_lead_contact = st.text_input('Contact Method', placeholder='e.g., call, whatsapp')
                new_lead_case = st.text_area('Case Description', placeholder='Describe the case or inquiry')
            with col2:
                new_lead_agent = st.selectbox('Sales Agent', 
                                            options=[''] + [u.username for u in get_session().query(User).filter(User.role=='salesman').order_by(User.username.asc()).all()])
                new_lead_status = st.selectbox('Status', options=['new', 'contacted', 'qualified', 'lost', 'won'])
                new_lead_feedback = st.text_area('Feedback', placeholder='Any feedback or notes')
                new_lead_assigned = st.selectbox('Assigned To', 
                                               options=[''] + [u.username for u in get_session().query(User).filter(User.role=='salesman').order_by(User.username.asc()).all()])
            
            if st.form_submit_button('➕ Add New Lead'):
                if new_lead_name and new_lead_agent:
                    with get_session() as db:
                        new_lead = Lead(
                            number=new_lead_number if new_lead_number else None,
                            name=new_lead_name,
                            sales_agent=new_lead_agent,
                            contact=new_lead_contact if new_lead_contact else None,
                            case_desc=new_lead_case if new_lead_case else None,
                            feedback=new_lead_feedback if new_lead_feedback else None,
                            status=new_lead_status,
                            assigned_to=new_lead_assigned if new_lead_assigned else new_lead_agent,
                            uploaded_by=current_user.username,
                            uploaded_by_id=current_user.id,
                            uploaded_at=datetime.utcnow()
                        )
                        db.add(new_lead)
                        db.commit()
                        log_activity(db, new_lead.id, current_user.username, 'create', detail='CTO manual lead creation')
                        st.success(f'✅ Lead "{new_lead_name}" added successfully!')
                        st.rerun()
                else:
                    st.error('❌ Please provide at least Lead Name and Sales Agent')

    df_all, total = read_leads_df(limit=10000, offset=0)
    if df_all.empty:
        st.info('No data yet. Upload leads (Sales tab) or generate demo leads below.')
        with st.expander('Demo tools (dev)'):
            n = st.number_input('How many demo leads to generate?', min_value=10, max_value=2000, value=200, step=10)
            if st.button('Generate demo leads'):
                try:
                    with get_session() as db:
                        sales_users = [u.username for u in db.query(User).filter(User.role=='salesman').all()]
                        if not sales_users:
                            u = create_user(db, 'sales_auto', 'pass', role='salesman', name='Sales Auto')
                            sales_users = [u.username]
                        contact_opts = ['call', 'call and whatsapp', "didn't reach"]
                        case_opts = ['general', 'pricing', 'technical', 'support', 'complaint', 'other']
                        feedback_opts = ['positive', 'neutral', 'negative', 'not interested', 'call later', 'wrong number', 'closed won', 'closed lost', 'other']
                        status_opts = ['new','contacted','qualified','won','lost']
                        now = datetime.utcnow()
                        for i in range(int(n)):
                            agent = random.choice(sales_users)
                            lead = Lead(
                                number=str(100000 + random.randint(0, 999999)),
                                name=f"Lead {i+1}",
                                sales_agent=agent,
                                contact=random.choice(contact_opts),
                                case_desc=random.choice(case_opts),
                                feedback=random.choice(feedback_opts),
                                status=random.choice(status_opts),
                                assigned_to=agent,
                                uploaded_by='demo',
                                uploaded_at=now - pd.Timedelta(days=random.randint(0, 60), hours=random.randint(0,23))
                            )
                            db.add(lead)
                        db.commit()
                    st.success('Demo leads generated. Reloading dashboard...')
                    st.rerun()
                except Exception as e:
                    st.error(f'Failed to generate demo leads: {e}')
    else:
        # Helpers for categorical charts
        def _normalize_text(value: str):
            if value is None:
                return None
            try:
                text = str(value).strip().lower()
            except Exception:
                return None
            if text in {'', 'na', 'n/a', 'none', 'null', '-'}:
                return None
            return text

        def _plot_top_categories(df: pd.DataFrame, column: str, title: str, top_n: int = 20):
            series = df[column].dropna().astype(str).map(_normalize_text).dropna()
            if series.empty:
                st.info(f'No values in {column} to plot.')
                return
            counts = series.value_counts().head(top_n).reset_index()
            counts.columns = [column, 'count']
            fig_local = px.bar(counts, x=column, y='count', title=title)
            st.plotly_chart(fig_local, use_container_width=True)

        df_all['uploaded_at'] = pd.to_datetime(df_all['uploaded_at'])
        df_all['date'] = df_all['uploaded_at'].dt.date

        # Filters
        min_date = df_all['date'].min()
        max_date = df_all['date'].max()
        c_f1, c_f2, c_f3 = st.columns([2,2,3])
        with c_f1:
            date_range = st.date_input('Date range', value=(min_date, max_date))
        with c_f2:
            agents_all = sorted([a for a in df_all['sales_agent'].dropna().unique().tolist()])
            sel_agents = st.multiselect('Agents', options=agents_all, default=agents_all)
        with c_f3:
            status_all = ['new','contacted','qualified','won','lost']
            sel_statuses = st.multiselect('Statuses', options=status_all, default=status_all)

        # Apply filters
        df_f = df_all.copy()
        if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
            start_d, end_d = pd.to_datetime(date_range[0]), pd.to_datetime(date_range[1])
            df_f = df_f[(df_f['uploaded_at'] >= start_d) & (df_f['uploaded_at'] <= (end_d + pd.Timedelta(days=1)))]
        if sel_agents:
            df_f = df_f[df_f['sales_agent'].isin(sel_agents)]
        if sel_statuses:
            df_f = df_f[df_f['status'].isin(sel_statuses)]

        st.subheader('Leads over time')
        daily = df_f.groupby('date').size().reset_index(name='count')
        fig = px.line(daily, x='date', y='count', title='Leads per day')
        st.plotly_chart(fig, use_container_width=True)

        # Lead Assignment (CTO)
        st.markdown('---')
        st.subheader('Lead Assignment')
        # Toggle uploads lock for salesmen
        with get_session() as _db_lock:
            lock_row = _db_lock.query(Setting).filter(Setting.key=='uploads_locked').first()
            cur_locked = (lock_row and (lock_row.value or '').lower() in ('1','true','yes','on'))
        new_locked = st.toggle('Lock salesman uploads', value=cur_locked, help='Prevent salesmen from uploading files')
        if new_locked != cur_locked:
            with get_session() as _db_lock2:
                row = _db_lock2.query(Setting).filter(Setting.key=='uploads_locked').first()
                if not row:
                    row = Setting(key='uploads_locked', value='1' if new_locked else '0')
                    _db_lock2.add(row)
                else:
                    row.value = '1' if new_locked else '0'
                _db_lock2.commit()
            st.success('Upload lock setting updated')
        with get_session() as _db_assign:
            # consider unassigned as sales_agent is null/empty/'unassigned' OR assigned_to is null
            unassigned = (
                _db_assign.query(Lead)
                .filter(
                    sa.or_(
                        Lead.sales_agent.is_(None),
                        func.trim(Lead.sales_agent) == '',
                        func.lower(func.trim(Lead.sales_agent)) == 'unassigned',
                        Lead.assigned_to.is_(None)
                    )
                )
                .order_by(Lead.uploaded_at.desc())
                .all()
            )
            st.write(f"Unassigned leads: {len(unassigned)}")
        with get_session() as _db_users:
            salesman_users = [u.username for u in _db_users.query(User).filter(User.role == 'salesman').order_by(User.username.asc()).all()]
        target_agents = st.multiselect('Select salesmen to distribute to', options=salesman_users, default=salesman_users)
        max_to_assign = st.number_input('Max leads to distribute (0 = all)', min_value=0, value=0, step=1)
        if st.button('Distribute unassigned leads equally'):
            if not target_agents:
                st.warning('Select at least one salesman.')
            else:
                with get_session() as db:
                    leads = (
                        db.query(Lead)
                        .filter(
                            sa.or_(
                                Lead.sales_agent.is_(None),
                                func.trim(Lead.sales_agent) == '',
                                func.lower(func.trim(Lead.sales_agent)) == 'unassigned',
                                Lead.assigned_to.is_(None)
                            )
                        )
                        .order_by(Lead.uploaded_at.asc())
                        .all()
                    )
                    if max_to_assign and max_to_assign > 0:
                        leads = leads[:max_to_assign]
                    if not leads:
                        st.info('No unassigned leads to distribute.')
                    else:
                        random.shuffle(leads)
                        assigned_count = {a: 0 for a in target_agents}
                        for idx, lead in enumerate(leads):
                            agent = target_agents[idx % len(target_agents)]
                            lead.sales_agent = agent
                            lead.assigned_to = agent
                            db.add(lead)
                            db.flush()
                            log_activity(db, lead.id, current_user.username, 'assign', detail=f'Assigned to {agent} by CTO')
                            assigned_count[agent] += 1
                        db.commit()
                        st.success('Distribution completed: ' + ', '.join([f"{a}: {n}" for a, n in assigned_count.items()]))
                        st.rerun()

        # Demo: randomize statuses to see distribution (dev use)
        with st.expander('Demo tools (dev)'):
            if st.button('Randomize lead statuses for demo'):
                with get_session() as db:
                    statuses = ['new','contacted','qualified','lost','won']
                    all_leads = db.query(Lead).all()
                    for lead in all_leads:
                        lead.status = random.choice(statuses)
                    db.commit()
                st.success('Statuses randomized. Refresh charts above.')

        st.subheader('Leads by agent')
        agent_counts = df_f['sales_agent'].value_counts().reset_index()
        agent_counts.columns = ['sales_agent','count']
        fig2 = px.bar(agent_counts, x='sales_agent', y='count', title='Leads by Agent')
        st.plotly_chart(fig2, use_container_width=True)

        # Login events (visibility for CTO)
        st.markdown('---')
        st.subheader('User Login Activity (last 7 days)')
        try:
            with get_session() as db:
                seven_days_ago = datetime.utcnow() - pd.Timedelta(days=7)
                logins = db.query(LoginEvent).filter(LoginEvent.logged_in_at >= seven_days_ago).order_by(LoginEvent.logged_in_at.desc()).all()
                if logins:
                    df_login = pd.DataFrame([
                        {
                            'username': e.username,
                            'role': e.role,
                            'logged_in_at': e.logged_in_at
                        } for e in logins
                    ])
                    st.dataframe(df_login)
                    agg = df_login.groupby('username').size().reset_index(name='logins')
                    fig_login = px.bar(agg, x='username', y='logins', title='Logins per user (7 days)')
                    st.plotly_chart(fig_login, use_container_width=True)
                else:
                    st.info('No login events yet.')
        except Exception:
            st.info('Login activity not available yet.')

        st.subheader('Status breakdown')
        status_counts = df_f['status'].value_counts().reset_index()
        status_counts.columns = ['status','count']
        fig3 = px.pie(status_counts, names='status', values='count', title='Leads by Status')
        st.plotly_chart(fig3, use_container_width=True)

        # Leads by status per salesman (stacked)
        st.subheader('Leads by Status per Salesman')
        status_per_sales = (
            df_f.dropna(subset=['sales_agent','status'])
                  .groupby(['sales_agent','status'])
                  .size()
                  .reset_index(name='count')
        )
        if not status_per_sales.empty:
            fig_status_sales = px.bar(
                status_per_sales,
                x='sales_agent', y='count', color='status',
                barmode='stack', title='Leads by Status per Salesman'
            )
            st.plotly_chart(fig_status_sales, use_container_width=True)
        else:
            st.info('No data for status per salesman yet.')

        st.subheader('Feedback word cloud-ish (top words)')
        feedbacks = df_f['feedback'].dropna().astype(str)
        if not feedbacks.empty:
            all_text = ' '.join(feedbacks.tolist()).lower()
            words = [w.strip('.,!?:;()"\'') for w in all_text.split() if len(w) > 3]
            if words:
                wc = pd.Series(words).value_counts().head(30).reset_index()
                wc.columns = ['word','count']
                fig4 = px.bar(wc, x='word', y='count', title='Top feedback words')
                st.plotly_chart(fig4, use_container_width=True)
            else:
                st.info('No keywords found in feedback yet.')
        else:
            st.info('No feedback text yet to analyze.')

        st.subheader('HOW TO CONTACT — top categories')
        try:
            _plot_top_categories(df_f, 'contact', 'Contact method — top categories')
        except Exception:
            st.info('No values in contact to plot.')

        st.subheader('CASE — top categories')
        try:
            _plot_top_categories(df_f, 'case_desc', 'Case — top categories')
        except Exception:
            st.info('No values in case_desc to plot.')

        st.subheader('FEED BACK — top categories')
        try:
            _plot_top_categories(df_f, 'feedback', 'Feedback — top categories')
        except Exception:
            st.info('No values in feedback to plot.')

        # Additional charts
        st.subheader('Sales Pipeline — Funnel')
        try:
            ordered_status = ['new','contacted','qualified','won','lost']
            counts_map = {s: int((df_f['status'] == s).sum()) for s in ordered_status}
            funnel_df = pd.DataFrame({'status': list(counts_map.keys()), 'count': list(counts_map.values())})
            fig_funnel = px.funnel(funnel_df, x='count', y='status', title='Lead Funnel')
            st.plotly_chart(fig_funnel, use_container_width=True)
        except Exception:
            pass

        st.subheader('Contact method by agent')
        cm = df_f.dropna(subset=['sales_agent','contact'])
        if not cm.empty:
            cm_counts = cm.groupby(['sales_agent','contact']).size().reset_index(name='count')
            fig_cm = px.bar(cm_counts, x='sales_agent', y='count', color='contact', barmode='stack', title='Contact methods by agent')
            st.plotly_chart(fig_cm, use_container_width=True)

        st.subheader('Lead uploads heatmap (weekday x hour)')
        if not df_f.empty:
            tmp = df_f.copy()
            tmp['weekday'] = tmp['uploaded_at'].dt.day_name()
            tmp['hour'] = tmp['uploaded_at'].dt.hour
            heat = tmp.groupby(['weekday','hour']).size().reset_index(name='count')
            if not heat.empty:
                fig_heat = px.density_heatmap(heat, x='hour', y='weekday', z='count', histfunc='avg', title='Uploads heatmap')
                st.plotly_chart(fig_heat, use_container_width=True)

        st.subheader('Leads per day (7-day rolling avg)')
        if not daily.empty:
            daily_ra = daily.sort_values('date').copy()
            daily_ra['rolling_7d'] = daily_ra['count'].rolling(window=7, min_periods=1).mean()
            fig_ra = px.line(daily_ra, x='date', y=['count','rolling_7d'], labels={'value':'leads','variable':'series'}, title='Daily leads and 7d avg')
            st.plotly_chart(fig_ra, use_container_width=True)

        st.subheader('Export filtered leads')
        csv_bytes = df_f.to_csv(index=False).encode('utf-8')
        st.download_button('Download CSV (filtered)', data=csv_bytes, file_name='leads_filtered.csv', mime='text/csv')
        xls_buf = io.BytesIO()
        with pd.ExcelWriter(xls_buf, engine='openpyxl') as writer:
            df_f.to_excel(writer, index=False, sheet_name='leads_filtered')
        st.download_button('Download Excel (filtered)', data=xls_buf.getvalue(), file_name='leads_filtered.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # CTO Analytics Package Download
        st.markdown('---')
        st.subheader('📦 Download Complete Analytics Package')
        
        from datetime import datetime
        import zipfile
        
        # Generate filename with current day and date
        now = datetime.now()
        day_name = now.strftime('%A')  # Full day name (e.g., 'Monday')
        date_str = now.strftime('%Y-%m-%d')  # Date format (e.g., '2024-08-09')
        zip_filename = f"CTO_Analytics_{day_name}_{date_str}.zip"
        
        if st.button('📊 Generate & Download CTO Analytics Package'):
            try:
                # Create ZIP file in memory
                zip_buffer = io.BytesIO()
                
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    # Add all chart data as Excel sheets
                    excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                        # Main filtered data
                        df_f.to_excel(writer, sheet_name='Filtered_Leads', index=False)
                        
                        # Chart data from all the charts above
                        charts_data = {}
                        
                        # Daily leads
                        if not daily.empty:
                            daily.to_excel(writer, sheet_name='Daily_Leads', index=False)
                            charts_data['daily_leads'] = daily
                        
                        # Agent breakdown
                        if not agent_counts.empty:
                            agent_counts.to_excel(writer, sheet_name='Agent_Breakdown', index=False)
                            charts_data['agent_breakdown'] = agent_counts
                        
                        # Status breakdown
                        if not status_counts.empty:
                            status_counts.to_excel(writer, sheet_name='Status_Breakdown', index=False)
                            charts_data['status_breakdown'] = status_counts
                        
                        # Sales funnel
                        if 'funnel_df' in locals():
                            funnel_df.to_excel(writer, sheet_name='Sales_Funnel', index=False)
                            charts_data['sales_funnel'] = funnel_df
                        
                        # Contact methods
                        if 'cm_counts' in locals() and not cm_counts.empty:
                            cm_counts.to_excel(writer, sheet_name='Contact_Methods', index=False)
                            charts_data['contact_methods'] = cm_counts
                        
                        # Activity heatmap
                        if 'heat' in locals() and not heat.empty:
                            heat.to_excel(writer, sheet_name='Activity_Heatmap', index=False)
                            charts_data['activity_heatmap'] = heat
                        
                        # Trends analysis
                        if 'daily_ra' in locals():
                            daily_ra.to_excel(writer, sheet_name='Trends_Analysis', index=False)
                            charts_data['trends'] = daily_ra
                        
                        # CTO Summary statistics
                        summary_data = {
                            'Metric': ['Total Leads', 'Filtered Leads', 'Active Agents', 'Date Range', 'Generated By'],
                            'Value': [
                                len(df_all),
                                len(df_f),
                                len(agent_counts) if 'agent_breakdown' in charts_data else 0,
                                f"{date_range[0] if isinstance(date_range, (list, tuple)) else 'All'} to {date_range[1] if isinstance(date_range, (list, tuple)) else 'All'}",
                                'CTO Dashboard'
                            ]
                        }
                        pd.DataFrame(summary_data).to_excel(writer, sheet_name='CTO_Summary', index=False)
                    
                    # Add Excel file to ZIP
                    zip_file.writestr(f'CTO_Analytics_{date_str}.xlsx', excel_buffer.getvalue())
                    
                    # Generate and add PDF graphs
                    pdf_graphs = generate_analytics_graphs(df_f, charts_data, date_str, "CTO Analytics")
                    zip_file.writestr(f'CTO_Graphs_{date_str}.pdf', pdf_graphs)
                    
                    # Generate and add interactive HTML graphs
                    html_graphs = generate_plotly_graphs(df_f, charts_data, date_str, "CTO Analytics")
                    for filename, html_content in html_graphs.items():
                        zip_file.writestr(f'graphs/{filename}', html_content)
                    
                    # Add deals data if available
                    deals_df, _ = read_deals_df(limit=100000)
                    if not deals_df.empty:
                        deals_buffer = io.BytesIO()
                        with pd.ExcelWriter(deals_buffer, engine='openpyxl') as deals_writer:
                            deals_df.to_excel(deals_writer, sheet_name='All_Deals', index=False)
                            
                            # Deals summary
                            deals_summary = deals_df.groupby('uploaded_by').size().reset_index(name='deals_count')
                            deals_summary.to_excel(deals_writer, sheet_name='Deals_by_Agent', index=False)
                        
                        zip_file.writestr(f'CTO_Deals_Report_{date_str}.xlsx', deals_buffer.getvalue())
                    
                    # Add a README file
                    readme_content = f"""CTO Analytics Package
Generated on: {now.strftime('%A, %B %d, %Y at %H:%M')}
Generated for: CTO Dashboard

Contents:
- CTO_Analytics_{date_str}.xlsx: Complete analytics with all charts data
- CTO_Graphs_{date_str}.pdf: Static PDF graphs and charts
- graphs/ folder: Interactive HTML graphs (open in browser)
  * daily_leads_trend.html: Daily leads trend analysis
  * agent_performance.html: Agent performance breakdown
  * status_distribution.html: Lead status distribution
  * sales_funnel.html: Sales funnel visualization
  * contact_methods.html: Contact methods analysis
  * activity_heatmap.html: Activity heatmap
  * interactive_dashboard.html: Combined interactive dashboard
- CTO_Deals_Report_{date_str}.xlsx: Deals tracking and performance data

Chart Data Included:
- Daily leads trends
- Agent performance breakdown
- Lead status distribution
- Sales funnel analysis
- Contact method analysis
- Activity heatmaps
- Rolling averages and trends
- Filtered data based on current selections

Graphs Available:
- Static PDF graphs for printing and sharing
- Interactive HTML graphs for detailed analysis
- Combined dashboard view for executive overview

CTO Summary:
- Total Leads: {len(df_all)}
- Filtered Leads: {len(df_f)}
- Active Agents: {len(agent_counts) if 'agent_breakdown' in charts_data else 0}

This package contains comprehensive technical analytics for system management.
"""
                    zip_file.writestr('README.txt', readme_content)
                
                # Offer download
                st.download_button(
                    label=f'📥 Download {zip_filename}',
                    data=zip_buffer.getvalue(),
                    file_name=zip_filename,
                    mime='application/zip',
                    help=f'Download complete CTO analytics package for {day_name}, {date_str}'
                )
                
                st.success(f'✅ CTO Analytics package ready! Contains all dashboard charts and technical reports for {day_name}, {date_str}')
                
            except Exception as e:
                st.error(f'Error creating CTO analytics package: {str(e)}')

        # Done deals charts (handles empty safely)
        st.markdown('---')
        st.subheader('Done Deals — Analytics')
        deals_df, _ = read_deals_df(limit=100000)
        if deals_df.empty:
            st.info('No deals yet')
        else:
            deals_df['created_at'] = pd.to_datetime(deals_df['created_at'])
            deals_df['date'] = deals_df['created_at'].dt.date
            if not deals_df.empty:
                dd_daily = deals_df.groupby('date').size().reset_index(name='deals')
                if not dd_daily.empty:
                    dd_fig = px.line(dd_daily, x='date', y='deals', title='Deals per day')
                    st.plotly_chart(dd_fig, use_container_width=True)
                dd_sales = deals_df['uploaded_by'].value_counts().reset_index()
                dd_sales.columns = ['salesman','deals']
                if not dd_sales.empty:
                    dd_fig2 = px.bar(dd_sales, x='salesman', y='deals', title='Deals by salesman')
                    st.plotly_chart(dd_fig2, use_container_width=True)

    # CTO Lead Archiving Section
    st.markdown('---')
    st.subheader('📦 CTO Lead Archiving Management')
    
    # Create tabs for different archiving functions
    archive_tab1, archive_tab2, archive_tab3, archive_tab4 = st.tabs(['Archive Leads', 'View Archived', 'Bulk Archive', 'Archive by Date'])
    
    with archive_tab1:
        st.write('**Archive Individual Leads**')
        
        # Get active leads for archiving
        active_leads_df, active_total = read_leads_df(limit=1000, offset=0, include_archived=False)
        
        if not active_leads_df.empty:
            # Filter options
            col1, col2, col3 = st.columns(3)
            with col1:
                agent_filter = st.selectbox('Filter by Sales Agent', 
                                          options=['All'] + sorted(active_leads_df['sales_agent'].dropna().unique().tolist()))
            with col2:
                status_filter = st.selectbox('Filter by Status', 
                                           options=['All'] + sorted(active_leads_df['status'].dropna().unique().tolist()))
            with col3:
                search_term = st.text_input('Search leads', placeholder='Name, number, or contact')
            
            # Apply filters
            filtered_df = active_leads_df.copy()
            if agent_filter != 'All':
                filtered_df = filtered_df[filtered_df['sales_agent'] == agent_filter]
            if status_filter != 'All':
                filtered_df = filtered_df[filtered_df['status'] == status_filter]
            if search_term:
                mask = (filtered_df['name'].str.contains(search_term, case=False, na=False) |
                       filtered_df['number'].str.contains(search_term, case=False, na=False) |
                       filtered_df['contact'].str.contains(search_term, case=False, na=False))
                filtered_df = filtered_df[mask]
            
            st.write(f'**Found {len(filtered_df)} leads to archive**')
            
            if not filtered_df.empty:
                # Display leads in a selectable format
                selected_leads = st.multiselect(
                    'Select leads to archive:',
                    options=filtered_df.apply(lambda x: f"ID: {x['id']} - {x['name']} ({x['sales_agent']}) - {x['status']}", axis=1).tolist(),
                    help='Select multiple leads to archive'
                )
                
                if selected_leads:
                    # Extract lead IDs from selection
                    lead_ids = []
                    for selection in selected_leads:
                        lead_id = int(selection.split(' - ')[0].replace('ID: ', ''))
                        lead_ids.append(lead_id)
                    
                    # Archive form
                    with st.form('archive_leads_form'):
                        archive_reason = st.selectbox('Archive Reason', [
                            'Completed/Closed',
                            'No longer relevant',
                            'Duplicate lead',
                            'Wrong information',
                            'Customer request',
                            'System cleanup',
                            'Other'
                        ], key='archive_reason_individual')
                        custom_reason = st.text_area('Custom reason (optional)', 
                                                   placeholder='Add additional details about why this lead is being archived')
                        archive_date = st.date_input('Archive Date', value=date.today(), key='archive_date_individual')
                        
                        if st.form_submit_button('🗄️ Archive Selected Leads'):
                            if lead_ids:
                                with get_session() as db:
                                    final_reason = f"{archive_reason}"
                                    if custom_reason.strip():
                                        final_reason += f" - {custom_reason.strip()}"
                                    
                                    archived_count = bulk_archive_leads(
                                        db, lead_ids, current_user.username, 
                                        final_reason, datetime.combine(archive_date, datetime.min.time())
                                    )
                                    
                                    if archived_count > 0:
                                        st.success(f'✅ Successfully archived {archived_count} leads')
                                        st.rerun()
                                    else:
                                        st.error('❌ Failed to archive leads')
        else:
            st.info('No active leads found to archive')
    
    with archive_tab2:
        st.write('**View Archived Leads**')
        
        # Get archived leads
        archived_leads_df, archived_total = read_leads_df(
            filters={'is_archived': 'yes'}, 
            limit=1000, 
            offset=0, 
            include_archived=True
        )
        
        if not archived_leads_df.empty:
            # Add archive information columns
            archived_leads_df['archived_date'] = pd.to_datetime(archived_leads_df['archived_at']).dt.strftime('%Y-%m-%d %H:%M')
            archived_leads_df['archive_reason_short'] = archived_leads_df['archive_reason'].str[:50] + '...'
            
            # Display archived leads
            st.dataframe(
                archived_leads_df[['id', 'name', 'sales_agent', 'status', 'archived_by', 'archived_date', 'archive_reason_short']],
                use_container_width=True
            )
            
            # Unarchive option
            st.write('**Unarchive Leads**')
            unarchive_selection = st.multiselect(
                'Select leads to unarchive:',
                options=archived_leads_df.apply(lambda x: f"ID: {x['id']} - {x['name']} (Archived: {x['archived_date']})", axis=1).tolist()
            )
            
            if unarchive_selection and st.button('🔄 Unarchive Selected Leads'):
                unarchive_ids = []
                for selection in unarchive_selection:
                    lead_id = int(selection.split(' - ')[0].replace('ID: ', ''))
                    unarchive_ids.append(lead_id)
                
                with get_session() as db:
                    unarchived_count = 0
                    for lead_id in unarchive_ids:
                        if unarchive_lead(db, lead_id, current_user.username):
                            unarchived_count += 1
                    
                    if unarchived_count > 0:
                        st.success(f'✅ Successfully unarchived {unarchived_count} leads')
                        st.rerun()
                    else:
                        st.error('❌ Failed to unarchive leads')
            
            # Bulk unarchive options
            st.write('**Bulk Unarchive Options:**')
            col1, col2 = st.columns(2)
            
            with col1:
                bulk_unarchive_reason = st.selectbox('Unarchive by Archive Reason', 
                                                   options=['All'] + sorted(archived_leads_df['archive_reason'].dropna().unique().tolist()) if not archived_leads_df.empty else ['All'], key='bulk_unarchive_reason')
                bulk_unarchive_agent = st.selectbox('Unarchive by Archived By', 
                                                  options=['All'] + sorted(archived_leads_df['archived_by'].dropna().unique().tolist()) if not archived_leads_df.empty else ['All'], key='bulk_unarchive_agent')
            
            with col2:
                bulk_unarchive_days = st.number_input('Unarchive leads archived within (days)', min_value=1, value=7, key='bulk_unarchive_days')
                bulk_unarchive_limit = st.number_input('Max leads to unarchive (0 = no limit)', min_value=0, value=0, key='bulk_unarchive_limit')
            
            if st.button('🔄 Bulk Unarchive by Criteria'):
                with get_session() as db:
                    q = db.query(Lead).filter(Lead.is_archived == 'yes')
                    
                    if bulk_unarchive_reason != 'All':
                        q = q.filter(Lead.archive_reason == bulk_unarchive_reason)
                    if bulk_unarchive_agent != 'All':
                        q = q.filter(Lead.archived_by == bulk_unarchive_agent)
                    
                    # Filter by archive date
                    cutoff_date = datetime.utcnow() - pd.Timedelta(days=bulk_unarchive_days)
                    q = q.filter(Lead.archived_at >= cutoff_date)
                    
                    if bulk_unarchive_limit > 0:
                        q = q.limit(bulk_unarchive_limit)
                    
                    leads_to_unarchive = q.all()
                    
                    if leads_to_unarchive:
                        unarchived_count = 0
                        for lead in leads_to_unarchive:
                            if unarchive_lead(db, lead.id, current_user.username):
                                unarchived_count += 1
                        
                        st.success(f'✅ Successfully unarchived {unarchived_count} leads')
                        st.rerun()
                    else:
                        st.info('No leads match the unarchive criteria')
            
            # Database Deletion Section (Danger Zone)
            st.markdown('---')
            st.write('**🗑️ Database Deletion (Danger Zone)**')
            st.warning('⚠️ **WARNING**: This will permanently delete leads from the database. This action cannot be undone!')
            
            # Individual lead deletion
            st.write('**Delete Individual Leads:**')
            delete_selection = st.multiselect(
                'Select leads to permanently delete:',
                options=archived_leads_df.apply(lambda x: f"ID: {x['id']} - {x['name']} (Archived: {x['archived_date']})", axis=1).tolist(),
                key='delete_selection'
            )
            
            if delete_selection:
                delete_reason = st.text_area('Deletion reason (required)', 
                                           placeholder='Explain why these leads are being permanently deleted', 
                                           key='delete_reason')
                
                if st.button('🗑️ Permanently Delete Selected Leads', type='secondary', key='delete_selected'):
                    if delete_reason.strip():
                        delete_ids = []
                        for selection in delete_selection:
                            lead_id = int(selection.split(' - ')[0].replace('ID: ', ''))
                            delete_ids.append(lead_id)
                        
                        with get_session() as db:
                            deleted_count = bulk_delete_leads_from_db(db, delete_ids, current_user.username, delete_reason.strip())
                            
                            if deleted_count > 0:
                                st.success(f'✅ Successfully deleted {deleted_count} leads from database')
                                st.rerun()
                            else:
                                st.error('❌ Failed to delete leads')
                    else:
                        st.error('❌ Please provide a deletion reason')
            
            # Bulk deletion options
            st.write('**Bulk Delete Options:**')
            col1, col2 = st.columns(2)
            
            with col1:
                bulk_delete_reason = st.selectbox('Delete by Archive Reason', 
                                                options=['All'] + sorted(archived_leads_df['archive_reason'].dropna().unique().tolist()) if not archived_leads_df.empty else ['All'], key='bulk_delete_reason')
                bulk_delete_agent = st.selectbox('Delete by Archived By', 
                                               options=['All'] + sorted(archived_leads_df['archived_by'].dropna().unique().tolist()) if not archived_leads_df.empty else ['All'], key='bulk_delete_agent')
            
            with col2:
                bulk_delete_days = st.number_input('Delete leads archived within (days)', min_value=1, value=30, key='bulk_delete_days')
                bulk_delete_limit = st.number_input('Max leads to delete (0 = no limit)', min_value=0, value=0, key='bulk_delete_limit')
            
            bulk_delete_reason_text = st.text_area('Bulk deletion reason (required)', 
                                                 placeholder='Explain why these leads are being permanently deleted', 
                                                 key='bulk_delete_reason_text')
            
            if st.button('🗑️ Bulk Delete by Criteria', type='secondary', key='bulk_delete_criteria'):
                if bulk_delete_reason_text.strip():
                    with get_session() as db:
                        q = db.query(Lead).filter(Lead.is_archived == 'yes')
                        
                        if bulk_delete_reason != 'All':
                            q = q.filter(Lead.archive_reason == bulk_delete_reason)
                        if bulk_delete_agent != 'All':
                            q = q.filter(Lead.archived_by == bulk_delete_agent)
                        
                        # Filter by archive date
                        cutoff_date = datetime.utcnow() - pd.Timedelta(days=bulk_delete_days)
                        q = q.filter(Lead.archived_at >= cutoff_date)
                        
                        if bulk_delete_limit > 0:
                            q = q.limit(bulk_delete_limit)
                        
                        leads_to_delete = q.all()
                        
                        if leads_to_delete:
                            lead_ids = [lead.id for lead in leads_to_delete]
                            deleted_count = bulk_delete_leads_from_db(db, lead_ids, current_user.username, bulk_delete_reason_text.strip())
                            
                            if deleted_count > 0:
                                st.success(f'✅ Successfully deleted {deleted_count} leads from database')
                                st.rerun()
                            else:
                                st.error('❌ Failed to delete leads')
                        else:
                            st.info('No leads match the deletion criteria')
                else:
                    st.error('❌ Please provide a deletion reason')
            
            # Quick delete options
            st.write('**Quick Delete Options:**')
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button('🗑️ Delete All Archived Leads', type='secondary', key='delete_all_archived'):
                    st.error('⚠️ This will delete ALL archived leads! Are you sure?')
                    confirm_delete_all = st.checkbox('I understand this will permanently delete ALL archived leads', key='confirm_delete_all')
                    
                    if confirm_delete_all:
                        delete_all_reason = st.text_area('Reason for deleting all archived leads', 
                                                       placeholder='Explain why all archived leads should be deleted', 
                                                       key='delete_all_reason')
                        
                        if st.button('🗑️ Confirm Delete All', type='primary', key='confirm_delete_all_btn'):
                            if delete_all_reason.strip():
                                with get_session() as db:
                                    all_archived = db.query(Lead).filter(Lead.is_archived == 'yes').all()
                                    if all_archived:
                                        lead_ids = [lead.id for lead in all_archived]
                                        deleted_count = bulk_delete_leads_from_db(db, lead_ids, current_user.username, delete_all_reason.strip())
                                        st.success(f'✅ Successfully deleted {deleted_count} leads from database')
                                        st.rerun()
                                    else:
                                        st.info('No archived leads found')
                            else:
                                st.error('❌ Please provide a deletion reason')
            
            with col2:
                quick_delete_days = st.selectbox('Delete archives older than', [30, 60, 90, 180, 365], key='quick_delete_days')
                if st.button(f'🗑️ Delete Archives Older Than {quick_delete_days} Days', type='secondary', key='quick_delete_old'):
                    cutoff_date = datetime.utcnow() - pd.Timedelta(days=quick_delete_days)
                    
                    with get_session() as db:
                        old_archived = db.query(Lead).filter(
                            Lead.is_archived == 'yes',
                            Lead.archived_at <= cutoff_date
                        ).all()
                        
                        if old_archived:
                            lead_ids = [lead.id for lead in old_archived]
                            deleted_count = bulk_delete_leads_from_db(db, lead_ids, current_user.username, f'Automatic cleanup - archives older than {quick_delete_days} days')
                            st.success(f'✅ Successfully deleted {deleted_count} old archived leads')
                            st.rerun()
                        else:
                            st.info(f'No archived leads older than {quick_delete_days} days found')
            
            # Enhanced Export Archived Leads
            st.write('**📊 Enhanced Export Archived Leads**')
            
            # Export options tabs
            export_tab1, export_tab2, export_tab3 = st.tabs(['📅 Date Range Export', '📋 Quick Exports', '📈 Analytics Export'])
            
            with export_tab1:
                st.write('**Export by Date Range**')
                col1, col2, col3 = st.columns(3)
                with col1:
                    export_start_date = st.date_input('Export from date', value=date.today() - pd.Timedelta(days=30), key='export_start_date')
                    export_format = st.selectbox('Export format', ['Excel', 'CSV'], key='export_format')
                with col2:
                    export_end_date = st.date_input('Export to date', value=date.today(), key='export_end_date')
                    export_include_analytics = st.checkbox('Include analytics summary', value=True, key='export_analytics')
                with col3:
                    export_filter_reason = st.selectbox('Filter by archive reason', 
                                                      options=['All'] + sorted(archived_leads_df['archive_reason'].dropna().unique().tolist()) if not archived_leads_df.empty else ['All'], key='export_filter_reason')
                    export_filter_agent = st.selectbox('Filter by archived by', 
                                                     options=['All'] + sorted(archived_leads_df['archived_by'].dropna().unique().tolist()) if not archived_leads_df.empty else ['All'], key='export_filter_agent')
                
                if st.button('📊 Export Archived Leads Report', key='export_date_range'):
                    with get_session() as db:
                        start_datetime = datetime.combine(export_start_date, datetime.min.time())
                        end_datetime = datetime.combine(export_end_date, datetime.max.time())
                        
                        # Build enhanced query with filters
                        q = db.query(Lead).filter(
                            Lead.is_archived == 'yes',
                            Lead.archived_at >= start_datetime,
                            Lead.archived_at <= end_datetime
                        )
                        
                        if export_filter_reason != 'All':
                            q = q.filter(Lead.archive_reason == export_filter_reason)
                        if export_filter_agent != 'All':
                            q = q.filter(Lead.archived_by == export_filter_agent)
                        
                        filtered_leads = q.all()
                        
                        if filtered_leads:
                            # Create enhanced export data
                            export_data = []
                            for lead in filtered_leads:
                                export_data.append({
                                    'Lead ID': lead.id,
                                    'Lead Number': lead.number,
                                    'Customer Name': lead.name,
                                    'Sales Agent': lead.sales_agent,
                                    'Contact Method': lead.contact,
                                    'Case Description': lead.case_desc,
                                    'Feedback': lead.feedback,
                                    'Status': lead.status,
                                    'Assigned To': lead.assigned_to,
                                    'Original Upload Date': lead.uploaded_at.strftime('%Y-%m-%d %H:%M') if lead.uploaded_at else '',
                                    'Archived By': lead.archived_by,
                                    'Archive Date': lead.archived_at.strftime('%Y-%m-%d %H:%M') if lead.archived_at else '',
                                    'Archive Reason': lead.archive_reason,
                                    'Scheduled Archive Date': lead.archive_date.strftime('%Y-%m-%d %H:%M') if lead.archive_date else '',
                                    'Days Since Archive': (datetime.utcnow() - lead.archived_at).days if lead.archived_at else 0
                                })
                            
                            df_export = pd.DataFrame(export_data)
                            
                            if export_format == 'Excel':
                                buffer = io.BytesIO()
                                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                                    df_export.to_excel(writer, sheet_name='Archived_Leads', index=False)
                                    
                                    if export_include_analytics:
                                        # Analytics summary sheet
                                        analytics_data = {
                                            'Metric': [
                                                'Total Archived Leads',
                                                'Date Range',
                                                'Filtered By Reason',
                                                'Filtered By Agent',
                                                'Average Days Since Archive',
                                                'Most Common Archive Reason',
                                                'Most Active Archiver',
                                                'Export Generated By',
                                                'Export Generated On'
                                            ],
                                            'Value': [
                                                len(filtered_leads),
                                                f"{export_start_date} to {export_end_date}",
                                                export_filter_reason,
                                                export_filter_agent,
                                                f"{df_export['Days Since Archive'].mean():.1f} days",
                                                df_export['Archive Reason'].mode().iloc[0] if not df_export['Archive Reason'].mode().empty else 'N/A',
                                                df_export['Archived By'].mode().iloc[0] if not df_export['Archived By'].mode().empty else 'N/A',
                                                current_user.username,
                                                datetime.now().strftime('%Y-%m-%d %H:%M')
                                            ]
                                        }
                                        pd.DataFrame(analytics_data).to_excel(writer, sheet_name='Analytics_Summary', index=False)
                                        
                                        # Archive reasons breakdown
                                        reason_counts = df_export['Archive Reason'].value_counts().reset_index()
                                        reason_counts.columns = ['Archive Reason', 'Count']
                                        reason_counts.to_excel(writer, sheet_name='Archive_Reasons', index=False)
                                
                                filename = f"Archived_Leads_{export_start_date}_{export_end_date}.xlsx"
                                st.download_button(
                                    label=f'📥 Download {filename}',
                                    data=buffer.getvalue(),
                                    file_name=filename,
                                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                                )
                            else:
                                csv_data = df_export.to_csv(index=False)
                                filename = f"Archived_Leads_{export_start_date}_{export_end_date}.csv"
                                st.download_button(
                                    label=f'📥 Download {filename}',
                                    data=csv_data,
                                    file_name=filename,
                                    mime='text/csv'
                                )
                            
                            st.success(f'✅ Export ready! {len(filtered_leads)} leads included.')
                        else:
                            st.info('No archived leads found matching the criteria.')
            
            with export_tab2:
                st.write('**Quick Export Options**')
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write('**Export All Archived:**')
                    if st.button('📊 Export All Archived Leads', key='export_all'):
                        with get_session() as db:
                            all_archived = db.query(Lead).filter(Lead.is_archived == 'yes').all()
                            if all_archived:
                                export_data = []
                                for lead in all_archived:
                                    export_data.append({
                                        'Lead ID': lead.id,
                                        'Customer Name': lead.name,
                                        'Sales Agent': lead.sales_agent,
                                        'Status': lead.status,
                                        'Archived By': lead.archived_by,
                                        'Archive Date': lead.archived_at.strftime('%Y-%m-%d %H:%M') if lead.archived_at else '',
                                        'Archive Reason': lead.archive_reason
                                    })
                                
                                df_all = pd.DataFrame(export_data)
                                buffer = io.BytesIO()
                                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                                    df_all.to_excel(writer, sheet_name='All_Archived_Leads', index=False)
                                
                                filename = f"All_Archived_Leads_{datetime.now().strftime('%Y%m%d')}.xlsx"
                                st.download_button(
                                    label=f'📥 Download {filename}',
                                    data=buffer.getvalue(),
                                    file_name=filename,
                                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                                )
                                st.success(f'✅ All {len(all_archived)} archived leads exported!')
                            else:
                                st.info('No archived leads found.')
                    
                    st.write('**Export Recent Archives:**')
                    recent_days = st.selectbox('Export archives from last', [7, 14, 30, 60, 90], key='recent_days')
                    if st.button('📊 Export Recent Archives', key='export_recent'):
                        with get_session() as db:
                            cutoff_date = datetime.utcnow() - pd.Timedelta(days=recent_days)
                            recent_archived = db.query(Lead).filter(
                                Lead.is_archived == 'yes',
                                Lead.archived_at >= cutoff_date
                            ).all()
                            
                            if recent_archived:
                                export_data = []
                                for lead in recent_archived:
                                    export_data.append({
                                        'Lead ID': lead.id,
                                        'Customer Name': lead.name,
                                        'Sales Agent': lead.sales_agent,
                                        'Status': lead.status,
                                        'Archived By': lead.archived_by,
                                        'Archive Date': lead.archived_at.strftime('%Y-%m-%d %H:%M') if lead.archived_at else '',
                                        'Archive Reason': lead.archive_reason
                                    })
                                
                                df_recent = pd.DataFrame(export_data)
                                buffer = io.BytesIO()
                                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                                    df_recent.to_excel(writer, sheet_name=f'Recent_Archives_{recent_days}d', index=False)
                                
                                filename = f"Recent_Archives_{recent_days}days_{datetime.now().strftime('%Y%m%d')}.xlsx"
                                st.download_button(
                                    label=f'📥 Download {filename}',
                                    data=buffer.getvalue(),
                                    file_name=filename,
                                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                                )
                                st.success(f'✅ {len(recent_archived)} recent archives exported!')
                            else:
                                st.info(f'No archives found in the last {recent_days} days.')
                
                with col2:
                    st.write('**Export by Archive Reason:**')
                    quick_reason = st.selectbox('Select archive reason', 
                                              options=sorted(archived_leads_df['archive_reason'].dropna().unique().tolist()) if not archived_leads_df.empty else [], key='quick_reason')
                    if st.button('📊 Export by Reason', key='export_by_reason'):
                        with get_session() as db:
                            reason_archived = db.query(Lead).filter(
                                Lead.is_archived == 'yes',
                                Lead.archive_reason == quick_reason
                            ).all()
                            
                            if reason_archived:
                                export_data = []
                                for lead in reason_archived:
                                    export_data.append({
                                        'Lead ID': lead.id,
                                        'Customer Name': lead.name,
                                        'Sales Agent': lead.sales_agent,
                                        'Status': lead.status,
                                        'Archived By': lead.archived_by,
                                        'Archive Date': lead.archived_at.strftime('%Y-%m-%d %H:%M') if lead.archived_at else '',
                                        'Archive Reason': lead.archive_reason
                                    })
                                
                                df_reason = pd.DataFrame(export_data)
                                buffer = io.BytesIO()
                                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                                    df_reason.to_excel(writer, sheet_name=f'Reason_{quick_reason[:20]}', index=False)
                                
                                filename = f"Archives_Reason_{quick_reason[:20]}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                                st.download_button(
                                    label=f'📥 Download {filename}',
                                    data=buffer.getvalue(),
                                    file_name=filename,
                                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                                )
                                st.success(f'✅ {len(reason_archived)} leads exported for reason: {quick_reason}')
                            else:
                                st.info(f'No archives found for reason: {quick_reason}')
            
            with export_tab3:
                st.write('**Analytics Export**')
                st.write('Generate comprehensive analytics reports for archived data.')
                
                col1, col2 = st.columns(2)
                with col1:
                    analytics_start = st.date_input('Analytics start date', value=date.today() - pd.Timedelta(days=90), key='analytics_start')
                    analytics_end = st.date_input('Analytics end date', value=date.today(), key='analytics_end')
                with col2:
                    include_charts = st.checkbox('Include chart data', value=True, key='include_charts')
                    include_summary = st.checkbox('Include executive summary', value=True, key='include_summary')
                
                if st.button('📊 Generate Analytics Report', key='generate_analytics'):
                    with get_session() as db:
                        start_datetime = datetime.combine(analytics_start, datetime.min.time())
                        end_datetime = datetime.combine(analytics_end, datetime.max.time())
                        
                        archived_in_period = db.query(Lead).filter(
                            Lead.is_archived == 'yes',
                            Lead.archived_at >= start_datetime,
                            Lead.archived_at <= end_datetime
                        ).all()
                        
                        if archived_in_period:
                            # Create comprehensive analytics report
                            buffer = io.BytesIO()
                            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                                # Main data
                                export_data = []
                                for lead in archived_in_period:
                                    export_data.append({
                                        'Lead ID': lead.id,
                                        'Customer Name': lead.name,
                                        'Sales Agent': lead.sales_agent,
                                        'Status': lead.status,
                                        'Archived By': lead.archived_by,
                                        'Archive Date': lead.archived_at.strftime('%Y-%m-%d %H:%M') if lead.archived_at else '',
                                        'Archive Reason': lead.archive_reason
                                    })
                                
                                df_analytics = pd.DataFrame(export_data)
                                df_analytics.to_excel(writer, sheet_name='Archived_Data', index=False)
                                
                                if include_summary:
                                    # Executive summary
                                    summary_data = {
                                        'Metric': [
                                            'Total Archives in Period',
                                            'Period Start',
                                            'Period End',
                                            'Most Common Archive Reason',
                                            'Most Active Archiver',
                                            'Average Archives per Day',
                                            'Peak Archive Day',
                                            'Report Generated By',
                                            'Report Generated On'
                                        ],
                                        'Value': [
                                            len(archived_in_period),
                                            analytics_start,
                                            analytics_end,
                                            df_analytics['Archive Reason'].mode().iloc[0] if not df_analytics['Archive Reason'].mode().empty else 'N/A',
                                            df_analytics['Archived By'].mode().iloc[0] if not df_analytics['Archived By'].mode().empty else 'N/A',
                                            f"{len(archived_in_period) / ((end_datetime - start_datetime).days + 1):.1f}",
                                            df_analytics['Archive Date'].mode().iloc[0] if not df_analytics['Archive Date'].mode().empty else 'N/A',
                                            current_user.username,
                                            datetime.now().strftime('%Y-%m-%d %H:%M')
                                        ]
                                    }
                                    pd.DataFrame(summary_data).to_excel(writer, sheet_name='Executive_Summary', index=False)
                                
                                if include_charts:
                                    # Archive reasons breakdown
                                    reason_counts = df_analytics['Archive Reason'].value_counts().reset_index()
                                    reason_counts.columns = ['Archive Reason', 'Count']
                                    reason_counts.to_excel(writer, sheet_name='Archive_Reasons', index=False)
                                    
                                    # Archive by agent
                                    agent_counts = df_analytics['Archived By'].value_counts().reset_index()
                                    agent_counts.columns = ['Archived By', 'Count']
                                    agent_counts.to_excel(writer, sheet_name='Archive_by_Agent', index=False)
                                    
                                    # Daily archive trends
                                    df_analytics['Archive Date Only'] = pd.to_datetime(df_analytics['Archive Date']).dt.date
                                    daily_counts = df_analytics['Archive Date Only'].value_counts().reset_index()
                                    daily_counts.columns = ['Date', 'Archives']
                                    daily_counts = daily_counts.sort_values('Date')
                                    daily_counts.to_excel(writer, sheet_name='Daily_Trends', index=False)
                            
                            filename = f"Analytics_Report_{analytics_start}_{analytics_end}.xlsx"
                            st.download_button(
                                label=f'📥 Download Analytics Report',
                                data=buffer.getvalue(),
                                file_name=filename,
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            )
                            st.success(f'✅ Analytics report generated with {len(archived_in_period)} records!')
                        else:
                            st.info('No archived data found for the selected period.')
        else:
            st.info('No archived leads found')
    
    with archive_tab3:
        st.write('**🚀 Advanced Bulk Archiving System**')
        
        # Create sub-tabs for different bulk archiving methods
        bulk_tab1, bulk_tab2, bulk_tab3, bulk_tab4 = st.tabs(['📊 Smart Bulk Archive', '📅 Date-Based Bulk', '👥 Agent-Based Bulk', '⚡ Quick Bulk Actions'])
        
        with bulk_tab1:
            st.write('**Smart Bulk Archive - Multiple Criteria**')
            
            # Advanced filtering options
            col1, col2 = st.columns(2)
            with col1:
                smart_agents = st.multiselect('Select Sales Agents', 
                                            options=sorted(active_leads_df['sales_agent'].dropna().unique().tolist()) if not active_leads_df.empty else [],
                                            help='Leave empty to select all agents')
                smart_statuses = st.multiselect('Select Statuses', 
                                              options=['new', 'contacted', 'qualified', 'lost', 'won'],
                                              default=['new', 'contacted'],
                                              help='Select which statuses to archive')
                smart_contact_methods = st.multiselect('Select Contact Methods',
                                                     options=sorted(active_leads_df['contact'].dropna().unique().tolist()) if not active_leads_df.empty else [],
                                                     help='Leave empty to include all contact methods')
            
            with col2:
                smart_days_old = st.number_input('Archive leads older than (days)', min_value=1, value=30, key='smart_days_old')
                smart_max_leads = st.number_input('Maximum leads to archive (0 = no limit)', min_value=0, value=0, help='Limit the number of leads to prevent accidental bulk operations', key='smart_max_leads')
                smart_reason = st.selectbox('Archive Reason', [
                    'Smart cleanup - old leads',
                    'Completed/Closed leads',
                    'No longer relevant',
                    'Duplicate cleanup',
                    'System maintenance',
                    'Performance optimization',
                    'Data cleanup',
                    'Other'
                ], key='smart_reason')
                custom_reason = st.text_area('Custom reason details', placeholder='Add specific details about this bulk archive operation')
            
            # Preview functionality
            if st.button('🔍 Preview Leads to Archive'):
                with get_session() as db:
                    q = db.query(Lead).filter(sa.or_(Lead.is_archived.is_(None), Lead.is_archived == 'no'))
                    
                    if smart_agents:
                        q = q.filter(Lead.sales_agent.in_(smart_agents))
                    if smart_statuses:
                        q = q.filter(Lead.status.in_(smart_statuses))
                    if smart_contact_methods:
                        q = q.filter(Lead.contact.in_(smart_contact_methods))
                    
                    # Filter by age
                    cutoff_date = datetime.utcnow() - pd.Timedelta(days=smart_days_old)
                    q = q.filter(Lead.uploaded_at <= cutoff_date)
                    
                    preview_leads = q.limit(100).all()
                    
                    if preview_leads:
                        preview_df = pd.DataFrame([{
                            'ID': lead.id,
                            'Name': lead.name,
                            'Agent': lead.sales_agent,
                            'Status': lead.status,
                            'Contact': lead.contact,
                            'Uploaded': lead.uploaded_at.strftime('%Y-%m-%d') if lead.uploaded_at else '',
                            'Age (days)': (datetime.utcnow() - lead.uploaded_at).days if lead.uploaded_at else 0
                        } for lead in preview_leads])
                        
                        st.write(f'**Preview of leads to archive (showing first 100):**')
                        st.dataframe(preview_df, use_container_width=True)
                        
                        total_count = q.count()
                        st.info(f'📊 Total leads matching criteria: {total_count}')
                        
                        if smart_max_leads > 0 and total_count > smart_max_leads:
                            st.warning(f'⚠️ Will archive only {smart_max_leads} leads (limit set)')
                    else:
                        st.info('No leads match the selected criteria')
            
            # Execute bulk archive
            if st.button('🗄️ Execute Smart Bulk Archive'):
                with get_session() as db:
                    q = db.query(Lead).filter(sa.or_(Lead.is_archived.is_(None), Lead.is_archived == 'no'))
                    
                    if smart_agents:
                        q = q.filter(Lead.sales_agent.in_(smart_agents))
                    if smart_statuses:
                        q = q.filter(Lead.status.in_(smart_statuses))
                    if smart_contact_methods:
                        q = q.filter(Lead.contact.in_(smart_contact_methods))
                    
                    # Filter by age
                    cutoff_date = datetime.utcnow() - pd.Timedelta(days=smart_days_old)
                    q = q.filter(Lead.uploaded_at <= cutoff_date)
                    
                    if smart_max_leads > 0:
                        q = q.limit(smart_max_leads)
                    
                    leads_to_archive = q.all()
                    
                    if leads_to_archive:
                        lead_ids = [lead.id for lead in leads_to_archive]
                        final_reason = smart_reason
                        if custom_reason.strip():
                            final_reason += f" - {custom_reason.strip()}"
                        
                        archived_count = bulk_archive_leads(
                            db, lead_ids, current_user.username, 
                            final_reason, datetime.utcnow()
                        )
                        
                        st.success(f'✅ Successfully archived {archived_count} leads!')
                        st.info(f'📋 Archive reason: {final_reason}')
                        st.rerun()
                    else:
                        st.info('No leads match the selected criteria')
        
        with bulk_tab2:
            st.write('**Date-Based Bulk Archiving**')
            
            col1, col2 = st.columns(2)
            with col1:
                date_archive_start = st.date_input('Start Date', value=date.today() - pd.Timedelta(days=60), key='date_archive_start_tab2')
                date_archive_end = st.date_input('End Date', value=date.today(), key='date_archive_end_tab2')
                date_archive_agents = st.multiselect('Filter by Agents (optional)',
                                                   options=sorted(active_leads_df['sales_agent'].dropna().unique().tolist()) if not active_leads_df.empty else [], key='date_archive_agents_tab2')
            with col2:
                date_archive_statuses = st.multiselect('Filter by Status (optional)',
                                                     options=['new', 'contacted', 'qualified', 'lost', 'won'],
                                                     default=['new', 'contacted'], key='date_archive_statuses_tab2')
                date_archive_reason = st.selectbox('Date Archive Reason', [
                    'Date-based cleanup',
                    'Old leads cleanup',
                    'System maintenance',
                    'Historical data management',
                    'Performance optimization',
                    'Other'
                ], key='date_archive_reason')
            
            # Preview date-based archive
            if st.button('📅 Preview Date-Based Archive'):
                with get_session() as db:
                    start_datetime = datetime.combine(date_archive_start, datetime.min.time())
                    end_datetime = datetime.combine(date_archive_end, datetime.max.time())
                    
                    q = db.query(Lead).filter(
                        sa.or_(Lead.is_archived.is_(None), Lead.is_archived == 'no'),
                        Lead.uploaded_at >= start_datetime,
                        Lead.uploaded_at <= end_datetime
                    )
                    
                    if date_archive_agents:
                        q = q.filter(Lead.sales_agent.in_(date_archive_agents))
                    if date_archive_statuses:
                        q = q.filter(Lead.status.in_(date_archive_statuses))
                    
                    preview_leads = q.limit(100).all()
                    
                    if preview_leads:
                        preview_df = pd.DataFrame([{
                            'ID': lead.id,
                            'Name': lead.name,
                            'Agent': lead.sales_agent,
                            'Status': lead.status,
                            'Uploaded': lead.uploaded_at.strftime('%Y-%m-%d %H:%M') if lead.uploaded_at else ''
                        } for lead in preview_leads])
                        
                        st.write(f'**Preview of leads to archive from {date_archive_start} to {date_archive_end}:**')
                        st.dataframe(preview_df, use_container_width=True)
                        
                        total_count = q.count()
                        st.info(f'📊 Total leads in date range: {total_count}')
                    else:
                        st.info('No leads found in the specified date range')
            
            # Execute date-based archive
            if st.button('🗄️ Execute Date-Based Archive'):
                with get_session() as db:
                    start_datetime = datetime.combine(date_archive_start, datetime.min.time())
                    end_datetime = datetime.combine(date_archive_end, datetime.max.time())
                    
                    q = db.query(Lead).filter(
                        sa.or_(Lead.is_archived.is_(None), Lead.is_archived == 'no'),
                        Lead.uploaded_at >= start_datetime,
                        Lead.uploaded_at <= end_datetime
                    )
                    
                    if date_archive_agents:
                        q = q.filter(Lead.sales_agent.in_(date_archive_agents))
                    if date_archive_statuses:
                        q = q.filter(Lead.status.in_(date_archive_statuses))
                    
                    leads_in_range = q.all()
                    
                    if leads_in_range:
                        lead_ids = [lead.id for lead in leads_in_range]
                        archived_count = bulk_archive_leads(
                            db, lead_ids, current_user.username, 
                            f"{date_archive_reason} ({date_archive_start} to {date_archive_end})", 
                            datetime.utcnow()
                        )
                        st.success(f'✅ Successfully archived {archived_count} leads from {date_archive_start} to {date_archive_end}')
                        st.rerun()
                    else:
                        st.info(f'No leads found in date range {date_archive_start} to {date_archive_end}')
        
        with bulk_tab3:
            st.write('**Agent-Based Bulk Archiving**')
            
            # Get all sales agents
            with get_session() as db:
                all_agents = [u.username for u in db.query(User).filter(User.role=='salesman').order_by(User.username.asc()).all()]
            
            col1, col2 = st.columns(2)
            with col1:
                agent_bulk_selection = st.multiselect('Select Agents to Archive From',
                                                    options=all_agents,
                                                    help='Select which agents\' leads to archive')
                agent_bulk_statuses = st.multiselect('Statuses to Archive',
                                                   options=['new', 'contacted', 'qualified', 'lost', 'won'],
                                                   default=['new', 'contacted'])
                agent_bulk_days = st.number_input('Archive leads older than (days)', min_value=1, value=30, key='agent_bulk_days')
            
            with col2:
                agent_bulk_reason = st.selectbox('Agent Archive Reason', [
                    'Agent performance cleanup',
                    'Agent reassignment',
                    'Agent inactive leads',
                    'Agent data cleanup',
                    'Other'
                ], key='agent_bulk_reason')
                agent_bulk_limit = st.number_input('Max leads per agent (0 = no limit)', min_value=0, value=0, key='agent_bulk_limit')
            
            if st.button('👥 Preview Agent-Based Archive'):
                if not agent_bulk_selection:
                    st.warning('Please select at least one agent')
                else:
                    with get_session() as db:
                        cutoff_date = datetime.utcnow() - pd.Timedelta(days=agent_bulk_days)
                        
                        agent_summary = []
                        for agent in agent_bulk_selection:
                            q = db.query(Lead).filter(
                                sa.or_(Lead.is_archived.is_(None), Lead.is_archived == 'no'),
                                Lead.sales_agent == agent,
                                Lead.status.in_(agent_bulk_statuses),
                                Lead.uploaded_at <= cutoff_date
                            )
                            count = q.count()
                            agent_summary.append({'Agent': agent, 'Leads to Archive': count})
                        
                        summary_df = pd.DataFrame(agent_summary)
                        st.write('**Agent Archive Summary:**')
                        st.dataframe(summary_df, use_container_width=True)
                        
                        total_leads = sum(item['Leads to Archive'] for item in agent_summary)
                        st.info(f'📊 Total leads to archive: {total_leads}')
            
            if st.button('🗄️ Execute Agent-Based Archive'):
                if not agent_bulk_selection:
                    st.warning('Please select at least one agent')
                else:
                    with get_session() as db:
                        cutoff_date = datetime.utcnow() - pd.Timedelta(days=agent_bulk_days)
                        total_archived = 0
                        
                        for agent in agent_bulk_selection:
                            q = db.query(Lead).filter(
                                sa.or_(Lead.is_archived.is_(None), Lead.is_archived == 'no'),
                                Lead.sales_agent == agent,
                                Lead.status.in_(agent_bulk_statuses),
                                Lead.uploaded_at <= cutoff_date
                            )
                            
                            if agent_bulk_limit > 0:
                                q = q.limit(agent_bulk_limit)
                            
                            leads_to_archive = q.all()
                            
                            if leads_to_archive:
                                lead_ids = [lead.id for lead in leads_to_archive]
                                archived_count = bulk_archive_leads(
                                    db, lead_ids, current_user.username, 
                                    f"{agent_bulk_reason} - Agent: {agent}", 
                                    datetime.utcnow()
                                )
                                total_archived += archived_count
                        
                        st.success(f'✅ Successfully archived {total_archived} leads from {len(agent_bulk_selection)} agents')
                        st.rerun()
        
        with bulk_tab4:
            st.write('**Quick Bulk Actions**')
            
            # Quick action buttons
            col1, col2 = st.columns(2)
            
            with col1:
                st.write('**Quick Archive Actions:**')
                
                if st.button('🗄️ Archive All Old Leads (60+ days)'):
                    with get_session() as db:
                        cutoff_date = datetime.utcnow() - pd.Timedelta(days=60)
                        q = db.query(Lead).filter(
                            sa.or_(Lead.is_archived.is_(None), Lead.is_archived == 'no'),
                            Lead.uploaded_at <= cutoff_date
                        )
                        leads_to_archive = q.all()
                        
                        if leads_to_archive:
                            lead_ids = [lead.id for lead in leads_to_archive]
                            archived_count = bulk_archive_leads(
                                db, lead_ids, current_user.username, 
                                'Quick action - Old leads (60+ days)', 
                                datetime.utcnow()
                            )
                            st.success(f'✅ Archived {archived_count} old leads')
                            st.rerun()
                        else:
                            st.info('No old leads found')
                
                if st.button('🗄️ Archive All "Lost" Status'):
                    with get_session() as db:
                        q = db.query(Lead).filter(
                            sa.or_(Lead.is_archived.is_(None), Lead.is_archived == 'no'),
                            Lead.status == 'lost'
                        )
                        leads_to_archive = q.all()
                        
                        if leads_to_archive:
                            lead_ids = [lead.id for lead in leads_to_archive]
                            archived_count = bulk_archive_leads(
                                db, lead_ids, current_user.username, 
                                'Quick action - All lost leads', 
                                datetime.utcnow()
                            )
                            st.success(f'✅ Archived {archived_count} lost leads')
                            st.rerun()
                        else:
                            st.info('No lost leads found')
            
            with col2:
                st.write('**Quick Archive by Status:**')
                
                quick_status = st.selectbox('Select Status to Archive', ['new', 'contacted', 'qualified', 'won', 'lost'])
                
                if st.button(f'🗄️ Archive All "{quick_status}" Status'):
                    with get_session() as db:
                        q = db.query(Lead).filter(
                            sa.or_(Lead.is_archived.is_(None), Lead.is_archived == 'no'),
                            Lead.status == quick_status
                        )
                        leads_to_archive = q.all()
                        
                        if leads_to_archive:
                            lead_ids = [lead.id for lead in leads_to_archive]
                            archived_count = bulk_archive_leads(
                                db, lead_ids, current_user.username, 
                                f'Quick action - All {quick_status} leads', 
                                datetime.utcnow()
                            )
                            st.success(f'✅ Archived {archived_count} {quick_status} leads')
                            st.rerun()
                        else:
                            st.info(f'No {quick_status} leads found')
                
                # Archive by age ranges
                st.write('**Quick Archive by Age:**')
                age_ranges = [
                    ('30+ days', 30),
                    ('60+ days', 60),
                    ('90+ days', 90),
                    ('180+ days', 180)
                ]
                
                for age_label, age_days in age_ranges:
                    if st.button(f'🗄️ Archive {age_label}'):
                        with get_session() as db:
                            cutoff_date = datetime.utcnow() - pd.Timedelta(days=age_days)
                            q = db.query(Lead).filter(
                                sa.or_(Lead.is_archived.is_(None), Lead.is_archived == 'no'),
                                Lead.uploaded_at <= cutoff_date
                            )
                            leads_to_archive = q.all()
                            
                            if leads_to_archive:
                                lead_ids = [lead.id for lead in leads_to_archive]
                                archived_count = bulk_archive_leads(
                                    db, lead_ids, current_user.username, 
                                    f'Quick action - {age_label} old leads', 
                                    datetime.utcnow()
                                )
                                st.success(f'✅ Archived {archived_count} leads ({age_label})')
                                st.rerun()
                            else:
                                st.info(f'No leads {age_label} old found')
            
            # Quick Delete Actions
            st.write('**🗑️ Quick Delete Actions**')
            st.warning('⚠️ **WARNING**: These actions will permanently delete leads from the database!')
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.write('**Delete by Status:**')
                delete_status = st.selectbox('Select status to delete', ['new', 'contacted', 'qualified', 'won', 'lost'], key='delete_status')
                
                if st.button(f'🗑️ Delete All "{delete_status}" Status', type='secondary', key='delete_by_status'):
                    delete_status_reason = st.text_area(f'Reason for deleting all {delete_status} leads', 
                                                      placeholder='Explain why these leads should be deleted', 
                                                      key='delete_status_reason')
                    
                    if st.button(f'🗑️ Confirm Delete All {delete_status}', type='primary', key='confirm_delete_status'):
                        if delete_status_reason.strip():
                            with get_session() as db:
                                q = db.query(Lead).filter(
                                    sa.or_(Lead.is_archived.is_(None), Lead.is_archived == 'no'),
                                    Lead.status == delete_status
                                )
                                leads_to_delete = q.all()
                                
                                if leads_to_delete:
                                    lead_ids = [lead.id for lead in leads_to_delete]
                                    deleted_count = bulk_delete_leads_from_db(db, lead_ids, current_user.username, delete_status_reason.strip())
                                    st.success(f'✅ Successfully deleted {deleted_count} {delete_status} leads')
                                    st.rerun()
                                else:
                                    st.info(f'No {delete_status} leads found')
                        else:
                            st.error('❌ Please provide a deletion reason')
            
            with col2:
                st.write('**Delete by Age:**')
                delete_age_days = st.selectbox('Delete leads older than', [30, 60, 90, 180, 365], key='delete_age_days')
                
                if st.button(f'🗑️ Delete Leads Older Than {delete_age_days} Days', type='secondary', key='delete_by_age'):
                    cutoff_date = datetime.utcnow() - pd.Timedelta(days=delete_age_days)
                    
                    with get_session() as db:
                        old_leads = db.query(Lead).filter(
                            sa.or_(Lead.is_archived.is_(None), Lead.is_archived == 'no'),
                            Lead.uploaded_at <= cutoff_date
                        ).all()
                        
                        if old_leads:
                            lead_ids = [lead.id for lead in old_leads]
                            deleted_count = bulk_delete_leads_from_db(db, lead_ids, current_user.username, f'Automatic cleanup - leads older than {delete_age_days} days')
                            st.success(f'✅ Successfully deleted {deleted_count} old leads')
                            st.rerun()
                        else:
                            st.info(f'No leads older than {delete_age_days} days found')
    
    with archive_tab4:
        st.write('**Archive Leads by Date**')
        
        # Date-based archiving
        col1, col2 = st.columns(2)
        with col1:
            archive_start_date = st.date_input('Start Date', value=date.today() - pd.Timedelta(days=30), key='archive_start_date_tab4')
            archive_end_date = st.date_input('End Date', value=date.today(), key='archive_end_date_tab4')
        with col2:
            date_archive_reason = st.selectbox('Date Archive Reason', [
                'Date-based cleanup',
                'Old leads cleanup',
                'System maintenance',
                'Other'
            ], key='date_archive_reason_tab4')
        
        if st.button('🗄️ Archive Leads by Date Range'):
            with get_session() as db:
                start_datetime = datetime.combine(archive_start_date, datetime.min.time())
                end_datetime = datetime.combine(archive_end_date, datetime.max.time())
                
                # Get leads in date range
                q = db.query(Lead).filter(
                    sa.or_(Lead.is_archived.is_(None), Lead.is_archived == 'no'),
                    Lead.uploaded_at >= start_datetime,
                    Lead.uploaded_at <= end_datetime
                )
                
                leads_in_range = q.all()
                
                if leads_in_range:
                    lead_ids = [lead.id for lead in leads_in_range]
                    archived_count = bulk_archive_leads(
                        db, lead_ids, current_user.username, 
                        f"{date_archive_reason} ({archive_start_date} to {archive_end_date})", 
                        datetime.utcnow()
                    )
                    st.success(f'✅ Archived {archived_count} leads from {archive_start_date} to {archive_end_date}')
                    st.rerun()
                else:
                    st.info(f'No leads found in date range {archive_start_date} to {archive_end_date}')
        
        # View archived leads by date
        st.write('**View Archived Leads by Date**')
        view_date = st.date_input('View archived on date', value=date.today(), key='view_date_tab4')
        
        if st.button('📅 View Archived by Date'):
            with get_session() as db:
                archived_on_date = get_archived_leads_by_date(db, view_date.strftime('%Y-%m-%d'))
                
                if archived_on_date:
                    archived_df = pd.DataFrame([{
                        'id': lead.id,
                        'name': lead.name,
                        'sales_agent': lead.sales_agent,
                        'status': lead.status,
                        'archived_by': lead.archived_by,
                        'archived_at': lead.archived_at.strftime('%Y-%m-%d %H:%M') if lead.archived_at else '',
                        'archive_reason': lead.archive_reason
                    } for lead in archived_on_date])
                    
                    st.write(f'**Leads archived on {view_date}:**')
                    st.dataframe(archived_df, use_container_width=True)
                else:
                    st.info(f'No leads were archived on {view_date}')
    
    # Archive Summary Dashboard
    st.markdown('---')
    st.subheader('📊 Archive Summary Dashboard')
    
    with get_session() as db:
        # Get archive statistics
        total_archived = db.query(Lead).filter(Lead.is_archived == 'yes').count()
        total_active = db.query(Lead).filter(sa.or_(Lead.is_archived.is_(None), Lead.is_archived == 'no')).count()
        
        # Archive by reason
        archive_reasons = db.query(Lead.archive_reason, func.count(Lead.id)).filter(
            Lead.is_archived == 'yes'
        ).group_by(Lead.archive_reason).all()
        
        # Archive by date (last 30 days)
        thirty_days_ago = datetime.utcnow() - pd.Timedelta(days=30)
        recent_archives = db.query(func.date(Lead.archived_at), func.count(Lead.id)).filter(
            Lead.is_archived == 'yes',
            Lead.archived_at >= thirty_days_ago
        ).group_by(func.date(Lead.archived_at)).order_by(func.date(Lead.archived_at)).all()
        
        # Archive by agent
        archive_by_agent = db.query(Lead.archived_by, func.count(Lead.id)).filter(
            Lead.is_archived == 'yes'
        ).group_by(Lead.archived_by).all()
    
    # Display summary metrics
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric('Total Active Leads', total_active)
    with col2:
        st.metric('Total Archived Leads', total_archived)
    with col3:
        archive_rate = (total_archived / (total_active + total_archived) * 100) if (total_active + total_archived) > 0 else 0
        st.metric('Archive Rate', f'{archive_rate:.1f}%')
    with col4:
        recent_count = sum(count for _, count in recent_archives)
        st.metric('Archived (30 days)', recent_count)
    
    # Charts
    col1, col2 = st.columns(2)
    
    with col1:
        if archive_reasons:
            reasons_df = pd.DataFrame(archive_reasons, columns=['Reason', 'Count'])
            fig_reasons = px.pie(reasons_df, values='Count', names='Reason', title='Archives by Reason')
            st.plotly_chart(fig_reasons, use_container_width=True)
        else:
            st.info('No archive reasons data available')
    
    with col2:
        if archive_by_agent:
            agents_df = pd.DataFrame(archive_by_agent, columns=['Agent', 'Count'])
            fig_agents = px.bar(agents_df, x='Agent', y='Count', title='Archives by Agent')
            st.plotly_chart(fig_agents, use_container_width=True)
        else:
            st.info('No archive by agent data available')
    
    # Recent archive trend
    if recent_archives:
        recent_df = pd.DataFrame(recent_archives, columns=['Date', 'Count'])
        recent_df['Date'] = pd.to_datetime(recent_df['Date'])
        fig_trend = px.line(recent_df, x='Date', y='Count', title='Archive Trend (Last 30 Days)')
        st.plotly_chart(fig_trend, use_container_width=True)
    else:
        st.info('No recent archive data available')

elif role == 'ceo':
    st.header('CEO — Executive Dashboard & Reports')
    df_all, total = read_leads_df(limit=10000, offset=0)
    if df_all.empty:
        st.info('No data yet')
    else:
        st.subheader('Executive KPIs')
        total_leads = total
        unique_contacts = int(df_all['contact'].nunique())
        vc_all = df_all['sales_agent'].value_counts()
        top_agent = (vc_all.idxmax() if not vc_all.empty else '—')
        c1, c2, c3 = st.columns(3)
        c1.metric('Total Leads', total_leads)
        c2.metric('Unique Contacts', unique_contacts)
        c3.metric('Top Agent', top_agent)

        # CTO Analytics Section for CEO
        st.markdown('---')
        st.subheader('📊 CTO Analytics Dashboard')
        
        # Create all the charts that CTO sees
        df_f = df_all.copy()
        charts_data = {}
        
        try:
            # 1. Time series chart
            if not df_f.empty:
                df_f['uploaded_at'] = pd.to_datetime(df_f['uploaded_at'])
                daily = df_f.groupby(df_f['uploaded_at'].dt.date).size().reset_index(name='count')
                daily.columns = ['date', 'count']
                if not daily.empty:
                    fig_time = px.line(daily, x='date', y='count', title='Leads per day')
                    st.plotly_chart(fig_time, use_container_width=True)
                    charts_data['daily_leads'] = daily
            
            # 2. Agent breakdown
            agent_counts = df_f['sales_agent'].value_counts().reset_index()
            agent_counts.columns = ['agent', 'leads']
            if not agent_counts.empty:
                fig_agents = px.bar(agent_counts, x='agent', y='leads', title='Leads by agent')
                st.plotly_chart(fig_agents, use_container_width=True)
                charts_data['agent_breakdown'] = agent_counts
            
            # 3. Status breakdown
            status_counts = df_f['status'].value_counts().reset_index()
            status_counts.columns = ['status', 'count']
            if not status_counts.empty:
                fig_status = px.pie(status_counts, values='count', names='status', title='Lead status distribution')
                st.plotly_chart(fig_status, use_container_width=True)
                charts_data['status_breakdown'] = status_counts
            
            # 4. Sales Pipeline Funnel
            st.subheader('Sales Pipeline — Funnel Analysis')
            ordered_status = ['new','contacted','qualified','won','lost']
            counts_map = {s: int((df_f['status'] == s).sum()) for s in ordered_status}
            funnel_df = pd.DataFrame({'status': list(counts_map.keys()), 'count': list(counts_map.values())})
            fig_funnel = px.funnel(funnel_df, x='count', y='status', title='Lead Conversion Funnel')
            st.plotly_chart(fig_funnel, use_container_width=True)
            charts_data['sales_funnel'] = funnel_df
            
            # 5. Contact method analysis
            st.subheader('Contact Methods by Agent')
            cm = df_f.dropna(subset=['sales_agent','contact'])
            if not cm.empty:
                cm_counts = cm.groupby(['sales_agent','contact']).size().reset_index(name='count')
                fig_cm = px.bar(cm_counts, x='sales_agent', y='count', color='contact', barmode='stack', title='Contact methods by agent')
                st.plotly_chart(fig_cm, use_container_width=True)
                charts_data['contact_methods'] = cm_counts
            
            # 6. Heatmap analysis
            st.subheader('Lead Activity Heatmap')
            if not df_f.empty:
                tmp = df_f.copy()
                tmp['weekday'] = tmp['uploaded_at'].dt.day_name()
                tmp['hour'] = tmp['uploaded_at'].dt.hour
                heat = tmp.groupby(['weekday','hour']).size().reset_index(name='count')
                if not heat.empty:
                    fig_heat = px.density_heatmap(heat, x='hour', y='weekday', z='count', histfunc='avg', title='Lead uploads by time')
                    st.plotly_chart(fig_heat, use_container_width=True)
                    charts_data['activity_heatmap'] = heat
            
            # 7. Rolling average
            st.subheader('Trends Analysis')
            if not daily.empty:
                daily_ra = daily.sort_values('date').copy()
                daily_ra['rolling_7d'] = daily_ra['count'].rolling(window=7, min_periods=1).mean()
                fig_ra = px.line(daily_ra, x='date', y=['count','rolling_7d'], labels={'value':'leads','variable':'series'}, title='Daily leads with 7-day trend')
                st.plotly_chart(fig_ra, use_container_width=True)
                charts_data['trends'] = daily_ra
            
        except Exception as e:
            st.error(f'Error generating charts: {str(e)}')
        
        # Download all charts and data as ZIP
        st.markdown('---')
        st.subheader('📦 Download Complete Analytics Package')
        
        from datetime import datetime
        import zipfile
        
        # Generate filename with current day and date
        now = datetime.now()
        day_name = now.strftime('%A')  # Full day name (e.g., 'Monday')
        date_str = now.strftime('%Y-%m-%d')  # Date format (e.g., '2024-08-09')
        zip_filename = f"IQ_Stats_CRM_Analytics_{day_name}_{date_str}.zip"
        
        if st.button('📊 Generate & Download Analytics Package'):
            try:
                # Create ZIP file in memory
                zip_buffer = io.BytesIO()
                
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    # Add all chart data as Excel sheets
                    excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                        # Main data
                        df_all.to_excel(writer, sheet_name='All_Leads', index=False)
                        
                        # Chart data
                        for chart_name, chart_df in charts_data.items():
                            if isinstance(chart_df, pd.DataFrame) and not chart_df.empty:
                                chart_df.to_excel(writer, sheet_name=chart_name.replace('_', ' ').title()[:31], index=False)
                        
                        # Summary statistics
                        summary_data = {
                            'Metric': ['Total Leads', 'Unique Contacts', 'Top Agent', 'Active Salesmen', 'Conversion Rate'],
                            'Value': [
                                total_leads,
                                unique_contacts,
                                top_agent,
                                len(agent_counts) if 'agent_breakdown' in charts_data else 0,
                                f"{(df_f['status'] == 'won').sum() / len(df_f) * 100:.1f}%" if not df_f.empty else "0%"
                            ]
                        }
                        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Executive_Summary', index=False)
                    
                    # Add Excel file to ZIP
                    zip_file.writestr(f'CRM_Analytics_{date_str}.xlsx', excel_buffer.getvalue())
                    
                    # Generate and add PDF graphs
                    pdf_graphs = generate_analytics_graphs(df_f, charts_data, date_str, "CEO Analytics")
                    zip_file.writestr(f'CEO_Graphs_{date_str}.pdf', pdf_graphs)
                    
                    # Generate and add interactive HTML graphs
                    html_graphs = generate_plotly_graphs(df_f, charts_data, date_str, "CEO Analytics")
                    for filename, html_content in html_graphs.items():
                        zip_file.writestr(f'graphs/{filename}', html_content)
                    
                    # Add deals data if available
                    deals_df, _ = read_deals_df(limit=100000)
                    if not deals_df.empty:
                        deals_buffer = io.BytesIO()
                        with pd.ExcelWriter(deals_buffer, engine='openpyxl') as deals_writer:
                            deals_df.to_excel(deals_writer, sheet_name='All_Deals', index=False)
                            
                            # Deals summary
                            deals_summary = deals_df.groupby('uploaded_by').size().reset_index(name='deals_count')
                            deals_summary.to_excel(deals_writer, sheet_name='Deals_by_Agent', index=False)
                        
                        zip_file.writestr(f'Deals_Report_{date_str}.xlsx', deals_buffer.getvalue())
                    
                    # Add a README file
                    readme_content = f"""IQ Stats CRM Analytics Package
Generated on: {now.strftime('%A, %B %d, %Y at %H:%M')}
Generated for: CEO Dashboard

Contents:
- CRM_Analytics_{date_str}.xlsx: Complete leads analysis with all charts data
- CEO_Graphs_{date_str}.pdf: Static PDF graphs and charts
- graphs/ folder: Interactive HTML graphs (open in browser)
  * daily_leads_trend.html: Daily leads trend analysis
  * agent_performance.html: Agent performance breakdown
  * status_distribution.html: Lead status distribution
  * sales_funnel.html: Sales funnel visualization
  * contact_methods.html: Contact methods analysis
  * activity_heatmap.html: Activity heatmap
  * interactive_dashboard.html: Combined interactive dashboard
- Deals_Report_{date_str}.xlsx: Deals tracking and performance data

Chart Data Included:
- Daily leads trends
- Agent performance breakdown
- Lead status distribution
- Sales funnel analysis
- Contact method analysis
- Activity heatmaps
- Rolling averages and trends

Graphs Available:
- Static PDF graphs for printing and sharing
- Interactive HTML graphs for detailed analysis
- Combined dashboard view for executive overview

Executive Summary:
- Total Leads: {total_leads}
- Unique Contacts: {unique_contacts}
- Top Performing Agent: {top_agent}

This package contains comprehensive analytics for strategic decision making.
"""
                    zip_file.writestr('README.txt', readme_content)
                
                # Offer download
                st.download_button(
                    label=f'📥 Download {zip_filename}',
                    data=zip_buffer.getvalue(),
                    file_name=zip_filename,
                    mime='application/zip',
                    help=f'Download complete analytics package for {day_name}, {date_str}'
                )
                
                st.success(f'✅ Analytics package ready! Contains all CTO dashboard charts and executive reports for {day_name}, {date_str}')
                
            except Exception as e:
                st.error(f'Error creating analytics package: {str(e)}')

        st.markdown('---')
        st.subheader('Download reports')
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            df_all.to_excel(writer, index=False, sheet_name='leads')
        st.download_button('Download Excel report', data=buf.getvalue(), file_name='crm_report.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        st.subheader('Summary by agent')
        summary = df_all.groupby('sales_agent').agg(leads=('id','count'), unique_contacts=('contact','nunique')).reset_index()
        st.dataframe(summary)

        # Done Deals Reports
        st.markdown('---')
        st.subheader('Done Deals — Reports & Export')
        deals_df, deals_total = read_deals_df(limit=100000, offset=0)
        if deals_df.empty:
            st.info('No deals submitted yet')
        else:
            deals_df['created_at'] = pd.to_datetime(deals_df['created_at'])
            c1, c2, c3 = st.columns(3)
            c1.metric('Total Deals', deals_total)
            c2.metric('Unique Customers', int(deals_df['customer_name'].nunique()))
            c3.metric('Salesmen Submitting', int(deals_df['uploaded_by'].nunique()))

            # Summaries
            by_salesman = deals_df.groupby('uploaded_by').size().reset_index(name='deals')
            by_day = deals_df.groupby(deals_df['created_at'].dt.date).size().reset_index(name='deals')
            st.write('Deals by Salesman')
            st.dataframe(by_salesman)
            st.write('Deals per Day')
            st.dataframe(by_day)

            # Export comprehensive Excel for advanced analysis
            deals_buf = io.BytesIO()
            with pd.ExcelWriter(deals_buf, engine='openpyxl') as writer:
                deals_df.to_excel(writer, index=False, sheet_name='deals')
                by_salesman.to_excel(writer, index=False, sheet_name='by_salesman')
                by_day.to_excel(writer, index=False, sheet_name='by_day')
            st.download_button('Download Done Deals Excel', data=deals_buf.getvalue(), file_name='done_deals_report.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            # Optional: Excel with embedded images for CEO (may be large)
            try:
                db = get_session()
                all_deals = db.query(Deal).order_by(Deal.created_at.desc()).all()
                excel_with_images = build_deals_excel_with_images(all_deals)
                if excel_with_images:
                    st.download_button('Download Done Deals (Excel with images)', data=excel_with_images, file_name='done_deals_with_images.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                zip_all = build_deals_images_zip(all_deals)
                st.download_button('Download all deal screenshots (ZIP)', data=zip_all, file_name='done_deals_screenshots.zip', mime='application/zip')
            finally:
                try:
                    db.close()
                except Exception:
                    pass

        # Login activity (visibility for CEO)
        st.markdown('---')
        st.subheader('User Login Activity (last 30 days)')
        try:
            with get_session() as db:
                thirty_days_ago = datetime.utcnow() - pd.Timedelta(days=30)
                logins = db.query(LoginEvent).filter(LoginEvent.logged_in_at >= thirty_days_ago).order_by(LoginEvent.logged_in_at.desc()).all()
                if logins:
                    df_login = pd.DataFrame([
                        {
                            'username': e.username,
                            'role': e.role,
                            'logged_in_at': e.logged_in_at
                        } for e in logins
                    ])
                    st.dataframe(df_login)
                    agg = df_login.groupby(['role','username']).size().reset_index(name='logins')
                    fig_login = px.bar(agg, x='username', y='logins', color='role', title='Logins per user (30 days)')
                    st.plotly_chart(fig_login, use_container_width=True)
                else:
                    st.info('No login events yet.')
        except Exception:
            st.info('Login activity not available yet.')

else:
    st.write('Role not supported in UI')

# Recent Activity Section - Only for CEO and CTO
if role in ['ceo', 'cto']:
    st.sidebar.markdown('---')
    st.sidebar.markdown('### 📊 Recent Activity')
    if st.sidebar.checkbox('Show recent activity', value=True):
        db = get_session()
        try:
            # Get recent activities
            acts = db.query(Activity).order_by(Activity.timestamp.desc()).limit(20).all()
            if acts:
                st.sidebar.markdown('**Latest System Activities:**')
                for i, a in enumerate(acts, 1):
                    time_str = a.timestamp.strftime('%m/%d %H:%M')
                    st.sidebar.markdown(f"**{i}.** `{time_str}` - **{a.actor}** {a.action} Lead #{a.lead_id}")
            else:
                st.sidebar.info('No recent activities found.')
            
            # Get recent login events
            st.sidebar.markdown('---')
            st.sidebar.markdown('**Recent Logins:**')
            recent_logins = db.query(LoginEvent).order_by(LoginEvent.logged_in_at.desc()).limit(10).all()
            if recent_logins:
                for i, login in enumerate(recent_logins, 1):
                    time_str = login.logged_in_at.strftime('%m/%d %H:%M')
                    st.sidebar.markdown(f"**{i}.** `{time_str}` - **{login.username}** ({login.role})")
            else:
                st.sidebar.info('No recent logins found.')
                
        except Exception as e:
            st.sidebar.error(f'Error loading activities: {str(e)}')
        finally:
            db.close()


